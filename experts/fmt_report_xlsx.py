$extens("include.py")
include("import openpyxl", ["extella-pip install openpyxl"])

def fmt_report_xlsx(input_path: str = "", records_json: str = "", spec_json: str = "",
                    output_path: str = "", brand_name: str = "", accent: str = "",
                    footer_note: str = "", sections_json: str = "") -> dict:
    """Оформитель отчётов в Excel. Тот же контракт, что у PDF/Word/презентации.

    Excel — формат для тех, кто будет СЧИТАТЬ: фильтровать, сводить, добавлять свои формулы.
    Поэтому здесь не картинка отчёта, а рабочая книга:
      «Сводка»  — главное число, разрезы таблицами + РОДНАЯ диаграмма (правится в Excel);
      «Данные»  — все записи с автофильтром и закреплённой шапкой, чтобы работать сразу.

    Отличие от нынешней голой выгрузки: шапки, ширины, числовые форматы и диаграмма —
    файл открывается готовым к работе, а не требует получаса ручной доводки.

    Документ принадлежит КЛИЕНТУ: знаков Extella нет, в шапке имя владельца процесса.
    """
    import json
    import os
    from pathlib import Path
    from datetime import datetime, timezone
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    ready = []
    if sections_json and not sections_json.startswith("{{"):
        try:
            ready = [s for s in (json.loads(sections_json) or []) if isinstance(s, dict) and s.get("items")]
        except Exception:
            ready = []

    recs = []
    if records_json and not records_json.startswith("{{"):
        try:
            recs = json.loads(records_json)
        except Exception:
            return {"status": "error", "message": "records_json не разобрался"}
    elif input_path and not input_path.startswith("{{") and Path(input_path).exists():
        try:
            data = json.loads(Path(input_path).read_text(encoding="utf-8"))
        except Exception:
            return {"status": "error", "message": "вход не читается как JSON: " + str(input_path)[:120]}
        recs = data if isinstance(data, list) else (data.get("records") or data.get("rows") or [])
    recs = [r for r in (recs or []) if isinstance(r, dict)]
    if not ready and not recs:
        return {"status": "error", "message": "нет данных для отчёта (ни records_json, ни input_path, ни sections_json)"}

    spec = {}
    if spec_json and not spec_json.startswith("{{"):
        try:
            spec = json.loads(spec_json) or {}
        except Exception:
            spec = {}
    style = spec.get("style") or {}
    accent = (accent or style.get("accent") or "#2F6B66").lstrip("#")
    if len(accent) != 6:
        accent = "2F6B66"
    brand_name = brand_name or style.get("brand_name") or ""
    footer_note = footer_note or style.get("footer") or ""
    cols = list(recs[0].keys()) if recs else []

    sections = list(ready)
    if not sections:
        views = [v for v in (spec.get("views") or []) if isinstance(v, dict) and v.get("group_by") in cols]
        if not views:
            cand = []
            for c in cols:
                vals = [str(r.get(c, "")).strip() for r in recs if str(r.get(c, "")).strip()]
                uniq = len(set(vals))
                if vals and 1 < uniq <= max(2, min(12, len(recs))) and uniq < len(vals):
                    cand.append((uniq, c))
            views = [{"group_by": c, "title": "По полю «" + c + "»"} for _, c in sorted(cand)[:3]]
        for v in views[:4]:
            g = v["group_by"]
            agg = {}
            for r in recs:
                k = str(r.get(g, "")).strip()
                agg[k or "не заполнено"] = agg.get(k or "не заполнено", 0) + 1
            named = {k: n for k, n in agg.items() if k != "не заполнено"}
            items = dict(sorted(named.items(), key=lambda x: -x[1])[:12])
            if agg.get("не заполнено"):
                items["не заполнено"] = agg["не заполнено"]
            sections.append({"title": v.get("title") or g, "items": items})

    hl = spec.get("headline") or {}
    if ready and not recs:
        _t = spec.get("total")
        hv = _t if _t is not None else sum(sum(s["items"].values()) for s in ready)
    else:
        hv = len(recs)
    hlabel = str(hl.get("label") or "записей в отчёте").replace("\n", " ")

    ACC = PatternFill("solid", fgColor=accent)
    THIN = Border(bottom=Side(style="thin", color="EBE8E1"))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False

    ws["A1"] = brand_name or ""
    ws["A1"].font = Font(bold=True, size=12)
    ws["D1"] = datetime.now(timezone.utc).strftime("%d.%m.%Y")
    ws["D1"].font = Font(size=9, color="8C8C8C")
    ws["D1"].alignment = Alignment(horizontal="right")
    ws["A3"] = str(spec.get("report") or "Отчёт")
    ws["A3"].font = Font(bold=True, size=16)
    if spec.get("subtitle"):
        ws["A4"] = str(spec["subtitle"])
        ws["A4"].font = Font(size=10, color="2A2A2A")

    ws["A6"] = hv
    ws["A6"].font = Font(bold=True, size=24, color=accent)
    ws["B6"] = hlabel
    ws["B6"].font = Font(size=11)
    ws["B6"].alignment = Alignment(vertical="center")

    row = 8
    first_chart_ref = None
    for sec in sections:
        ws.cell(row=row, column=1, value=str(sec["title"]).upper()).font = Font(bold=True, size=9, color="8C8C8C")
        row += 1
        h1 = ws.cell(row=row, column=1, value="Значение")
        h2 = ws.cell(row=row, column=2, value="Количество")
        for c in (h1, h2):
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = ACC
        row += 1
        start = row
        for k, v in sec["items"].items():
            ws.cell(row=row, column=1, value=str(k)[:60]).border = THIN
            c = ws.cell(row=row, column=2, value=v)
            c.border = THIN
            c.alignment = Alignment(horizontal="right")
            row += 1
        # РОДНАЯ диаграмма для первого разреза: её можно править и пересчитывать в Excel
        if first_chart_ref is None and row > start:
            ch = BarChart()
            ch.type = "bar"
            ch.title = str(sec["title"])
            ch.legend = None
            ch.height, ch.width = 7, 12
            data = Reference(ws, min_col=2, min_row=start, max_row=row - 1)
            cats = Reference(ws, min_col=1, min_row=start, max_row=row - 1)
            ch.add_data(data, titles_from_data=False)
            ch.set_categories(cats)
            ws.add_chart(ch, "E8")
            first_chart_ref = True
        row += 1

    if footer_note:
        ws.cell(row=row + 1, column=1, value=footer_note).font = Font(size=8, color="8C8C8C")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 16

    # ── лист «Данные»: с ним сразу работают — автофильтр + закреплённая шапка ──
    if recs:
        ws2 = wb.create_sheet("Данные")
        ws2.append(cols)
        for i, c in enumerate(cols, start=1):
            cell = ws2.cell(row=1, column=i)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = ACC
            ws2.column_dimensions[get_column_letter(i)].width = min(38, max(12, len(str(c)) + 6))
        for r in recs:
            ws2.append([r.get(c, "") for c in cols])
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = "A1:" + get_column_letter(len(cols)) + str(len(recs) + 1)

    out = output_path or ("/tmp/report_" + datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S") + ".xlsx")
    try:
        wb.save(out)
    except Exception as e:
        return {"status": "error", "message": "XLSX не сохранился: " + str(e)[:140]}
    return {"status": "success", "path": out, "bytes": os.path.getsize(out),
            "records": len(recs), "preview": bool(ready and not recs),
            "sheets": wb.sheetnames, "sections": [s["title"] for s in sections]}
