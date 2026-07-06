# expert: wz_build_runner
# description: Эксперт wz_build_runner (Adoption Wizard).
# params: session_id, build_plan_path, task_id, api_token, api_key, base_url, model, max_attempts, api_base

$extens("include.py")
include("import requests", ["extella-pip install requests"])
include("import openpyxl", ["extella-pip install openpyxl"])

def wz_build_runner(
    session_id: str = "",
    build_plan_path: str = "",
    task_id: str = "",
    api_token: str = "",
    api_key: str = "",
    base_url: str = "https://api.openai.com/v1",
    model: str = "gpt-4o",
    max_attempts: int = 3,
    api_base: str = "https://api.extella.ai"
) -> dict:
    import json, re, time, ast, tempfile
    import requests
    from pathlib import Path
    from datetime import datetime, timezone

    def now():
        return datetime.now(timezone.utc).isoformat()

    # ── Resolve inputs ────────────────────────────────────────────────
    if not api_token:
        return {"status": "error", "message": "api_token is required"}
    if not api_key:
        return {"status": "error", "message": "api_key is required"}
    if not build_plan_path:
        if not session_id:
            return {"status": "error", "message": "session_id or build_plan_path is required"}
        build_plan_path = str(Path.home() / "extella_wizard" / "sessions" / (session_id + "_build_plan.json"))
    bpp = Path(build_plan_path)
    if not bpp.exists():
        return {"status": "error", "message": "build plan not found: " + str(bpp)}
    doc = json.loads(bpp.read_text(encoding="utf-8"))
    plan = doc.get("plan", doc)
    tasks = plan.get("tasks") or []
    if not tasks:
        return {"status": "error", "message": "build plan has no tasks"}

    manifest_path = bpp.parent / (bpp.stem.replace("_build_plan", "") + "_build_manifest.json")
    manifest = {"tasks": {}, "updated_at": now()}
    if manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            pass

    def save_manifest():
        manifest["updated_at"] = now()
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    def t_state(tid):
        return manifest["tasks"].get(tid, {}).get("status", "pending")

    # ── Pick the task: explicit id, or next runnable ──────────────────
    task = None
    if task_id:
        task = next((t for t in tasks if t.get("id") == task_id), None)
        if not task:
            return {"status": "error", "message": "task not found in plan: " + task_id}
    else:
        for t in tasks:
            if t_state(t["id"]) in ("done", "built_unverified"):
                continue
            deps_ok = all(t_state(d) in ("done", "built_unverified") for d in (t.get("depends_on") or []))
            if deps_ok:
                task = t
                break
        if not task:
            done = [tid for tid in manifest["tasks"] if t_state(tid) in ("done", "built_unverified")]
            return {"status": "success", "message": "no runnable tasks left",
                    "done_tasks": done, "total_tasks": len(tasks),
                    "all_done": len(done) >= len(tasks)}

    tid = task["id"]
    rec = manifest["tasks"].setdefault(tid, {"status": "pending", "attempts": []})
    expert_name = task.get("expert_name", "")
    acceptance = task.get("acceptance_test") or {}

    # ── REST helpers ──────────────────────────────────────────────────
    headers = {"X-Auth-Token": api_token, "Content-Type": "application/json",
               "X-Profile-Id": "default", "X-Agent-Id": "agent_extella_default"}

    def xapi(ep, payload, timeout=600):
        # КАНОН СКОУПИНГА: построенные эксперты живут в ОБЩЕЙ библиотеке (global),
        # чтобы оркестратор процесса и агент вызывали их с global:true. Без этого
        # эксперт оседает в скоупе сборщика и снаружи «не находится».
        if ep.startswith("/api/expert/") and "global" not in payload:
            payload = dict(payload)
            payload["global"] = True
        try:
            r = requests.post(api_base.rstrip("/") + ep, headers=headers, json=payload, timeout=timeout)
            return r.json() if r.status_code == 200 else {"status": "error", "message": "HTTP " + str(r.status_code) + ": " + r.text[:200]}
        except Exception as e:
            return {"status": "error", "message": str(e)[:200]}

    def parse_run(res):
        raw = res.get("result", res) if isinstance(res, dict) else res
        if isinstance(raw, str):
            for loader in (json.loads, ast.literal_eval):
                try:
                    v = loader(raw)
                    if isinstance(v, dict):
                        return v
                except Exception:
                    continue
            return {"status": "unknown", "raw": raw[:400]}
        return raw if isinstance(raw, dict) else {"status": "unknown"}

    # ── REUSE path: verify the library expert exists ──────────────────
    if task.get("action") in ("reuse", "parameterize") and task.get("reuse_of"):
        g = xapi("/api/expert/get", {"name": task["reuse_of"]})
        if g.get("status") == "success" and g.get("expert_code"):
            rec.update({"status": "done", "mode": task["action"], "expert_name": task["reuse_of"],
                        "note": "library expert verified to exist; acceptance to be run during vertical slice",
                        "finished_at": now()})
            save_manifest()
            return {"status": "success", "task_id": tid, "result": "reused",
                    "expert_name": task["reuse_of"], "manifest_path": str(manifest_path)}
        rec.setdefault("attempts", []).append({"at": now(), "error": "reuse_of not found: " + str(task["reuse_of"])})
        task["action"] = "build"

    # ── ГРАУНДИНГ: реальная структура загруженного файла клиента ──────
    # компоненты строим под ФАКТИЧЕСКИЕ колонки образца, а не под выдуманную схему.
    def inspect_sample():
        sid = session_id or bpp.stem.replace("_build_plan", "")
        fdir = Path.home() / "extella_wizard" / "sessions" / (sid + "_files")
        if not fdir.is_dir():
            return ""
        files = [p for p in sorted(fdir.iterdir()) if p.is_file()]
        if not files:
            return ""
        f = files[0]
        ext = f.suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = []
                for r in ws.iter_rows(min_row=1, max_row=15):
                    rows.append([("" if c.value is None else str(c.value)) for c in r])
                hdr_idx = 0
                best = -1
                for i, r in enumerate(rows):
                    filled = sum(1 for v in r if v.strip())
                    strs = sum(1 for v in r if v.strip() and not v.replace(".", "").replace("-", "").isdigit())
                    score = filled + strs
                    if score > best:
                        best = score; hdr_idx = i
                cols = [v for v in rows[hdr_idx] if v.strip()]
                sample = rows[hdr_idx + 1] if hdr_idx + 1 < len(rows) else []
                return ("\n\nФАКТИЧЕСКАЯ СТРУКТУРА ЗАГРУЖЕННОГО ФАЙЛА (СТРОЙ СТРОГО ПОД ЭТИ КОЛОНКИ, "
                        "НЕ ВЫДУМЫВАЙ ПОЛЯ вроде id/due_date, если их тут нет): формат " + ext +
                        ", лист '" + str(ws.title) + "', строка заголовков #" + str(hdr_idx + 1) +
                        ", колонки: " + json.dumps(cols, ensure_ascii=False) +
                        ", пример строки: " + json.dumps(sample, ensure_ascii=False)[:400] +
                        ". Парсер должен вернуть записи с ЭТИМИ ключами; анализ/отчёт — работать с ними же.")
            if ext == ".csv":
                import csv as _csv
                with open(str(f), "r", encoding="utf-8", errors="replace") as fh:
                    rd = list(_csv.reader(fh))
                cols = [v for v in (rd[0] if rd else []) if v.strip()]
                sample = rd[1] if len(rd) > 1 else []
                return ("\n\nФАКТИЧЕСКАЯ СТРУКТУРА ЗАГРУЖЕННОГО ФАЙЛА (СТРОЙ СТРОГО ПОД ЭТИ КОЛОНКИ): формат csv"
                        ", колонки: " + json.dumps(cols, ensure_ascii=False) +
                        ", пример строки: " + json.dumps(sample, ensure_ascii=False)[:400] +
                        ". Парсер должен вернуть записи с ЭТИМИ ключами; анализ/отчёт — работать с ними же.")
        except Exception as e:
            return "\n\n(структуру файла-образца прочитать не удалось: " + str(e)[:120] + ")"
        return ""

    data_schema_hint = inspect_sample()

    # ── BUILD path: LLM codegen -> save -> verify -> acceptance ───────
    QUALITY = """Обязательный стандарт эксперта Extella:
1) первая строка кода: $extens("include.py"); зависимости: include("import X", ["extella-pip install X"]) — стандартная библиотека БЕЗ install-команды: include("import json", []);
2) сигнатура: явные типы и дефолты, БЕЗ *args/**kwargs; все импорты продублировать внутри функции;
3) валидация входов, ранний return {"status":"error","message":...};
4) возврат ВСЕГДА dict со status ("success"/"error");
5) НИКАКОГО хардкода путей/ключей/ID/колонок — всё параметрами; НЕ обращаться к KV Store из кода; не возвращать бинарные данные (только пути к файлам)."""

    stage_ctx = json.dumps({k: task.get(k) for k in ("purpose", "params_spec", "cspl")}, ensure_ascii=False)
    accept_ctx = json.dumps(acceptance, ensure_ascii=False)

    SYSTEM = """Ты — генератор кода экспертов платформы Extella. Верни ТОЛЬКО JSON:
{"name": "...", "description": "<английский: что делает + ВСЕ параметры с назначением>",
 "code": "<полный код эксперта>", "kwargs": {<параметры с дефолтами>}, "cspl": "fython",
 "fixture_code": "<python: def make_fixture(tmp_dir: str) -> dict — создаёт в tmp_dir синтетические входные файлы для приёмочного теста и возвращает словарь параметров эксперта, указывающих на них; если файлы не нужны — верни {}>"}
""" + QUALITY + """
6) КРИТИЧНО: ровно ОДНА функция верхнего уровня — она и есть точка входа (рантайм исполняет первый def); все вспомогательные функции определяй ВНУТРИ главной функции.
7) КРИТИЧНО ПРОТИВ ГЛЮКА ВЫГРУЗКИ: если эксперт производит данные (разобранные строки, список записей, таблицу, отчёт) — он ОБЯЗАН принимать параметр output_path (str) и ПИСАТЬ результат в файл по этому пути (json/csv/xlsx/docx по смыслу). Возвращать ТОЛЬКО компактную сводку-dict: {"status":"success","output_path":<путь>, <несколько ключевых чисел: сколько строк/записей/сумма>}. НИКОГДА не клади сырые строки/записи/полный датасет в возвращаемый dict — платформа давится крупным результатом ("Task completed but upload result failed"). Параметр output_path добавь в kwargs и в приёмочные example_params.
Правила fixture_code: только стандартная библиотека Python (json, csv, random, pathlib, openpyxl НЕЛЬЗЯ — для xlsx используй csv или json вход у эксперта, если спецификация позволяет; если эксперт обязан читать xlsx — fixture может писать CSV, а эксперт должен поддерживать оба формата параметром). Числа в фикстуре осмысленные (суммы, даты, просрочки). Никаких сетевых вызовов в fixture_code.
8) fixture_code ОБЯЗАН вернуть значения для ВСЕХ параметров сигнатуры эксперта (см. список имён из приёмочного теста). Если параметр — список записей (напр. request_data/subscription_data), верни его как список из 3-5 словарей с реалистичными полями (используй колонки из «ФАКТИЧЕСКОЙ СТРУКТУРЫ ФАЙЛА», если она дана). Ни один обязательный параметр не должен остаться незаполненным."""

    def gen(prompt_extra):
        req_keys = sorted((acceptance.get("example_params") or {}).keys())
        user = ("ЗАДАЧА СТРОЙКИ (JSON):\n" + stage_ctx +
                "\n\nИмя эксперта (СТРОГО): " + expert_name +
                "\nCSPL (СТРОГО): " + str(task.get("cspl", "fython")) +
                "\nСигнатура ОБЯЗАНА принимать параметры с ТОЧНО этими именами (из приёмочного теста): " + json.dumps(req_keys) +
                "\n\nПРИЁМОЧНЫЙ ТЕСТ (JSON):\n" + accept_ctx + data_schema_hint + prompt_extra +
                "\n\nСгенерируй эксперта и fixture_code.")
        r = requests.post(base_url.rstrip("/") + "/chat/completions",
                          headers={"Authorization": "Bearer " + api_key, "Content-Type": "application/json"},
                          json={"model": model,
                                "messages": [{"role": "system", "content": SYSTEM},
                                             {"role": "user", "content": user}],
                                "temperature": 0, "response_format": {"type": "json_object"},
                                "max_tokens": 4000},
                          timeout=180)
        if r.status_code != 200:
            raise RuntimeError("LLM " + str(r.status_code) + ": " + r.text[:150])
        return json.loads(r.json()["choices"][0]["message"]["content"])

    feedback = ""
    for attempt in range(1, max(1, int(max_attempts)) + 1):
        att = {"at": now(), "attempt": attempt}
        try:
            spec = gen(feedback)
            code = spec.get("code", "")
            if not code or "def " not in code:
                raise RuntimeError("LLM returned empty/invalid code")
            # HYGIENE: модель иногда переопределяет служебные функции include.py
            # (load_module/include) на верхнем уровне — вырезаем их блоки, они уже
            # предоставлены $extens("include.py") и ломают правило «один top-level def».
            code = re.sub(r"(?ms)^def\s+(?:load_module|include)\s*\(.*?(?=^\S|\Z)", "", code)
            # PLATFORM RULE: the runtime executes the FIRST top-level def in the code.
            # Enforce exactly ONE top-level function; helpers must be nested inside it.
            top_defs = re.findall(r"^def\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", code, flags=re.M)
            if len(top_defs) != 1:
                raise RuntimeError("code must contain exactly ONE top-level def (entry point); "
                                   "found " + str(top_defs) + ". Define helpers INSIDE the main function.")
            if top_defs[0] != expert_name:
                code = re.sub(r"^def\s+" + re.escape(top_defs[0]) + r"\s*\(",
                              "def " + expert_name + "(", code, count=1, flags=re.M)
            spec["name"] = expert_name
            # save expert
            sv = xapi("/api/expert/save", {"name": expert_name,
                                           "description": str(spec.get("description", task.get("purpose", "")))[:900],
                                           "code": code,
                                           "kwargs": spec.get("kwargs") or {},
                                           "cspl": task.get("cspl", "fython")})
            if sv.get("status") != "success":
                raise RuntimeError("save_expert failed: " + str(sv.get("message"))[:150])
            # independent existence check: the saved code must define the planned function
            g = xapi("/api/expert/get", {"name": expert_name})
            if g.get("status") != "success" or not g.get("expert_code"):
                raise RuntimeError("get_expert after save: expert missing")
            if ("def " + expert_name) not in g.get("expert_code", ""):
                raise RuntimeError("saved code does not define function " + expert_name)
            att["saved"] = True

            if task.get("cspl", "fython") != "fython":
                rec.update({"status": "built_unverified", "mode": "build", "expert_name": expert_name,
                            "note": "cspl=" + str(task.get("cspl")) + ": acceptance run deferred to vertical slice",
                            "finished_at": now()})
                rec.setdefault("attempts", []).append(att)
                save_manifest()
                return {"status": "success", "task_id": tid, "result": "built_unverified",
                        "expert_name": expert_name, "manifest_path": str(manifest_path)}

            # actual signature keys from the saved expert (independent source of truth)
            actual_keys = set((g.get("expert_params") or {}).keys())

            # fixture
            params = dict(acceptance.get("example_params") or {})
            fx = str(spec.get("fixture_code") or "")
            if fx.strip():
                tmp_dir = tempfile.mkdtemp(prefix="wzbuild_" + tid + "_")
                ns = {}
                exec(fx, ns)  # generated fixture, same trust domain as the expert itself
                if "make_fixture" in ns:
                    overrides = ns["make_fixture"](tmp_dir)
                    if isinstance(overrides, dict):
                        params.update(overrides)
                att["fixture_dir"] = tmp_dir

            # real uploaded sample files from the session take priority over the
            # synthetic fixture for file-input experts (xlsx/csv/doc can't be
            # reliably synthesized — urok stroyki podpisok 05.07).
            sid_for_files = session_id or bpp.stem.replace("_build_plan", "")
            files_dir = Path.home() / "extella_wizard" / "sessions" / (sid_for_files + "_files")
            samples = [p for p in sorted(files_dir.iterdir()) if p.is_file()] if files_dir.is_dir() else []
            if samples:
                by_ext = {}
                for p in samples:
                    by_ext.setdefault(p.suffix.lower(), p)
                for k, v in list(params.items()):
                    kl = k.lower()
                    if "output" in kl or kl.endswith("_dir") or "out_" in kl:
                        continue
                    vs = str(v).lower()
                    ext = next((e for e in (".xlsx", ".xls", ".csv", ".docx", ".pdf", ".json")
                                if vs.endswith(e)), None)
                    looks_file = ext is not None or any(t in kl for t in
                        ("file", "path", "xlsx", "excel", "csv", "source", "input", "doc"))
                    if looks_file:
                        pick = by_ext.get(ext) if ext else samples[0]
                        if not pick:
                            pick = samples[0]
                        params[k] = str(pick)
                        att.setdefault("sample_overrides", {})[k] = str(pick)

            # output_path: направляем результат в файл (эксперт должен писать туда,
            # а не возвращать крупный payload — иначе платформенный upload-глюк).
            out_file = None
            for k in actual_keys:
                kl = k.lower()
                if ("output" in kl or "out_" in kl) and ("path" in kl or "file" in kl or kl == "output_path"):
                    if not params.get(k):
                        base_tmp = att.get("fixture_dir") or tempfile.mkdtemp(prefix="wzacc_" + tid + "_")
                        out_file = str(Path(base_tmp) / ("acc_out" + (".xlsx" if "xlsx" in kl or "report" in kl else ".json")))
                        params[k] = out_file
                    else:
                        out_file = str(params[k])
                    break

            # drop params the saved signature does not accept (load_module rejects unknown kwargs)
            if actual_keys:
                dropped = [k for k in params if k not in actual_keys]
                if dropped:
                    att["params_dropped"] = dropped
                params = {k: v for k, v in params.items() if k in actual_keys}

            # acceptance run
            rr = xapi("/api/expert/run", {"expert_name": expert_name, "params": params})
            run_out = parse_run(rr)
            att["run_status"] = run_out.get("status")
            att["run_summary"] = {k: v for k, v in run_out.items() if not isinstance(v, (dict, list))}
            # платформенный глюк выгрузки: задача выполнилась, но результат не отдался —
            # считаем приёмку пройденной, если ожидаемый output-файл реально появился.
            upload_glitch = (run_out.get("status") != "success"
                             and "upload result failed" in str(run_out).lower())
            if upload_glitch and out_file and Path(out_file).exists() and Path(out_file).stat().st_size > 0:
                att["upload_glitch_recovered"] = out_file
                run_out = {"status": "success", "output_path": out_file,
                           "note": "result upload glitched; output file verified on disk"}
                att["run_status"] = "success"
            if run_out.get("status") == "success":
                # persistence re-check: expert records have shown eventual-consistency
                # glitches after save; re-save the verified code if the record vanished
                persisted = False
                for _ in range(3):
                    g2 = xapi("/api/expert/get", {"name": expert_name})
                    if g2.get("status") == "success" and ("def " + expert_name) in g2.get("expert_code", ""):
                        persisted = True
                        break
                    xapi("/api/expert/save", {"name": expert_name,
                                              "description": str(spec.get("description", ""))[:900],
                                              "code": code, "kwargs": spec.get("kwargs") or {},
                                              "cspl": task.get("cspl", "fython")})
                    time.sleep(3)
                att["persisted"] = persisted
                rec.update({"status": "done", "mode": "build", "expert_name": expert_name,
                            "acceptance": att["run_summary"], "persisted": persisted,
                            "finished_at": now()})
                rec.setdefault("attempts", []).append(att)
                save_manifest()
                return {"status": "success", "task_id": tid, "result": "built_and_verified",
                        "expert_name": expert_name, "attempts_used": attempt,
                        "acceptance": att["run_summary"], "manifest_path": str(manifest_path)}
            raise RuntimeError("acceptance run failed: " + str(run_out)[:300])

        except Exception as e:
            att["error"] = str(e)[:400]
            rec.setdefault("attempts", []).append(att)
            save_manifest()
            feedback = ("\n\nПРЕДЫДУЩАЯ ПОПЫТКА ПРОВАЛИЛАСЬ, ИСПРАВЬ:\n" + str(e)[:500] +
                        "\nСгенерируй ИСПРАВЛЕННУЮ версию целиком.")
            time.sleep(2)

    rec.update({"status": "failed", "finished_at": now()})
    save_manifest()
    return {"status": "error", "task_id": tid, "expert_name": expert_name,
            "message": "all attempts failed; see manifest",
            "last_error": rec["attempts"][-1].get("error") if rec.get("attempts") else None,
            "manifest_path": str(manifest_path)}