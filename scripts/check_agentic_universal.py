#!/usr/bin/env python3
"""Универсальная матрица Task Contract → Source Model → draft/run/repair/memory.

Живой Qwen здесь намеренно заменён детерминированными mocks: тест доказывает инварианты харнесса,
а не стабильность внешнего провайдера или конкретные слова одной отрасли.
"""
import importlib.util
import json
import sys
import tempfile
import types
import zipfile
from pathlib import Path


ROOT = Path(__file__).resolve().parent.parent
MODULE = ROOT / "ui" / "wz_agentic.py"


def load_module():
    platform = types.ModuleType("wz_platform")
    platform.BASE = "https://example.invalid"
    platform.CONFIG = {"auth_token": "test"}
    platform.api = lambda *a, **k: {}
    platform.qwen_agent = lambda: "agent_qwen"
    platform.run_expert = lambda *a, **k: {}
    llm = types.ModuleType("wz_llm")
    llm.design_agent = lambda: "agent_judge"
    sys.modules["wz_platform"] = platform
    sys.modules["wz_llm"] = llm
    spec = importlib.util.spec_from_file_location("wz_agentic_universal", MODULE)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def profile(name, sections, extension=".xlsx"):
    return {"name": name, "extension": extension, "bytes": 100, "sha256": name + "-sha",
            "workbook": [{"title": s[0], "max_row": 4, "max_column": len(s[1]), "header_row": 1,
                          "columns": s[1], "sample_rows": [s[1], s[2] if len(s) > 2 else ["x"]]}
                         for s in sections]}


def package(inputs, goal="Обработать входы и доказать результат", capabilities=None):
    contract = {"version": 1, "business_goal": goal, "original_request": goal,
                "required_result": {"success_criteria": ["результат доказан"]},
                "inputs": [{"name": x["name"], "profile": x} for x in inputs], "sha256": "contract"}
    return {"task_contract": contract, "original_request": goal, "inputs": inputs,
            "available_plugins_and_experts": capabilities or [], "working_memory": {"version": 1, "entries": []}}


def source_for(pkg, strategy="holistic_build", operation="обработать", normalizations=None,
               selected=None):
    sources = []
    for item in pkg["inputs"]:
        sections = [{"name": sh["title"], "role": "табличная сущность",
                     "entities": ["records"],
                     "identifier_fields": (sh.get("columns") or [])[:1],
                     "evidence": ["колонки: " + ", ".join(sh.get("columns") or [])]}
                    for sh in item.get("workbook") or []]
        if item.get("text_sample") is not None:
            sections.append({"name": "document", "role": "текстовый документ",
                             "entities": ["document"], "evidence": ["извлечён текст"]})
        elif item.get("columns") or item.get("sample_rows"):
            sections.append({"name": "data", "role": "табличная сущность", "entities": ["records"],
                             "identifier_fields": (item.get("columns") or [])[:1],
                             "evidence": ["фактические колонки"]})
        sources.append({"name": item["name"], "role": "обязательный вход",
                        "entities": ["business records"], "sections": sections,
                        "evidence": ["фактический профиль"]})
    return {"status": "ready", "strategy": strategy, "reason": "следует из контракта и профиля",
            "sources": sources,
            "operations": [{"name": operation, "inputs": [x["name"] for x in sources],
                            "normalizations": normalizations or [], "evidence": ["Task Contract"]}],
            "acceptance_criteria": ["все обязательные входы обработаны"],
            "selected_capabilities": selected or [], "missing_capability": "", "question": ""}


def assert_matrix(mod):
    # 1. Один Excel: фильтрация/агрегация/отчёт; название и колонки не отраслевые.
    one = package([profile("delta.xlsx", [("Ledger", ["category", "amount"], ["A", 12])])],
                  "Отфильтровать записи и агрегировать сумму")
    valid_one = source_for(one, strategy="build", operation="filter and aggregate")
    assert mod._normalize_source_model(valid_one, one)["ok"]
    # JSON-mode drift is repaired only when the remaining body is already complete and then still
    # passes every deterministic source/operation/evidence check.
    missing_contract_fields = json.loads(json.dumps(valid_one))
    missing_contract_fields.pop("status")
    missing_contract_fields.pop("strategy")
    inferred = mod._normalize_source_model(missing_contract_fields, one)
    assert inferred["ok"] and inferred["model"]["status"] == "ready"
    assert "status:<missing>→ready" in inferred["model"]["contract_repairs"]
    wrapped = mod._normalize_source_model({"source_model": valid_one}, one)
    assert wrapped["ok"] and "unwrapped:source_model" in wrapped["model"]["contract_repairs"]
    assert mod._normalize_source_model({"reason": "без контракта"}, one)["ok"] is False
    nonsense = json.loads(json.dumps(valid_one)); nonsense["status"] = "maybe"
    assert mod._normalize_source_model(nonsense, one)["ok"] is False
    # Невалидная первая модель не проходит молча: второй вызов получает точную ошибку валидатора.
    prompts = []
    invalid_one = json.loads(json.dumps(valid_one)); invalid_one["sources"][0]["role"] = ""
    replies = [invalid_one, valid_one]
    original_llm_json = mod._llm_json
    mod._llm_json = lambda llm, prompt, **kw: prompts.append(prompt) or replies.pop(0)
    retried = mod.build_source_model(one, {}, max_tries=2)
    mod._llm_json = original_llm_json
    assert retried["ok"] and len(prompts) == 2 and "не определена роль" in prompts[1]

    # Provider may spend its whole response on reasoning and return no final JSON (Gulzhan 21.07).
    # Only that empty-shape failure receives a neutral identity model; semantic invalidity remains
    # fail-closed and is never papered over.
    empty_calls = []
    mod._llm_json = lambda *a, **k: empty_calls.append(True) or {}
    neutral = mod.build_source_model(one, {}, max_tries=2)
    assert neutral["ok"] and neutral["model"]["strategy"] == "holistic_build"
    assert len(empty_calls) == 1
    assert neutral["model"]["sources"][0]["source_id"] == "source_001"
    assert "fallback:deterministic_identity_after_empty_qwen" in neutral["model"]["contract_repairs"]
    invalid_ref = source_for(one)
    invalid_ref["sources"][0]["name"] = "invented.xlsx"
    mod._llm_json = lambda *a, **k: invalid_ref
    still_rejected = mod.build_source_model(one, {}, max_tries=2)
    mod._llm_json = lambda *a, **k: {"reason": "не могу доказать роли входов"}
    reason_only = mod.build_source_model(one, {}, max_tries=1)
    mod._llm_json = original_llm_json
    assert not still_rejected["ok"] and "неизвестный вход" in still_rejected["why"]
    assert not reason_only["ok"] and "<missing>" in reason_only["why"]

    # 2. Несколько таблиц и различные ID: преобразование допустимо только с evidence.
    multi = package([profile("left.xlsx", [("Items", ["external_id", "qty"], ["0012", 2])]),
                     profile("right.xlsx", [("Export", ["reference", "value"], ["12", 2])])],
                    "Сопоставить связанные сущности")
    norm = [{"field": "external_id/reference", "method": "канонический числовой ключ",
             "evidence": "обе колонки состоят из цифровых строк, владелец разрешил", "requires_owner": False}]
    assert mod._normalize_source_model(source_for(multi, normalizations=norm), multi)["ok"]
    unsafe = source_for(multi, normalizations=[dict(norm[0], requires_owner=True)])
    assert mod._normalize_source_model(unsafe, multi)["ok"] is False

    # macOS NFD и описательное имя Qwen не участвуют в identity: authoritative source/section ID
    # разрешаются детерминированно, а наружу возвращается человекочитаемый NFC.
    mac = package([profile("леи\u0306блы.xlsx", [("Данные", ["номер", "расход"], ["00015", 1])])],
                  "Сопоставить учётные записи")
    mac_raw = source_for(mac, strategy="build", operation="read source by stable ids")
    mac_raw["sources"][0].update({"source_id": "source_001", "name": "1C Excel exports (леи_блы)"})
    mac_raw["sources"][0]["sections"][0].update({
        "section_id": "source_001_section_001", "name": "описательное имя"})
    mac_raw["operations"][0]["inputs"] = ["source_001_section_001"]
    mac_checked = mod._normalize_source_model(mac_raw, mac)
    assert mac_checked["ok"], mac_checked
    assert mac_checked["model"]["sources"][0]["name"] == "лейблы.xlsx"
    assert mac_checked["model"]["operations"][0]["inputs"] == ["source_001_section_001"]
    legacy_nfd = source_for(mac, strategy="build", operation="legacy exact name")
    assert mod._normalize_source_model(legacy_nfd, mac)["ok"]

    # 3. Excel + PDF и 4. документный поток: роли следуют из схемы/текста, не из порядка.
    excel_pdf = package([profile("numbers.xlsx", [("Data", ["key", "metric"], ["K1", 8])]),
                         {"name": "scan.pdf", "extension": ".pdf", "bytes": 20,
                          "sha256": "pdf-sha", "text_sample": "Certificate key K1"}],
                        "Извлечь и сверить данные")
    assert mod._normalize_source_model(source_for(excel_pdf, operation="extract and reconcile"), excel_pdf)["ok"]
    docs = package([{"name": "memo.docx", "extension": ".docx", "bytes": 20,
                     "sha256": "doc-sha", "text_sample": "Request type: service"}],
                   "Классифицировать документ и вернуть структурированный результат")
    assert mod._normalize_source_model(source_for(docs, strategy="build", operation="semantic classification"), docs)["ok"]

    # 5. Смысловой LLM-шаг и 6. ветвление/объединение выбирают стратегию, а не жёсткую цепочку.
    semantic = source_for(docs, strategy="holistic_build", operation="Qwen semantic decision")
    assert mod._normalize_source_model(semantic, docs)["model"]["strategy"] == "holistic_build"
    branches = package([profile("a.xlsx", [("A", ["id", "a"], ["x", 1])]),
                        profile("b.xlsx", [("B", ["id", "b"], ["x", 2])]),
                        profile("c.xlsx", [("C", ["id", "c"], ["x", 3])])], "Объединить независимые ветки")
    assert mod._normalize_source_model(source_for(branches, operation="parallel branches then join"), branches)["ok"]

    # 7. Недостаток данных: ровно один вопрос; 8. capability: кандидат не становится installed/ready.
    need = {"status": "need_human", "strategy": "need_human", "reason": "не задана политика",
            "sources": [], "operations": [], "acceptance_criteria": [], "selected_capabilities": [],
            "missing_capability": "", "question": "Как трактовать записи без даты?"}
    assert mod._normalize_source_model(need, one)["model"]["status"] == "need_human"
    missing_question = dict(need, question="")
    assert mod._normalize_source_model(missing_question, one)["ok"] is False
    owner_session = {"answers": {}, "decisions": [], "waiting_build": {"build_id": "old"}}
    owner_id = mod.apply_owner_clarification(
        owner_session, "Использованными считать строки с расходом?", "Да, расход > 0", "build_old",
        "2026-07-21T00:00:00+00:00")
    owner_contract = mod.make_task_contract(owner_session, {}, "", one["inputs"])
    assert owner_id in owner_session["answers"] and not owner_session.get("waiting_build")
    assert any(x["answer"] == "Да, расход > 0" for x in owner_contract["interview"])
    assert owner_contract["owner_decisions"][-1]["type"] == "builder_clarification"
    acquire = {"status": "acquire", "strategy": "acquire", "reason": "нужна OCR-модель",
               "sources": [], "operations": [], "acceptance_criteria": [], "selected_capabilities": [],
               "missing_capability": "OCR handwriting model",
               "question": "Подключить отдельно проверенную OCR-модель?"}
    acq = mod._normalize_source_model(acquire, one)
    assert acq["ok"] and acq["model"]["status"] == "acquire"

    # Reuse/compose разрешены только для capability из релевантного каталога.
    cap_pkg = package(one["inputs"], capabilities=[{"expert": "sheet_reader", "purpose": "read tables"}])
    reuse = source_for(cap_pkg, strategy="reuse", selected=["sheet_reader"])
    assert mod._normalize_source_model(reuse, cap_pkg)["ok"]
    assert mod._normalize_source_model(source_for(cap_pkg, strategy="reuse", selected=["invented"]), cap_pkg)["ok"] is False

    # 9. Смена схемы: ссылка на пропавший лист fail-closed; возможна честная остановка.
    changed = json.loads(json.dumps(source_for(one)))
    changed["sources"][0]["sections"][0]["name"] = "OldSheet"
    assert mod._normalize_source_model(changed, one)["ok"] is False
    assert mod._normalize_source_model(dict(need, question="Какое новое поле заменяет category?"), one)["ok"]


def acceptance_cases(mod, root):
    import openpyxl
    inp = root / "input.csv"
    inp.write_text("id,value\nA,1\n", encoding="utf-8")
    out = root / "out"
    out.mkdir(exist_ok=True)
    md = out / "report.md"
    md.write_text("# Result\n\nNo matching records; checked 1 input and 1 row with criterion X. "
                  "The empty result is expected and supported by the recorded acceptance evidence.\n",
                  encoding="utf-8")
    xlsx = out / "report.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["criterion", "result"]); ws.append(["X", 0]); wb.save(xlsx)
    # 10. Корректный пустой результат принимается с доказательством.
    empty = {"status": "success", "summary": {"processed_files": 1, "total_count": 0, "matched": 0},
             "evidence": {"files_used": ["input.csv"], "acceptance_checks": [
                 {"criterion": "проверены все строки", "passed": True, "evidence": "1 строка, 0 совпадений"}]},
             "report_md": str(md), "report_xlsx": str(xlsx)}
    assert mod.validate_result(empty, [inp])["ok"]
    # 11. Формальный success без фактов отклоняется.
    false_success = {"status": "success", "summary": {"processed_files": 1, "total_count": 0}}
    rejected = mod.validate_result(false_success, [inp])
    assert not rejected["ok"] and any("acceptance_checks" in x for x in rejected["issues"])
    return inp, empty, false_success


def gulzhan_regression(mod, root):
    """Исторический 4/4 — только регрессионные факты, никаких отраслевых правил в production."""
    sid = "wz_gulzhan_regression"
    paths = [root / "certificates_register.xlsx", root / "labels_register.xlsx", root / "erp_export.xlsx"]
    for index, path in enumerate(paths):
        path.write_text("synthetic anonymized fixture %d\n" % index, encoding="utf-8")
    inputs = [profile(paths[0].name, [("Registry-A", ["identifier", "state"], ["0000001", "ok"])]),
              profile(paths[1].name, [("Registry-B", ["identifier", "state"], ["0000002", "ok"])]),
              profile(paths[2].name, [("Turnover", ["reference", "quantity"], ["0000001", 1])])]
    pkg = package(inputs, "Сверить несколько реестров с выгрузкой учётной системы")
    pkg.update({"rules": ["не выполнять внешние записи"], "fields": {}, "source_config": {}})
    normalized = mod._normalize_source_model(
        source_for(pkg, strategy="holistic_build", operation="reconcile all registries"), pkg)
    assert normalized["ok"]
    pkg["source_model"] = normalized["model"]
    mod._refresh_package(pkg)
    prepared = {"ok": True, "package": pkg, "source_model": normalized["model"],
                "source_memory_ids": []}
    (root / (sid + ".json")).write_text(json.dumps({
        "session_id": sid, "client_name": "Synthetic regression",
        "agentic_memory": {"version": 1, "entries": []}}, ensure_ascii=False), encoding="utf-8")

    report_md = root / "gulzhan_report.md"
    report_md.write_text("# Synthetic reconciliation\n\nAll three inputs processed. No matches were reported. "
                         "This is deliberately the historical false-success fixture.\n", encoding="utf-8")
    import openpyxl
    report_xlsx = root / "gulzhan_report.xlsx"
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["entity", "source_count", "matched"])
    ws.append(["type_a", 1063, 0]); ws.append(["type_b", 4889, 0]); wb.save(report_xlsx)
    historical = {
        "status": "success",
        "summary": {"processed_files": 3, "erp_rows": 18500, "type_a_count": 1063,
                    "type_b_count": 4889, "type_a_matches": 0, "type_b_matches": 0},
        "evidence": {"files_used": [p.name for p in paths], "acceptance_checks": [
            {"criterion": "все входы прочитаны", "passed": True, "evidence": "обработано 3 файла"}]},
        "report_md": str(report_md), "report_xlsx": str(report_xlsx),
    }
    validation = mod.validate_result(historical, paths, pkg)
    assert validation["ok"] and validation["summary"]["type_a_matches"] == 0

    repaired = json.loads(json.dumps(historical))
    repaired["summary"]["type_a_matches"] = 71
    repaired["summary"]["type_b_matches"] = 203
    repaired["evidence"]["acceptance_checks"].append({
        "criterion": "сопоставление доказано", "passed": True,
        "evidence": "нормализация ключей подтверждена наблюдаемыми парами"})

    feedbacks, events, promotions = [], [], []

    def create(name, task_package, feedback, llm):
        feedbacks.append(json.loads(json.dumps(feedback, default=str)) if feedback else "")
        if len(feedbacks) <= 3:
            return {"ok": False, "why": "synthetic generation failure %d" % len(feedbacks)}
        return {"ok": True, "code_sha256": "changed-%d" % len(feedbacks)}

    verdicts = [{
        "verdict": "fail", "confidence": 0.99,
        "issues": ["нулевые совпадения против ожидаемой связи источников не доказаны"],
        "owner_question": "", "memory": {"concepts": [], "rules": []},
        "rejected_hypotheses": [
            {"text": "идентификаторы совпадают как строки", "evidence": "0 совпадений из 1063 и 4889",
             "rejection_reason": "возможны разные представления идентификаторов"},
            {"text": "выбраны правильные разделы", "evidence": "профили содержат разные сущности",
             "rejection_reason": "возможно неверно выбраны листы и бизнес-сущности"}],
    }, {
        "verdict": "pass", "confidence": 0.96, "issues": [], "owner_question": "",
        "memory": {"concepts": [], "rules": []}, "rejected_hypotheses": [],
    }]
    runs = [historical, repaired]
    mod._create_or_update = create
    mod.run_expert = lambda *a, **k: runs.pop(0)
    mod.judge_result = lambda *a, **k: verdicts.pop(0)
    mod._promote_expert = lambda draft, stable, task_package: (
        promotions.append((draft, stable)) or {"ok": True, "code_sha256": "stable"})
    mod._delete_draft = lambda *a, **k: None
    mod.time.sleep = lambda *_: None
    result = mod.build_agentic_solution(
        sid, "gulzhan_legacy_4_of_4", "regression", paths, root, root, {"agent_id": "agent_qwen"},
        progress=lambda *args: events.append(args), max_creation_attempts=4,
        max_run_repairs=2, max_acceptance_repairs=2, max_total_attempts=4,
        prepared_context=prepared)

    # Старая система остановилась на критической диагностике 4/4. Новая сохраняет отдельный repair
    # budget: первый runnable draft на общем шаге 4 реально получает шаг 5 с переданным уроком.
    assert result["ok"] and result["attempts"][3]["attempt"] == 4
    assert result["attempts"][3]["result"]["summary"]["type_a_matches"] == 0
    assert result["attempts"][3]["judge"]["verdict"] == "fail"
    assert result["attempts"][4]["attempt"] == 5 and result["last_applied_lesson"]
    assert "нулевые совпадения" in json.dumps(feedbacks[4], ensure_ascii=False)
    titles = [str(event[1]) for event in events]
    assert not any("Qwen исправляет" in title for title in titles)
    assert any("урок передан следующему ремонту" in title for title in titles)
    assert promotions


def repair_and_memory(mod, root, inp, good, bad):
    sid = "wz_universal"
    (root / (sid + ".json")).write_text(json.dumps({
        "session_id": sid, "client_name": "Universal", "questionnaire_task": "Обработать пакет",
        "answers": {"adaptive_91": {"question": "Какой результат?", "answer": "Проверенный отчёт"}},
        "rules": ["не выполнять внешние записи"]}, ensure_ascii=False), encoding="utf-8")
    (root / (sid + "_blueprint.json")).write_text(json.dumps({"blueprint": {
        "process_name": "Universal", "goal": "Обработать пакет", "sample_test_plan": {
            "success_criteria": ["отчёт доказан"], "steps": ["прочитать", "проверить"]}}},
        ensure_ascii=False), encoding="utf-8")
    (root / (sid + "_build_plan.json")).write_text('{"plan":{"tasks":[]}}', encoding="utf-8")

    def ready(pkg, llm, max_tries=2):
        raw = source_for(pkg, strategy="build", operation="read and report")
        return mod._normalize_source_model(raw, pkg)

    mod.build_source_model = ready
    create_calls, packages, promotions, deletions = [], [], [], []

    def create(name, pkg, feedback, llm):
        create_calls.append(feedback)
        packages.append(json.loads(json.dumps(pkg, default=str)))
        if len(create_calls) <= 3:
            return {"ok": False, "why": "generation unavailable %d" % len(create_calls)}
        return {"ok": True, "code_sha256": "changed-%d" % (len(create_calls) - 3)}

    runs = [bad, bad, good]
    mod._create_or_update = create
    mod.run_expert = lambda *a, **k: runs.pop(0)
    mod.judge_result = lambda *a, **k: {
        "verdict": "pass", "confidence": 0.96, "issues": [], "owner_question": "",
        "memory": {"concepts": [{"text": "Вход является реестром", "evidence": "полный прогон"}],
                   "rules": [{"text": "Проверять все строки", "evidence": "acceptance"}]},
        "rejected_hypotheses": []}
    mod._promote_expert = lambda draft, stable, pkg: (
        promotions.append((draft, stable)) or {"ok": True, "code_sha256": "stable"})
    mod._delete_draft = lambda draft, agent="": deletions.append(draft)
    mod.time.sleep = lambda *_: None
    result = mod.build_agentic_solution(
        sid, "budget", "uni", [inp], root, root, {"agent_id": "agent_qwen"},
        max_creation_attempts=4, max_run_repairs=2, max_acceptance_repairs=2, max_total_attempts=6)
    # 12. Три create-fail + первый runnable fail + две реально изменённые repair-версии.
    assert result["ok"] and len(create_calls) == 6, result
    runnable = [x for x in result["attempts"] if x.get("runnable_attempt")]
    assert len(runnable) == 3 and runnable[-1]["budgets"]["run_repair"] == 2
    assert promotions and promotions[0][1] == "uni_run_process"
    assert all("__draft_" in x[0] for x in promotions)
    # 14. Следующая попытка получила rejected-урок; success публикует только verified.
    assert any(x.get("status") == "rejected" for x in packages[4]["working_memory"]["entries"])
    assert result["verified_memory"] and all(x["status"] == "verified" for x in result["verified_memory"])
    assert all(x["status"] == "verified" for x in result["verified_memory"])

    # 13. Неизменившийся код — no_progress, не repair; итог честно stalled.
    sid_fail = "wz_universal_fail"
    base_session = json.loads((root / (sid + ".json")).read_text(encoding="utf-8"))
    base_session.pop("agentic_memory", None); base_session["session_id"] = sid_fail
    (root / (sid_fail + ".json")).write_text(json.dumps(base_session, ensure_ascii=False), encoding="utf-8")
    for suffix in ("_blueprint.json", "_build_plan.json"):
        (root / (sid_fail + suffix)).write_text((root / (sid + suffix)).read_text(encoding="utf-8"),
                                                encoding="utf-8")
    same_calls = []
    mod._create_or_update = lambda name, pkg, feedback, llm: (
        same_calls.append(feedback) or {"ok": True, "code_sha256": "same"})
    mod.run_expert = lambda *a, **k: bad
    promotions.clear()
    stalled = mod.build_agentic_solution(
        sid_fail, "stalled", "uni", [inp], root, root, {"agent_id": "agent_qwen"},
        max_creation_attempts=1, max_run_repairs=2, max_total_attempts=4)
    assert not stalled["ok"] and stalled["code"] == "agentic_stalled"
    assert sum(1 for x in stalled["attempts"] if x.get("phase") == "no_progress") == 3
    assert stalled["budgets"]["run_repair"]["used"] == 0 and not promotions
    # Failed memory survives this session but does not become verified process memory.
    rebuilt = mod.make_task_package(sid_fail, [inp], root, llm={"agent_id": "agent_qwen"})
    assert any(x.get("status") == "rejected" for x in rebuilt["working_memory"]["entries"])
    assert mod._verified_memory(stalled["working_memory"]) == []

    # need_human/acquire stop before any draft and cannot publish memory.
    called = []
    mod._create_or_update = lambda *a, **k: called.append(True) or {"ok": True, "code_sha256": "x"}
    mod.build_source_model = lambda *a, **k: {"ok": True, "model": {
        "status": "need_human", "strategy": "need_human", "reason": "ambiguous",
        "sources": [], "operations": [], "question": "Как трактовать пустое поле?", "sha256": "n"}}
    stopped = mod.build_agentic_solution(sid_fail, "human", "uni", [inp], root, root, {"agent_id": "agent_qwen"})
    assert stopped["code"] == "needs_owner_input" and not stopped["draft_created"] and not called
    mod.build_source_model = lambda *a, **k: {"ok": True, "model": {
        "status": "acquire", "strategy": "acquire", "reason": "missing OCR",
        "sources": [], "operations": [], "question": "Подключить OCR отдельно?",
        "missing_capability": "OCR", "sha256": "a"}}
    stopped = mod.build_agentic_solution(sid_fail, "acquire", "uni", [inp], root, root, {"agent_id": "agent_qwen"})
    assert stopped["code"] == "capability_missing" and stopped["verified_memory"] == [] and not called


def profile_variants(mod, root):
    # Варьируем порядок листов/названия/колонки и доказываем, что профилировщик не берёт только первые четыре.
    import openpyxl
    book = root / "renamed.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for i in range(6):
        ws = wb.create_sheet("Section-%d" % i)
        ws.append(["field_%d" % i, "measure_%d" % i]); ws.append(["id-%d" % i, i])
    wb.save(book)
    prof = mod.profile_file(book)
    assert len(prof["workbook"]) == 6 and prof["workbook"][5]["columns"] == ["field_5", "measure_5"]
    doc = root / "flow.docx"
    with zipfile.ZipFile(doc, "w") as zf:
        zf.writestr("word/document.xml", "<w:document><w:t>Document request alpha</w:t></w:document>")
    assert "Document request alpha" in mod.profile_file(doc).get("text_sample", "")


def adaptive_gulzhan_timeout(mod, root):
    """13 mixed inputs must split after one measured workload timeout, not run 4×15 minutes."""
    sid = "wz_gulzhan_adaptive"
    paths = []
    inputs = []
    for index in range(13):
        suffix = ".pdf" if index < 5 else ".xlsx"
        path = root / ("input_%02d%s" % (index, suffix))
        path.write_bytes(("fixture-%d" % index).encode("utf-8"))
        paths.append(path)
        inputs.append(profile(path.name, [("Data", ["id", "value"], [str(index), index])], suffix))
    pkg = package(inputs, "Сопоставить смешанный пакет и сформировать проверенный отчёт")
    normalized = mod._normalize_source_model(source_for(pkg, operation="reconcile mixed sources"), pkg)
    assert normalized["ok"]
    pkg["source_model"] = normalized["model"]
    mod._refresh_package(pkg)
    prepared = {"ok": True, "package": pkg, "source_model": normalized["model"],
                "source_memory_ids": []}
    (root / (sid + ".json")).write_text(json.dumps({"session_id": sid}), encoding="utf-8")
    created, ran, events = [], [], []
    mod._create_or_update = lambda *a, **k: (
        created.append(True) or {"ok": True, "code_sha256": "oversized-v1",
                                 "strategy_sha256": "strategy-v1"})
    mod.run_expert = lambda *a, **k: (
        ran.append(k.get("wait")) or {"status": "error", "message": "task timed out"})
    mod._delete_draft = lambda *a, **k: None
    mod.time.sleep = lambda *_: None
    step_contract = {"id": "reconcile", "version": 1, "title": "Сверить пакет",
                     "purpose": "Сопоставить все источники", "acceptance": {
                         "semantic_criteria": ["сверка доказана"], "minimum_confidence": 0.7}}
    result = mod.build_agentic_solution(
        sid, "adaptive", "gul", paths, root, root, {"agent_id": "agent_qwen"},
        progress=lambda *args: events.append(args), max_creation_attempts=2,
        max_run_repairs=2, max_acceptance_repairs=2, max_total_attempts=6,
        prepared_context=prepared, step_contract=step_contract,
        expert_name_override="gul_reconcile_v1")
    assert not result["ok"] and result["control_action"] == "split_step", result
    assert result["failure_decision"]["failure_class"] == "workload_timeout"
    assert len(created) == 1 and len(ran) == 1 and ran[0] == 240
    assert not result["owner_question"]
    assert any("дробит" in str(event[1]).casefold() for event in events)


def transient_retry_reuses_draft_once(mod, root, inp, good):
    """A first transport timeout retries the saved expert, without another code generation."""
    sid = "wz_transient_retry"
    (root / (sid + ".json")).write_text(json.dumps({"session_id": sid}), encoding="utf-8")
    pkg = package([profile(inp.name, [("Data", ["id", "value"], ["A", 1])], ".csv")],
                  "Produce a verified report")
    normalized = mod._normalize_source_model(source_for(pkg, operation="read and report"), pkg)
    assert normalized["ok"]
    pkg["source_model"] = normalized["model"]
    mod._refresh_package(pkg)
    prepared = {"ok": True, "package": pkg, "source_model": normalized["model"],
                "source_memory_ids": []}
    create_calls, run_waits = [], []
    code = "def transient_retry(source_file='', output_dir='', **kwargs):\n    return {'status':'success'}\n"

    def create(*args, **kwargs):
        create_calls.append(True)
        return {"ok": True, "code": code, "code_sha256": "transient-code-v1",
                "strategy_sha256": "transient-strategy-v1"}

    results = [{"status": "error", "message": "The read operation timed out"}, good]
    mod._create_or_update = create
    mod.run_expert = lambda *a, **k: run_waits.append(k.get("wait")) or results.pop(0)
    mod.judge_result = lambda *a, **k: {
        "verdict": "pass", "confidence": 1.0, "issues": [], "owner_question": "",
        "memory": {"concepts": [], "rules": []}, "rejected_hypotheses": []}
    mod._promote_expert = lambda *a, **k: {"ok": True, "code_sha256": "stable-transient"}
    mod._delete_draft = lambda *a, **k: None
    mod.time.sleep = lambda *_: None
    result = mod.build_agentic_solution(
        sid, "transient", "transient", [inp], root, root, {"agent_id": "agent_qwen"},
        max_creation_attempts=1, max_run_repairs=1, max_acceptance_repairs=1,
        max_total_attempts=3, prepared_context=prepared)
    assert result["ok"], result
    assert len(create_calls) == 1 and run_waits == [600, 600]
    assert any(row.get("action") == "retry_transient" for row in result["controller_history"])
    assert not result.get("owner_question")

    repair_prompt = mod._build_prompt("repair_envelope", pkg, {
        "required_next_action": "repair_output_contract",
        "previous_expert_code": code,
        "structural_issues": ["missing required artifact: step_result_json"],
    }, "agent_qwen")
    assert "РЕМОНТ ТОЛЬКО ВЫХОДНОГО КОНТРАКТА" in repair_prompt
    assert "не меняй выбор входов" in repair_prompt and code.strip() in repair_prompt


def main():
    mod = load_module()
    assert_matrix(mod)
    with tempfile.TemporaryDirectory(prefix="wz_universal_") as td:
        root = Path(td)
        inp, good, bad = acceptance_cases(mod, root)
        profile_variants(mod, root)
        gulzhan_regression(mod, root)
        repair_and_memory(mod, root, inp, good, bad)
        adaptive_gulzhan_timeout(mod, root)
        transient_retry_reuses_draft_once(mod, root, inp, good)
    print("универсальная матрица: 15 классов + adaptive timeout/split + transient retry + repair budget + память + fail-closed stops ✓")


if __name__ == "__main__":
    main()
