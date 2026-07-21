"""Кластер СТРОЙКИ моста Визарда (Фаза 1, шов #3).

Выделено из server.py: аудит собранных экспертов, кодоген стадии (_build_one), шаблоны оркестратора и
kp-стадии, сборка оркестратора, главный оркестратор стройки _run_build (план→кодоген→аудит→продовый агент).
Зависит от wz_platform (CONFIG/BASE/api/run_expert/qwen_agent) и wz_llm (run_llm_expert/design_agent).
Наружу торчит только _run_build (его зовёт HTTP-хендлер /x/build).
"""
import json
import re
import time
import base64
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from wz_platform import CONFIG, BASE, api, run_expert, qwen_agent
from wz_llm import run_llm_expert, design_agent, gen_panel_manifest, llm_transient_error
from wz_agentic import build_agentic_solution, prepare_task_context, _builder_brief
from wz_process import (accept_step as universal_accept_step, atomic_write_json,
                        budget_preflight as universal_budget_preflight,
                        block_for_human as universal_block_for_human,
                        checkpoint as process_checkpoint, memory_entry as universal_memory_entry,
                        expand_subgraph as universal_expand_subgraph,
                        normalize_step_result, process_from_blueprint,
                        record_usage as universal_record_usage,
                        recover_after_restart as universal_recover_after_restart,
                        process_status as universal_process_status, ready_steps as universal_ready_steps,
                        step_map as universal_step_map, transition_step as universal_transition_step)

SESS_DIR = Path.home() / "extella_wizard" / "sessions"
RUNS_DIR = Path.home() / "extella_wizard" / "runs"
HOST_TARGET = "85800354-f7b7-449f-b526-9357cd91f780"  # managed-хостинг VPS (PS.kz)


def _audit_experts(names):
    """Детерминированный предзапусковый аудит кода построенных экспертов."""
    import re as _re
    issues = []
    for n in names:
        e = api("/api/expert/get", {"name": n, "global": True})
        code = e.get("expert_code", "") if isinstance(e, dict) else ""
        if not code:
            continue
        checks = {
            "секрет в коде": bool(_re.search(r"(sk-[A-Za-z0-9]{20}|api[_-]?key\s*=\s*['\"][A-Za-z0-9]{16})", code)),
            "отправка почты": ("smtplib" in code or "sendmail" in code),
            "внешняя запись": bool(_re.search(r"https?://(?!api\.extella\.ai|disnet\.extella\.ai)[a-z0-9.]+/", code)),
            "путь устройства": ("/Users/" in code or "/home/" in code),
        }
        for k, v in checks.items():
            if v:
                issues.append(n + ": " + k)
    verdict = "allow" if not issues else "allow-with-confirmation"
    return {"verdict": verdict, "issues": issues}


def sample_preflight(session_id):
    """WZ-07 (ТЗ v2 §9.5): проверка образца ДО стройки. Файловые архетипы без примера данных
    раньше падали «нет входа для стадии» ПОСЛЕ минут сборки — теперь честный отказ сразу."""
    _FILE_ARCHETYPES = ("document_processing", "recurring_report", "flow_quality_control")
    _, sample_file = _inspect_sample(session_id)
    if sample_file:
        return {"ok": True, "sample_file": str(sample_file)}
    arch = ""
    try:
        bp = json.loads((SESS_DIR / (session_id + "_blueprint.json")).read_text(encoding="utf-8")).get("blueprint", {})
        a = bp.get("archetype")
        arch = str(a.get("id") if isinstance(a, dict) else (a or ""))
    except Exception:
        pass
    src = None
    try:
        src = json.loads((SESS_DIR / (session_id + ".json")).read_text(encoding="utf-8")).get("source")
    except Exception:
        pass
    if arch in _FILE_ARCHETYPES and not src:
        return {"ok": False,
                "message": "Нет файла-образца данных — стройка упрётся в первую стадию. "
                           "Приложите пример (вкладка «Файлы») или подключите источник данных."}
    return {"ok": True, "sample_file": ""}


def _inspect_sample(session_id):
    """Реальные колонки загруженного файла-образца для грануднинга кодогена."""
    files = _sample_files(session_id)
    if not files:
        return "", None
    f = files[0]
    ext = f.suffix.lower()
    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[("" if c.value is None else str(c.value)) for c in r]
                    for r in ws.iter_rows(min_row=1, max_row=15)]
            hdr, best = 0, -1
            for i, r in enumerate(rows):
                sc = sum(1 for v in r if v.strip()) + sum(1 for v in r if v.strip() and not v.replace(".", "").replace("-", "").isdigit())
                if sc > best:
                    best, hdr = sc, i
            cols = [v for v in rows[hdr] if v.strip()]
            sample = rows[hdr + 1] if hdr + 1 < len(rows) else []
            hint = ("\n\nФАКТИЧЕСКАЯ СТРУКТУРА ФАЙЛА (СТРОЙ СТРОГО ПОД ЭТИ КОЛОНКИ, не выдумывай поля): "
                    + "лист '" + str(ws.title) + "', заголовки в строке #" + str(hdr + 1)
                    + ", колонки: " + json.dumps(cols, ensure_ascii=False)
                    + ", пример: " + json.dumps(sample, ensure_ascii=False)[:300])
            return hint, str(f)
        if ext == ".csv":
            import csv as _csv
            rd = list(_csv.reader(open(str(f), "r", encoding="utf-8", errors="replace")))
            cols = [v for v in (rd[0] if rd else []) if v.strip()]
            return ("\n\nФАКТИЧЕСКАЯ СТРУКТУРА ФАЙЛА (СТРОЙ ПОД ЭТИ КОЛОНКИ): csv, колонки: "
                    + json.dumps(cols, ensure_ascii=False)), str(f)
    except Exception:
        return "", str(f)
    return "", str(f)


def _sample_files(session_id):
    """Все приложенные образцы в стабильном порядке.

    Старый Строитель молча брал files[0]. Для процесса, которому одновременно нужны Excel и PDF,
    это превращало настоящий DAG в фиктивную линейную цепочку и давало случайный результат в
    зависимости от имени файла. Список нужен гейту топологии: неоднозначный вход нельзя скрывать.
    """
    fdir = SESS_DIR / (session_id + "_files")
    if not fdir.is_dir():
        return []
    return [p for p in sorted(fdir.iterdir()) if p.is_file()]


def _entity_columns(session_id):
    """Колонки-СУЩНОСТИ файла-образца — по которым МОЖНО осмысленно искать во внешних источниках
    (названия товаров, компаний, поставщиков). Возвращает (entity_cols, all_cols). Классификация
    детерминированная по нескольким строкам: колонка = сущность, если её заголовок не число и в
    примерах преобладают НЕЧИСЛОВЫЕ текстовые значения длиной ≥3 (не суммы, не даты, не id).
    Нужна, чтобы закрепить ключ веб-поиска за реальной текстовой колонкой, а не дать шагу искать
    по суммам/заголовкам (мусор Гульжан, 20.07). Пусто = искать не по чему → шаг не строим."""
    fdir = SESS_DIR / (session_id + "_files")
    if not fdir.is_dir():
        return [], []
    files = [p for p in sorted(fdir.iterdir()) if p.is_file()]
    if not files:
        return [], []
    f = files[0]
    ext = f.suffix.lower()
    rows = []
    try:
        if ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[("" if c.value is None else str(c.value)) for c in r]
                    for r in ws.iter_rows(min_row=1, max_row=25)]
            wb.close()
        elif ext == ".csv":
            import csv as _csv
            rows = list(_csv.reader(open(str(f), "r", encoding="utf-8", errors="replace")))[:25]
    except Exception:
        return [], []
    if not rows:
        return [], []
    # строка-заголовок: максимум непустых нечисловых ячеек
    hdr = max(range(min(len(rows), 10)),
              key=lambda i: sum(1 for v in rows[i] if str(v).strip() and not _looks_numeric(v)), default=0)
    header = [str(v).strip() for v in rows[hdr]]
    body = rows[hdr + 1:hdr + 21]
    all_cols, entity_cols = [], []
    for ci, name in enumerate(header):
        if not name:
            continue
        all_cols.append(name)
        vals = [str(r[ci]).strip() for r in body if ci < len(r) and str(r[ci]).strip()]
        if not vals:
            continue
        texty = sum(1 for v in vals if not _looks_numeric(v) and not _looks_date(v) and len(v) >= 3
                    and re.search(r"[^\W\d_]", v, re.U))   # есть буква
        if texty >= max(2, int(len(vals) * 0.6)) and not _looks_numeric(name):
            entity_cols.append(name)
    return entity_cols, all_cols


def _looks_numeric(v):
    return bool(re.fullmatch(r"[\d\s.,%+\-()]+", str(v).strip() or "x"))


def _looks_date(v):
    s = str(v).strip()
    return bool(re.search(r"\d{1,4}[.\-/]\d{1,2}[.\-/]\d{1,4}", s) or
                re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", s))


_WEBSEARCH_TASK_MARKERS = ("веб-поиск", "веб поиск", "интернет", "в сети", "в интернете", "поиск постав",
                           "внешн", "обогащ", "enrich", "web_search", "web search", "websearch",
                           "external", "lookup", "duckduckgo", "google", "найти в", "search the")


def _is_websearch_task(t):
    """Шаг плана — это внешний веб-поиск/обогащение? По маркерам в назначении/описании/имени и по
    capability_ids из blueprint. Нужно, чтобы закрепить за таким шагом колонку-сущность или не строить."""
    if not isinstance(t, dict):
        return False
    caps = t.get("capability_ids") or t.get("capabilities") or []
    if isinstance(caps, list) and any("web" in str(c).lower() or "external" in str(c).lower() or "enrich" in str(c).lower() for c in caps):
        return True
    blob = " ".join(str(t.get(k, "")) for k in ("purpose", "description", "goal", "title", "name", "expert_name")).lower()
    return any(m in blob for m in _WEBSEARCH_TASK_MARKERS)


_BUILD_SYS = """Ты — генератор кода СТАДИИ КОНВЕЙЕРА для платформы Extella. Верни ТОЛЬКО JSON:
{"code":"<полный код>", "description":"<англ.: что делает>"}

ЖЁСТКИЙ КОНТРАКТ СТАДИИ (соблюдать точно):
- Сигнатура РОВНО: def <ИМЯ>(input_path: str = "", output_path: str = "", rules_json: str = "", fields_json: str = "") -> dict. НИКАКИХ других параметров.
- ПРАВИЛА ВЛАДЕЛЬЦА (F2-контракт): rules_json — JSON-список правил словами, fields_json — JSON-словарь полей.
  Если непусты и НЕ начинаются с "{{" — распарси (json.loads в try) и ПРИМЕНИ релевантные правила к своей работе
  (фильтры порогов, пометки, доп-колонки, сортировка, что исключить); нерелевантные твоей стадии — игнорируй молча.
  Пустые/непарсящиеся — работай как обычно (обратная совместимость).
- ВХОД: читай из input_path. %(INPUT_DESC)s
- РАБОТА: %(PURPOSE)s
- ВЫХОД: запиши результат в output_path как JSON. %(OUTPUT_DESC)s Если это отчётная стадия — можешь дополнительно писать .md/.docx рядом (import docx через include), но JSON в output_path обязателен.
- ВЕРНИ компактный dict: {"status":"success","output_path":output_path, ...ключевые счётчики}. НЕ клади крупные данные в возврат.

Стандарт: первая строка $extens("include.py"); зависимости через include("import X",["extella-pip install X"]) (openpyxl/ docx — так; стдлиб json/csv/datetime — include("import json",[])); РОВНО ОДНА top-level функция (имя строго заданное), хелперы ВНУТРИ неё, не переопределяй include/load_module; валидация входов с ранним return {"status":"error"}; без хардкода путей/ключей; не обращаться к KV.

ЕСЛИ СТАДИЯ ИЩЕТ ВО ВНЕШНИХ ИСТОЧНИКАХ (интернет/веб-поиск): ищи ТОЛЬКО по осмысленным сущностям — названиям товаров, компаний, поставщиков из ТЕКСТОВЫХ полей. НИКОГДА не ищи по числам, суммам, датам, id и служебным полям — это даёт мусор не по теме. Не перебирай подряд ВСЕ значения ячеек: возьми 1–2 значимых текстовых поля на запись. Если осмысленного текстового поля для поиска нет — верни пустой результат с честным status, а не выдумывай запросы."""


def _stage_sanity(title, purpose, out_path, llm):
    """СМЫСЛОВАЯ приёмка вывода стадии (в дополнение к структурной). Структурная ловит «пусто/не
    JSON/нет чисел», но пропускает СТРУКТУРНО ЦЕЛЫЙ МУСОР: шаг «поиск поставщиков» искал в вебе по
    суммам → вики про игру; форма цела, галочка зелёная, смысл — ноль (Гульжан, 20.07).

    Возвращает {"ok": bool, "why": str}. ok=False → стадия не публикуется: причина возвращается
    в тот же цикл генерации как фактический урок для ремонта. Если независимый Qwen-судья
    недоступен или не подтвердил результат явно, гейт закрыт (fail-closed)."""
    try:
        raw = Path(out_path).read_text(encoding="utf-8")
        data = json.loads(raw)
    except Exception:
        return {"ok": True, "why": ""}   # не разобрали — не наша забота здесь
    recs = data if isinstance(data, list) else (
        data.get("records") or data.get("rows") or data.get("items") or [])
    if not isinstance(recs, list):
        recs = []

    # 1) ДЕШЁВЫЙ детерминированный флаг: поиск/запрос по чистому числу = заведомо мусор.
    numq = 0
    for r in recs[:80]:
        if not isinstance(r, dict):
            continue
        for k in ("search_query", "query", "q", "matched_category"):
            v = str(r.get(k, "")).strip()
            if v and re.fullmatch(r"[\d\s.,%+-]+", v):
                numq += 1
                break
    if numq >= 2:
        return {"ok": False, "why": "шаг искал по числовым значениям (суммам), а не по осмысленным "
                                    "запросам — найденное не относится к задаче"}

    # 2) Независимый Qwen-судья по каждой стадии. Production-путь fail-closed: структурно целый
    #    JSON ещё не доказывает, что стадия выполнила бизнес-цель Task Contract.
    ag = design_agent() or (llm or {}).get("agent_id")
    if not ag:
        return {"ok": False, "why": "нет независимого Qwen для смысловой приёмки стадии"}
    sample = json.dumps(data, ensure_ascii=False, default=str)[:1600]
    context = json.dumps((llm or {}).get("task_context") or {}, ensure_ascii=False, default=str)[:9000]
    prompt = ("Ты приёмщик качества автоматизации. Шаг: «" + str(title)[:80] + "». Его задача: "
              + str(purpose)[:160] + ".\nОбразец вывода шага:\n" + sample +
              "\nАвторитетный Task Contract + Source Model:\n" + context +
              "\n\nВывод ОСМЫСЛЕН для задачи шага, или это мусор (веб-результаты не по теме, поиск "
              "по числам, заглушки, данные из чужой области)? Ответь ТОЛЬКО JSON: "
              '{"sensible": true|false, "why": "<если нет — одной короткой фразой по-русски, что не так>"}. '
              "Будь строг к очевидному мусору, но НЕ придирайся к нормальным данным.")
    try:
        res = api("/api/agent/run", {"agent_id": ag, "input": prompt, "run_timeout": 50,
                                     "store": False, "temperature": 0}, timeout=60)
    except Exception as exc:
        return {"ok": False, "why": "смысловой судья недоступен: " + str(exc)[:100]}
    text = ""
    for it in (res or {}).get("output", []):
        if isinstance(it, dict) and it.get("type") == "message":
            for c in it.get("content", []):
                if isinstance(c, dict) and c.get("type") == "output_text":
                    text += c.get("text", "")
    text = text or (res or {}).get("output_text", "")
    m = re.search(r"\{.*\}", text, re.S)
    if not m:
        return {"ok": False, "why": "смысловой судья не вернул JSON"}
    try:
        v = json.loads(m.group(0))
    except Exception:
        return {"ok": False, "why": "смысловой судья вернул невалидный JSON"}
    if v.get("sensible") is False:
        return {"ok": False, "why": str(v.get("why") or "результат шага не относится к задаче")[:160]}
    if v.get("sensible") is not True:
        return {"ok": False, "why": str(v.get("why") or "смысловой судья не подтвердил результат")[:160]}
    return {"ok": True, "why": ""}


def _human_title(t, ns):
    """Человеческое имя шага для прогресса. План не всегда даёт title → раньше в прогрессе
    светилось СЫРОЕ имя эксперта (eur_send_rfq_emails), клиенту непонятное (Гульжан, 20.07).
    Берём по порядку: явный title → первая фраза purpose → очеловеченное имя эксперта."""
    ttl = str(t.get("title") or "").strip()
    if ttl and not re.fullmatch(r"[a-z0-9_]+", ttl):   # не техническое имя
        return ttl[:72]
    p = str(t.get("purpose") or "").strip()
    if p:
        first = re.split(r"[.;\n]", p)[0].strip()
        if len(first) >= 4:
            return (first[0].upper() + first[1:])[:72]
    nm = re.sub(r"^" + re.escape(str(ns)) + r"_", "", str(t.get("expert_name") or "")).replace("_", " ").strip()
    return (nm[0].upper() + nm[1:])[:72] if nm else "Шаг процесса"


def _is_pipeline_data_task(t):
    """True только для шага, который преобразует данные внутри вертикального среза.

    Планировщик иногда добавляет в tasks техническую обвязку 24/7: watcher папки, daemon,
    launchd/systemd/autostart. У неё нет контракта input_path → output_path, поэтому кодоген
    закономерно падает и раньше блокировал уже готовый бизнес-процесс (Гульжан, 20.07).
    Такая обвязка настраивается после сборки через источник/расписание и не является стадией DAG.
    """
    name = str(t.get("expert_name") or "").lower().replace("-", "_")
    text = " ".join(str(t.get(k) or "") for k in ("title", "purpose")).lower()
    name_markers = (
        "schedule_", "_schedule", "cron", "orchestr", "pipeline",
        "notif_", "_notif", "send_", "_send_", "deliver_", "_deliver",
        "autostart", "launchd", "systemd", "daemon",
        "folder_monitor", "monitor_folder", "folder_watch", "watch_folder",
        "file_watcher", "setup_monitor",
    )
    text_markers = (
        "фоновый демон", "демон мониторинга", "автозапуск", "launchagent", "launchd",
        "systemd", "background daemon", "scheduled trigger", "watch folder",
        "отслеживает появление новых файлов",
    )
    return not any(m in name for m in name_markers) and not any(m in text for m in text_markers)


def _pipeline_topology(tasks):
    """Проверяет, можно ли честно исполнить план нынешним линейным оркестратором.

    Планировщик уже возвращает depends_on, но исторический исполнитель их игнорировал и всегда
    делал t1 -> t2 -> t3. Явное ветвление или объединение веток поэтому нельзя отправлять в
    кодоген: компоненты могут по одному позеленеть, но целого процесса из них не получится.
    """
    ids = [str(t.get("id") or "t%d" % (i + 1)) for i, t in enumerate(tasks or [])]
    known = set(ids)
    deps = {}
    children = {tid: [] for tid in ids}
    for i, t in enumerate(tasks or []):
        tid = ids[i]
        ds = [str(d) for d in (t.get("depends_on") or []) if str(d) in known and str(d) != tid]
        deps[tid] = list(dict.fromkeys(ds))
        for dep in deps[tid]:
            children.setdefault(dep, []).append(tid)
    joins = [tid for tid, ds in deps.items() if len(ds) > 1]
    branches = [tid for tid, ch in children.items() if len(ch) > 1]
    return {
        "supported": not joins and not branches,
        "joins": joins,
        "branches": branches,
        "dependencies": deps,
    }


def _build_one(expert_name, task, schema_hint, is_first, is_last, accept_input, llm):
    """Стройка СТАДИИ по контракту input_path->output_path. Keyless-путь: модель строит НАТИВНО
    (create-действием, исходник не в чат — уважает guard fine-tune), харнесс НЕЗАВИСИМО перечитывает
    созданное, усыновляет в global и ПРИНИМАЕТ прогоном на реальном входе. With-key путь — как было
    (модель отдаёт код текстом). Источник правды — get+прогон, не слово модели. Возвращает (ok, output_path, detail)."""
    import urllib.request as _u
    cspl = task.get("cspl", "fython")
    if is_first:
        input_desc = ("input_path — путь к ИСХОДНОМУ файлу данных клиента (xlsx/csv). Распарси его "
                      "(openpyxl для xlsx). ВАЖНО: строка заголовков даёт НАЗВАНИЯ колонок; ДАННЫЕ начинаются "
                      "со строки СРАЗУ ПОСЛЕ заголовков. Саму строку заголовков в записи НЕ включай. "
                      "Для КАЖДОЙ непустой строки данных собери словарь {название_колонки: ЗНАЧЕНИЕ ЯЧЕЙКИ (.value)}. "
                      "Значение — число/текст/дата, НЕ номер строки и НЕ индекс колонки. Пропускай пустые строки и "
                      "строки, повторяющие заголовки. Пиши json.dump(..., ensure_ascii=False, default=str)." + schema_hint)
        out_desc = "Запиши НОРМАЛИЗОВАННЫЙ список записей (list of dict со ЗНАЧЕНИЯМИ ячеек) как JSON."
    else:
        input_desc = ("input_path — путь к JSON-файлу от предыдущей стадии (список записей или "
                      "{\"records\":[...],\"summary\":{...}}). Прочитай json.load, работай с записями.")
        out_desc = ("ОБЯЗАТЕЛЬНО ВЫЧИСЛИ агрегаты из входных записей (НЕ копируй записи без обработки!): "
                    "определи числовую колонку ИТОГОВОЙ суммы — ПРЕДПОЧИТАЙ колонку, где в названии есть "
                    "'сумма'/'итог'/'стоимость'/'total'/'amount' (НЕ бери 'цена'/'price'/'цена за единицу', "
                    "если есть колонка итоговой суммы), и посчитай "
                    "total_count (число записей), total_sum (сумма по ней); построй разбивки — словари "
                    "{значение: сумма} по каждой НЕчисловой категориальной колонке (напр. Категория, Способ закупки). "
                    "Запиши JSON {\"summary\": {\"total_count\": N, \"total_sum\": X, \"by_<колонка>\": {...}, ...}, "
                    "\"records\": [...]}. "
                    + ("Это ФИНАЛЬНАЯ стадия — дополнительно собери человекочитаемый отчёт (.md рядом с output_path) "
                       "из summary." if is_last else ""))
    sysmsg = _BUILD_SYS % {"INPUT_DESC": input_desc, "PURPOSE": str(task.get("purpose", "обработай данные")),
                           "OUTPUT_DESC": out_desc}
    user = ("Имя эксперта (СТРОГО): " + expert_name + "\nCSPL: " + cspl +
            "\nНазначение: " + str(task.get("purpose", "")) +
            "\nСгенерируй код стадии строго по контракту (input_path, output_path).")
    if isinstance((llm or {}).get("task_context"), dict):
        user += ("\n\nЕДИНЫЙ АВТОРИТЕТНЫЙ КОНТЕКСТ (учти Task Contract, Source Model и память; "
                 "не угадывай схему по позиции листа):\n" +
                 json.dumps(llm["task_context"], ensure_ascii=False, default=str))
    out_path = "/tmp/stage_" + expert_name + ".json"
    last_err = None
    why = None   # #11: инициализация ДО цикла — иначе при провале всех попыток на этапе LLM/build (до приёмки, стр.243) финальный return читает несвязанную why → NameError
    # директива нативной стройки: модель СОЗДАЁТ эксперта действием, исходник — только в create, не в чат
    build_directive = ("\n\nПОСТРОЙ этого эксперта НА ПЛАТФОРМЕ действием (создай ИЛИ обнови эксперта) под именем "
                       "СТРОГО '" + expert_name + "', cspl=" + cspl + ", сигнатура РОВНО "
                       "def " + expert_name + "(input_path=\"\", output_path=\"\"). Исходный код помещай ТОЛЬКО в "
                       "действие создания эксперта — НЕ печатай исходник в чат. РОВНО одна top-level функция, хелперы внутри.")
    for attempt in range(3):
        code = ""
        _descr = ""
        if llm.get("api_key"):
            # managed-build по внешнему ключу (без guard): модель отдаёт код текстом, харнесс сохраняет
            try:
                rq = _u.Request(llm["base_url"].rstrip("/") + "/chat/completions",
                                data=json.dumps({"model": llm["model"], "temperature": 0,
                                                 "response_format": {"type": "json_object"},
                                                 "messages": [{"role": "system", "content": sysmsg},
                                                              {"role": "user", "content": user}],
                                                 "max_tokens": 3500}).encode(),
                                headers={"Authorization": "Bearer " + llm["api_key"], "Content-Type": "application/json"},
                                method="POST")
                with _u.urlopen(rq, timeout=150) as r:
                    _content = json.loads(r.read().decode())["choices"][0]["message"]["content"]
                spec = json.loads(_content)
                code = spec.get("code", "")
                _descr = str(spec.get("description", ""))[:900]
            except Exception as e:
                last_err = "LLM: " + str(e)[:150]; time.sleep(2 + attempt * 4); continue
        else:
            # НАТИВНАЯ стройка на нашей модели: модель СОЗДАЁТ эксперта действием (исходник не в чат,
            # guard fine-tune уважается), store:true — иначе нативные действия не исполняются.
            # Затем харнесс НЕЗАВИСИМО перечитывает созданное В СКОУПЕ МОДЕЛИ (X-Agent-Id = строитель).
            _b = llm.get("agent_id") or qwen_agent()
            _tok = llm.get("api_token", CONFIG["auth_token"])
            _base = llm.get("api_base", "https://api.extella.ai").rstrip("/")
            _hdr = {"X-Auth-Token": _tok, "Content-Type": "application/json",
                    "X-Profile-Id": "default", "X-Agent-Id": _b or "agent_extella_default"}
            try:
                rq = _u.Request(_base + "/api/agent/run",
                                data=json.dumps({"agent_id": _b, "input": sysmsg + "\n\n" + user + build_directive,
                                                 "run_timeout": 300, "store": True}).encode(),
                                headers=_hdr, method="POST")
                with _u.urlopen(rq, timeout=330) as r:
                    r.read()
            except Exception as e:
                last_err = "build-run: " + str(e)[:150]; time.sleep(3 + attempt * 4); continue
            try:  # независимый get созданного (не слово модели — реальная запись)
                gq = _u.Request(_base + "/api/expert/get", data=json.dumps({"name": expert_name}).encode(),
                                headers=_hdr, method="POST")
                with _u.urlopen(gq, timeout=60) as r:
                    g = json.loads(r.read().decode())
            except Exception as e:
                g = {}; last_err = "get after build: " + str(e)[:150]
            code = (g.get("expert_code") or "") if isinstance(g, dict) else ""
            _descr = str((g.get("description") if isinstance(g, dict) else "") or "")[:900]
        # общая валидация: РОВНО одна top-level функция, имя строго expert_name
        code = re.sub(r"(?ms)^def\s+(?:load_module|include)\s*\(.*?(?=^\S|\Z)", "", code)
        tops = re.findall(r"^def\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", code, flags=re.M)
        if len(tops) != 1:
            user += ("\n\nОШИБКА: нужна РОВНО ОДНА top-level функция def " + expert_name +
                     "(input_path,output_path), нашёл " + str(tops) + " (хелперы внутри); построй заново.")
            time.sleep(1); continue
        if tops[0] != expert_name:
            code = re.sub(r"^def\s+" + re.escape(tops[0]) + r"\s*\(", "def " + expert_name + "(", code, count=1, flags=re.M)
        # description ОБЯЗАН быть непустым: /api/expert/save валидирует min-длину (string_too_short),
        # а нативный create кладёт эксперта без description → get отдаёт пусто → save падал.
        if not _descr.strip():
            _descr = str(task.get("purpose", "")).strip()[:400]
        if len(_descr.strip()) < 8:
            _descr = "Стадия обработки данных: " + expert_name
        # усыновляем в GLOBAL: оркестратор ждёт global, а нативный create кладёт в скоуп модели
        sv = api("/api/expert/save", {"name": expert_name, "description": _descr,
                                      "code": code, "kwargs": {"input_path": "", "output_path": ""},
                                      "cspl": cspl, "global": True})
        if not (isinstance(sv, dict) and (sv.get("status") == "success" or sv.get("id"))):   # #25: приёмка сохранения только по явному успеху (раньше пропускался мусор без status/id/«error»)
            return False, None, "save: " + str(sv)[:150], None
        # приёмка = реальный прогон стадии на фактическом входе (это же и звено среза)
        if Path(out_path).exists():
            try: Path(out_path).unlink()
            except Exception: pass
        run_out = run_expert(expert_name, {"input_path": accept_input, "output_path": out_path}, wait=300)
        # СТРОГАЯ приёмка: выход должен быть валидным JSON с непустыми данными (не просто «файл есть»)
        why = None
        if not Path(out_path).exists() or Path(out_path).stat().st_size == 0:
            why = "output_path не создан/пуст; run=" + str(run_out)[:180]
        else:
            try:
                data = json.loads(Path(out_path).read_text(encoding="utf-8"))
            except Exception as e:
                why = "выход не валидный JSON (" + str(e)[:80] + ") — пиши json.dump(ensure_ascii=False, default=str)"
            else:
                recs = data if isinstance(data, list) else (data.get("records") or data.get("rows") or data.get("items") or [])
                if is_first:
                    if not isinstance(recs, list) or len(recs) == 0 or not isinstance(recs[0], dict):
                        why = "первая стадия должна вернуть НЕПУСТОЙ список словарей-записей"
                    else:
                        def _is_headerish(rec):
                            vals = list(rec.values())
                            if vals and all(isinstance(v, int) for v in vals) and (max(vals) - min(vals)) <= len(vals) + 2:
                                return True  # значения-индексы колонок
                            return sum(1 for k, v in rec.items() if str(v).strip() == str(k).strip()) >= max(2, len(rec) // 2)
                        # прагматично: САМИ чистим строки-заголовки/индексы из вывода, не заваливаем сборку
                        cleaned = [r for r in recs if isinstance(r, dict) and not _is_headerish(r)]
                        if not cleaned:
                            why = "после отсева заголовков не осталось записей-данных — парсер не извлёк реальные строки"
                        elif len(cleaned) != len(recs):
                            Path(out_path).write_text(json.dumps(cleaned, ensure_ascii=False, default=str), encoding="utf-8")
                else:
                    summ = data.get("summary") if isinstance(data, dict) else None
                    has_num = isinstance(summ, dict) and any(
                        isinstance(v, (int, float)) or (isinstance(v, dict) and any(isinstance(x, (int, float)) for x in v.values()))
                        for v in summ.values())
                    if not has_num:
                        why = ("стадия обязана ВЫЧИСЛИТЬ summary с числами (total_count/total_sum/by_<колонка>), "
                               "а не копировать записи; сейчас summary пуст или без чисел")
        if why is None:
            # Структурно принято → независимая смысловая проверка. FAIL возвращается в ремонт этой
            # же стадии; частичный эксперт не проходит в оркестратор как «требует доводки».
            _san = _stage_sanity(task.get("title") or expert_name, task.get("purpose") or "", out_path, llm)
            if not _san.get("ok"):
                why = _san.get("why") or "смысловая приёмка не пройдена"
                user += "\n\nСМЫСЛОВАЯ ПРИЁМКА УПАЛА: " + str(why) + ". Исправь первопричину по Task Contract."
                time.sleep(1)
                continue
            return True, out_path, "built+accepted", None
        user += "\n\nПРИЁМКА УПАЛА (вход " + str(accept_input) + "): " + why + ". Исправь под контракт."
    return False, None, "3 попытки: " + str(why or last_err or "не прошли приёмку")[:150], None


# ─────────── AC-06: СТРУКТУРНЫЕ ОШИБКИ СТРОЙКИ ───────────
# Сбой стройки должен отвечать владельцу на три вопроса: что не получилось, почему и что делать.
# Свободный текст на это не отвечает — и не даёт UI/агенту действовать по коду. Каталог кодов ниже
# = контракт: код + человеческая формулировка + КОНКРЕТНОЕ действие (remedy).
BUILD_ERRORS = {
    "no_sample_file": ("Стадии нечего обрабатывать: не приложен файл-образец",
                       "Приложите типовую выгрузку (2–3 строки достаточно) и запустите стройку заново — "
                       "по образцу собирается и проверяется каждая стадия."),
    "plan_failed": ("Не удалось составить план стройки",
                    "Уточните задачу словами в описании процесса: чем конкретнее цель и что на входе, "
                    "тем точнее план. Затем повторите стройку."),
    "plan_transport_failed": ("Связь прервалась во время составления плана",
                              "Интервью и сессия сохранены. Когда интернет станет стабильнее, нажмите "
                              "«Повторить сборку» — начинать заново и переписывать задачу не нужно."),
    "plan_not_saved": ("План стройки не сохранился",
                        "Похоже, платформа не приняла запись. Повторите стройку через минуту; "
                        "если повторится — это вопрос к платформе, а не к описанию процесса."),
    "unsupported_topology": ("Этот процесс требует нескольких входов или параллельных веток",
                             "Ничего переформулировать и повторно запускать не нужно. Это ограничение "
                             "текущего Строителя: он умеет проверять только один последовательный поток. "
                             "Кейс нужно передать разработчикам как задачу поддержки ветвления."),
    "stage_failed": ("Один из шагов не прошёл сборку и проверку",
                     "Следующие шаги не запускались, оркестратор не создавался, проверка допуска не "
                     "проводилась. Причина упавшего шага указана ниже; это задача диагностики сборки, "
                     "а не повод менять описание процесса наугад."),
    "agentic_acceptance_failed": ("Целостное решение не прошло приёмку на ваших данных",
                                  "Qwen уже получила фактические ошибки и пыталась исправить решение. "
                                  "Ни частичный эксперт, ни ложный результат не опубликованы. Если ниже "
                                  "есть вопрос владельцу — ответьте на него в интервью и повторите стройку; "
                                  "иначе это задача диагностики Строителя."),
    "agentic_run_failed": ("Draft-эксперт не прошёл техническую проверку на образцах",
                           "Результат не опубликован. История запуска и последний неприменённый урок "
                           "сохранены для следующей ручной стройки; это задача диагностики Строителя."),
    "agentic_build_failed": ("Qwen не удалось создать исполняемый draft-эксперт",
                             "Stable-версия не изменена. Ошибки генерации и сохранения перечислены ниже "
                             "и сохранены в журнале ремонта."),
    "agentic_stalled": ("Ремонт остановлен: код не менялся после той же ошибки",
                        "Неизменившийся код не засчитан как ремонт. Rejected-урок сохранён в сессии; "
                        "stable-эксперт не изменён."),
    "agentic_timeout": ("Сборка достигла общего лимита времени",
                        "Draft и история попыток сохранены в журнале, stable-версия не изменена. "
                        "Повторный ручной запуск продолжит с памятью этой сессии."),
    "source_model_failed": ("Не удалось доказанно понять роли входных данных",
                            "Source Model не прошла проверку фактических файлов и разделов. "
                            "Результат не выдуман и draft не опубликован; это задача диагностики Строителя."),
    "needs_owner_input": ("Для продолжения нужен один бизнес-ответ владельца",
                          "Ответьте на точный вопрос ниже в интервью и запустите стройку снова. "
                          "До ответа процесс завершён и ничего не публикует."),
    "capability_missing": ("Для задачи не хватает доступной проверенной способности",
                           "Кандидат можно рассмотреть отдельно, но Wizard ничего не устанавливал и "
                           "не записывал его в постоянный мозг."),
    "agent_publish_failed": ("Проверенный draft не удалось опубликовать как stable-эксперт",
                             "Рабочая stable-версия не была перезаписана. Повторите позже; если сбой "
                             "повторится, нужна диагностика API Extella."),
    "no_components_built": ("Ни один компонент не собрался",
                            "Чаще всего причина одна — нет файла-образца или он не читается. "
                            "Проверьте, что файл открывается и в нём есть заголовки колонок."),
    "stages_missing": ("Собрались не все шаги процесса",
                       "Незакрытые шаги перечислены ниже. Обычно помогает уточнить, что именно должен "
                       "делать этот шаг, и пересобрать процесс."),
    "slice_failed": ("Проверка на реальных данных не прошла",
                     "Процесс собран, но на Вашем файле не отработал. Проверьте образец: "
                     "совпадают ли колонки с тем, что описано в задаче."),
    "orchestrator_failed": ("Шаги собраны, но процесс не связался в цепочку",
                            "Это внутренняя ошибка сборки. Повторите стройку; если повторится — "
                            "нужна наша диагностика, данные не пострадали."),
    "crashed": ("Стройка прервалась",
                "Ничего необратимого не произошло. Повторите стройку; если ошибка повторяется — "
                "пришлите нам этот экран."),
}


def _process_manifest(ns, orchestrator, data_tasks, built_ok, sample_file, slice_summary,
                      kp_pack=None, build_id="", session_id="", skip_ids=None):
    """AC-06 МАНИФЕСТ: собранный процесс должен САМ СЕБЯ описывать — что ест, что отдаёт, из чего
    состоит и какие контракты держит. Без этого процесс существует только как код: его нельзя
    честно переиспользовать у другого клиента, положить в пак или проверить на совместимость.
    Манифест декларативен и строится из ФАКТОВ стройки, а не из обещаний плана."""
    inputs = []
    try:
        if sample_file and str(sample_file).lower().endswith((".xlsx", ".xlsm")):
            import openpyxl as _ox
            _ws = _ox.load_workbook(sample_file, read_only=True).active
            inputs = [str(c.value).strip() for c in next(_ws.iter_rows(min_row=1, max_row=1))
                      if c.value is not None and str(c.value).strip()]
    except Exception:
        inputs = []
    outs = sorted(slice_summary.keys()) if isinstance(slice_summary, dict) else []
    steps = []
    skip_ids = skip_ids or set()
    for i, tk in enumerate(data_tasks or []):
        nm = tk.get("expert_name") or (ns + "_" + tk.get("id", "t%d" % (i + 1)))
        if tk.get("id", "t%d" % (i + 1)) in skip_ids:   # намеренно не вошёл в каркас (искать не по чему)
            steps.append({"name": nm, "title": _human_title(tk, ns), "mode": "skipped", "ok": True})
            continue
        steps.append({"name": nm, "title": _human_title(tk, ns),
                      "mode": "reuse" if str(tk.get("action", "build")).lower() == "reuse" else "built",
                      "ok": nm in (built_ok or [])})
    return {
        "manifest_version": 1,
        "process": ns, "orchestrator": orchestrator, "build_id": build_id, "session_id": session_id,
        "built_at": datetime.now(timezone.utc).isoformat(),
        "input": {"kind": "file", "basename": Path(sample_file).name if sample_file else "",
                  "fields": inputs},
        "output": {"summary_keys": outs, "reports": ["md", "xlsx"]},
        "steps": steps,
        "knowledge_pack": kp_pack or None,
        # контракты = что этот процесс УМЕЕТ принимать. По ним мост решает, что можно ему передать,
        # не гадая по дате сборки (старым оркестраторам лишний kwarg = падение).
        "contracts": {"params": 1, "placement": 1, "adapter": 1},
    }


def _build_error(code, detail="", **extra):
    """Структурная ошибка стройки: код + что случилось + что делать. detail — фактура (какие шаги и т.п.)."""
    msg, remedy = BUILD_ERRORS.get(code, ("Стройка не завершилась", "Повторите стройку."))
    out = {"code": code, "message": msg, "remedy": remedy}
    if detail:
        out["detail"] = str(detail)[:300]
    out.update(extra)
    return out


_ORCH_TEMPLATE = '''$extens("include.py")
include("import requests", ["extella-pip install requests"])
include("import openpyxl", ["extella-pip install openpyxl"])
include("from cryptography.fernet import Fernet", ["extella-pip install cryptography"])

def %(NAME)s(source_file: str = "", work_dir: str = "%(WORKDIR)s", api_token: str = "", api_base: str = "https://api.extella.ai", target: str = "", source_key: str = "", rules_json: str = "", fields_json: str = "", run_id: str = "", placement_json: str = "", adapter_json: str = "", report_spec_json: str = "") -> dict:
    """Автосгенерированный оркестратор процесса. Гоняет контрактную цепочку стадий
    (input_path -> output_path) на исходном файле, чистит заголовки, возвращает сводку
    и рисует отчёт .md + .xlsx. Параметры: source_file, work_dir, api_token, target
    (пиннинг устройства для стадий), source_key (ключ файла в общем сторе для резолвера)."""
    import json
    import requests
    from pathlib import Path
    from datetime import datetime, timezone

    STAGES = %(STAGES)s
    KP_STAGES = %(KP_STAGES)s
    WEB_STAGES = %(WEB_STAGES)s   # Qwen-ведомые шаги веб-обогащения — им нужен api_token (agent_id вшит)
    # A1 КАРТА РАЗМЕЩЕНИЯ: {стадия: target}. Процесс больше не живёт целиком на одном устройстве —
    # чтение 1С на машине с 1С, отчёт и доставка на хостинге. Пусто = старое поведение (один target).
    PLACEMENT = {}
    if placement_json and not placement_json.startswith("{{"):
        try:
            PLACEMENT = json.loads(placement_json) or {}
        except Exception:
            PLACEMENT = {}
    if not api_token:
        cfg = Path.home() / "extella_wizard" / "app" / "config.json"
        try:
            api_token = json.loads(cfg.read_text(encoding="utf-8")).get("auth_token", "") if cfg.exists() else ""
        except Exception:
            api_token = ""
    if not api_token:
        return {"status": "error", "message": "api_token не передан и нет bridge-конфига"}
    if not source_file:
        return {"status": "error", "message": "source_file обязателен"}
    wd = Path(work_dir); wd.mkdir(parents=True, exist_ok=True)
    headers = {"X-Auth-Token": api_token, "Content-Type": "application/json",
               "X-Profile-Id": "default", "X-Agent-Id": "agent_extella_default"}

    # резолвер источника: если локального пути нет (процесс исполняется на хостинге) —
    # материализуем файл из общего стора (KV) чанками base64 в рабочую папку
    _resolve_err = ""
    if source_file and not Path(source_file).exists():
        import base64 as _b64, hashlib as _hl
        _bn = Path(source_file).name
        _base = source_key or ("file:%(SID)s:" + _hl.md5(_bn.encode("utf-8")).hexdigest()[:12])
        try:
            _mr = requests.post(api_base.rstrip("/") + "/api/kv/get", headers=headers, json={"key": _base + ":meta"}, timeout=120).json()
            _mv = _mr.get("value")
            if _mv:
                _meta = json.loads(_mv); _buf = ""
                for _i in range(int(_meta.get("chunks", 0))):
                    _cr = requests.post(api_base.rstrip("/") + "/api/kv/get", headers=headers, json={"key": _base + ":" + str(_i)}, timeout=120).json()
                    _buf += _cr.get("value") or ""
                if _buf:
                    if _meta.get("enc"):
                        # файл зашифрован — расшифровываем ЛОКАЛЬНЫМ vault.key устройства (не из KV)
                        _kc = [Path("/opt/extella-listener/extella_wizard/vault.key"), Path.home() / "extella_wizard/app/vault.key", Path.cwd() / "extella_wizard/vault.key"]
                        _kp = next((c for c in _kc if c.exists()), None)
                        if not _kp:
                            _resolve_err = "зашифрованный источник: локальный vault.key не найден на устройстве-хостинге"
                            _rawf = None
                        else:
                            _rawf = Fernet(_kp.read_bytes()).decrypt(_buf.encode())
                    else:
                        _rawf = _b64.b64decode(_buf)
                    if _rawf:
                        _tmp = wd / _bn
                        _tmp.write_bytes(_rawf)
                        source_file = str(_tmp)
        except Exception as _e:
            _resolve_err = _resolve_err or ("не удалось восстановить источник из стора: " + str(_e)[:120])
    # честный fail вместо тихой пропажи файла (напр. enc без ключа / битые чанки)
    if source_file and not Path(source_file).exists():
        return {"status": "error", "message": _resolve_err or ("источник не найден на устройстве и не восстановлен из стора: " + str(source_file))}

    def is_headerish(rec):
        if not isinstance(rec, dict): return True
        vals = list(rec.values())
        if vals and all(isinstance(v, int) for v in vals) and (max(vals) - min(vals)) <= len(vals) + 2: return True
        return sum(1 for k, v in rec.items() if str(v).strip() == str(k).strip()) >= max(2, len(rec) // 2)

    def clean_file(path):
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception:
            return
        if isinstance(data, list):
            cl = [r for r in data if not is_headerish(r)]
            if cl and len(cl) != len(data):
                Path(path).write_text(json.dumps(cl, ensure_ascii=False, default=str), encoding="utf-8")

    # AC-05 АДАПТЕР: клиент поменял выгрузку (переименовал/переставил колонки) — процесс не должен
    # ломаться и, что хуже, считать по чужой схеме. Адаптер — явный именованный маппинг
    # «колонка выгрузки → поле процесса», подтверждённый человеком. Применяем к ЗАГОЛОВКУ исходного
    # файла до первой стадии: дальше весь процесс видит привычные ему поля.
    _adapt = {}
    if adapter_json and not adapter_json.startswith("{{"):
        try:
            _adapt = (json.loads(adapter_json) or {}).get("map") or {}
        except Exception:
            _adapt = {}
    _adapt_applied = []
    if _adapt and str(source_file).lower().endswith((".xlsx", ".xlsm")):
        try:
            import openpyxl
            _wb = openpyxl.load_workbook(source_file)
            _ws = _wb.active
            for _c in next(_ws.iter_rows(min_row=1, max_row=1)):
                _old = str(_c.value or "").strip()
                if _old in _adapt and _adapt[_old] and _adapt[_old] != _old:
                    _c.value = _adapt[_old]
                    _adapt_applied.append(_old + " → " + _adapt[_old])
            if _adapt_applied:
                _wb.save(source_file)
        except Exception as _ae:
            return {"status": "error", "message": "адаптер источника не применился: " + str(_ae)[:150]}

    prev, last_out = source_file, None
    # A2 ЧЕКПОИНТЫ: длинный процесс не переигрывается с нуля. Чекпоинт привязан к run_id ЭТОГО прогона —
    # stage-файлы ЧУЖОГО/прошлого прогона не реюзаются НИКОГДА (урок data-integrity: упавшая стадия при
    # живом старом файле давала вчерашний отчёт как сегодняшний).
    import uuid as _uuid
    _ckpt_p = wd / "checkpoint.json"
    _resume_from = 0
    if not run_id or run_id.startswith("{{"):
        run_id = _uuid.uuid4().hex[:12]
        _ck = {"run_id": run_id, "done": []}
    else:
        try:
            _ck = json.loads(_ckpt_p.read_text(encoding="utf-8"))
        except Exception:
            _ck = {}
        if _ck.get("run_id") != run_id:
            _ck = {"run_id": run_id, "done": []}   # чекпоинт от другого прогона — начинаем с нуля
        else:
            _done = []
            for _d in (_ck.get("done") or []):     # доверяем только подряд идущим стадиям с живым файлом
                _op = Path(_d.get("out", ""))
                if _op.exists() and _op.stat().st_size > 0:
                    _done.append(_d)
                else:
                    break
            _ck["done"] = _done
            _resume_from = len(_done)
            if _done:
                prev = last_out = _done[-1]["out"]
    for i, name in enumerate(STAGES):
        outp = str(wd / ("stage%%d.json" %% i))
        if i < _resume_from:
            continue   # стадия уже успешно отработала В ЭТОМ прогоне — не переигрываем
        try:
            Path(outp).unlink()   # не тащим stage-файл прошлого прогона: упавшая стадия иначе «успешна» на чужих данных
        except OSError:
            pass
        _params = {"input_path": prev, "output_path": outp}
        # F2 (контракт параметров): правила/поля владельца доступны КАЖДОЙ стадии этой сборки
        # (кодоген генерит стадии с этими kwargs; менять поведение обязана только та, кому релевантно)
        if not (rules_json or "").startswith("{{") and rules_json:
            _params["rules_json"] = rules_json
        if not (fields_json or "").startswith("{{") and fields_json:
            _params["fields_json"] = fields_json
        _tgt = PLACEMENT.get(name) or target   # A1: устройство ЭТОЙ стадии
        if name in KP_STAGES:   # knowledge-стадия реюзает kp_ask — нужен target (где лежит база) и токен
            _params["target"] = _tgt
            _params["api_token"] = api_token
        if name in WEB_STAGES:   # веб-обогащение зовёт платформенную Qwen — нужен токен (agent_id вшит в стадию)
            _params["api_token"] = api_token
            _params["target"] = _tgt
        body = {"expert_name": name, "params": _params, "global": True}
        if _tgt:
            body["target"] = _tgt
        r = requests.post(api_base.rstrip("/") + "/api/expert/run", headers=headers, json=body, timeout=600)
        try:
            res = r.json().get("result", r.json())
        except Exception:
            res = {}
        ok = (isinstance(res, dict) and res.get("status") == "success") or (Path(outp).exists() and Path(outp).stat().st_size > 0)
        if ok and i == 0:
            clean_file(outp)
        if not ok or not Path(outp).exists():   # ok теперь честный (файл прошлого прогона удалён) — провал стадии не маскируется
            return {"status": "error", "failed_stage": name, "detail": str(res)[:200],
                    "run_id": run_id, "resumable": True, "done_stages": _resume_from}
        prev, last_out = outp, outp
        _ck["done"].append({"i": i, "name": name, "out": outp})   # A2: стадия зачтена в ЭТОМ прогоне
        try:
            _ckpt_p.write_text(json.dumps(_ck, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    summary = {}
    try:
        data = json.loads(Path(last_out).read_text(encoding="utf-8"))
        summary = data.get("summary", {}) if isinstance(data, dict) else {}
    except Exception:
        pass

    # F2: структурные правила владельца ({"field","op","value"} внутри rules_json) применяются
    # ЗДЕСЬ детерминированно — фильтр записей последней стадии + честный пересчёт сводки.
    # Текстовые правила из того же списка — дело кодогенных стадий; строки просто пропускаем.
    _structs = []
    try:
        if rules_json and not rules_json.startswith("{{"):
            _structs = [r for r in json.loads(rules_json)
                        if isinstance(r, dict) and r.get("field") and r.get("op")]
    except Exception:
        _structs = []
    if _structs:
        try:
            _fd = json.loads(Path(last_out).read_text(encoding="utf-8"))
            _recs = _fd if isinstance(_fd, list) else (_fd.get("records") if isinstance(_fd, dict) else None)
            if isinstance(_recs, list) and _recs and isinstance(_recs[0], dict):
                def _fkey(rec, fld):
                    fl = str(fld).casefold().strip()
                    for k in rec.keys():
                        if fl == str(k).casefold().strip() or fl in str(k).casefold():
                            return k
                    return None
                def _fnum(v):
                    try:
                        return float(str(v).replace(" ", "").replace(",", "."))
                    except Exception:
                        return None
                def _passes(rec):
                    for ru in _structs:
                        k = _fkey(rec, ru.get("field"))
                        if k is None:
                            continue   # поля нет в записи — правило её не душит (не наш разрез)
                        op, val = str(ru.get("op")), ru.get("value")
                        if op == "contains":
                            if str(val).casefold() not in str(rec.get(k, "")).casefold():
                                return False
                            continue
                        a, b = _fnum(rec.get(k)), _fnum(val)
                        if a is None or b is None:
                            continue
                        if (op == ">" and not a > b) or (op == ">=" and not a >= b) or \
                           (op == "<" and not a < b) or (op == "<=" and not a <= b) or \
                           (op in ("==", "=") and not a == b):
                            return False
                    return True
                _flt = [r for r in _recs if _passes(r)]
                if len(_flt) != len(_recs):
                    _new = {"total_count": len(_flt), "filtered_by_rules": len(_recs) - len(_flt)}
                    # колонка суммы: по имени + в ней реально есть числа (анонимизация могла замаскировать в ***)
                    _sumcol = next((c for c in _recs[0].keys()
                                    if ("сумм" in str(c).casefold() or "amount" in str(c).casefold() or "итог" in str(c).casefold())
                                    and any(_fnum(r.get(c)) is not None for r in _flt)), None)
                    if _sumcol:
                        _new["total_sum"] = sum((_fnum(r.get(_sumcol)) or 0) for r in _flt)
                    for k in list(summary.keys()):
                        if str(k).startswith("by_"):
                            _cnt = {}
                            for r in _flt:
                                ck = _fkey(r, str(k)[3:])
                                if ck is not None:
                                    v = str(r.get(ck))
                                    _cnt[v] = _cnt.get(v, 0) + 1
                            _new[k] = _cnt
                    summary = _new
        except Exception:
            pass   # применение правил не должно ронять прогон — сводка останется полной

    try:
        (wd / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, default=str), encoding="utf-8")
    except Exception:
        pass

    # детерминированный рендер отчёта .md + .xlsx из summary
    md = wd / "report.md"; xlsx = wd / "report.xlsx"
    lines = ["# Сводка процесса", "", "_Сформировано: " + datetime.now(timezone.utc).isoformat()[:16].replace("T", " ") + " UTC_", ""]
    tc = summary.get("total_count"); ts = summary.get("total_sum")
    if tc is not None: lines.append("- Всего позиций: **%%s**" %% tc)
    if ts is not None: lines.append("- Общая сумма: **%%s**" %% format(ts, ",").replace(",", " "))
    for k, v in summary.items():
        if k.startswith("by_") and isinstance(v, dict) and v:
            lines += ["", "## " + k[3:], "", "| Значение | Сумма |", "|---|---|"]
            for kk, vv in list(v.items())[:20]:
                lines.append("| %%s | %%s |" %% (kk, vv))
    md.write_text("\\n".join(lines), encoding="utf-8")
    try:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Сводка"
        ws.append(["Показатель", "Значение"])
        if tc is not None: ws.append(["Всего позиций", tc])
        if ts is not None: ws.append(["Общая сумма", ts])
        for k, v in summary.items():
            if k.startswith("by_") and isinstance(v, dict) and v:
                ws.append([]); ws.append([k[3:], ""])
                for kk, vv in list(v.items())[:50]:
                    ws.append([kk, vv])
        wb.save(str(xlsx))
    except Exception:
        pass

    # WZ-B02 (ТЗ v2 §17): пустой результат при непустом входе — НЕ success.
    # Легитимное исключение: записи отсеяны правилами владельца (filtered_by_rules).
    _in_count = None
    try:
        _d0 = json.loads((wd / "stage0.json").read_text(encoding="utf-8"))
        _r0 = _d0 if isinstance(_d0, list) else (_d0.get("records") if isinstance(_d0, dict) else None)
        _in_count = len(_r0) if isinstance(_r0, list) else None
    except Exception:
        pass
    _status = "success"
    _reason = ""
    if _in_count and not tc and not (isinstance(summary, dict) and summary.get("filtered_by_rules")):
        _status = "needs_review"
        _reason = ("вход содержит " + str(_in_count) +
                   " записей, а итоговая сводка пуста — процесс потерял данные, проверьте стадии")

    # ── ФИРМЕННЫЙ PDF: оформитель — отдельный эксперт (способность зашита в него, не в оркестратор).
    # Идёт на ТО ЖЕ устройство, что и стадии, и читает файл последней стадии по месту —
    # так не упираемся ни в размер полезной нагрузки, ни в наличие браузера.
    _pdf, _docx, _pptx = "", "", ""
    if report_spec_json and not report_spec_json.startswith("{{") and last_out:
        # формат выбирает владелец словами: PDF отправляют как есть, DOCX — дорабатывают
        try:
            _fmt = str((json.loads(report_spec_json) or {}).get("format") or "pdf").lower()
        except Exception:
            _fmt = "pdf"
        _jobs = []
        if _fmt in ("pdf", "both", "all"):
            _jobs.append(("fmt_report_pdf", str(wd / "report.pdf")))
        if _fmt in ("docx", "word", "both", "all"):
            _jobs.append(("fmt_report_docx", str(wd / "report.docx")))
        if _fmt in ("pptx", "slides", "all"):
            _jobs.append(("fmt_report_pptx", str(wd / "report.pptx")))
        for _exp, _dst in _jobs:
            try:
                _fb = {"expert_name": _exp, "global": True,
                       "params": {"input_path": last_out, "spec_json": report_spec_json,
                                  "output_path": _dst}}
                if target:
                    _fb["target"] = target
                requests.post(api_base.rstrip("/") + "/api/expert/run", headers=headers, json=_fb, timeout=180)
                if Path(_dst).exists() and Path(_dst).stat().st_size > 1000:
                    if _dst.endswith(".pdf"):
                        _pdf = _dst
                    elif _dst.endswith(".docx"):
                        _docx = _dst
                    else:
                        _pptx = _dst
            except Exception:
                pass   # .md/.xlsx уже собраны — оформленный документ не обязан ронять прогон

    result = {"status": _status, "summary": summary, "total_count": tc, "total_sum": ts,
              "report_pdf": _pdf, "report_docx": _docx, "report_pptx": _pptx,
              "adapter_applied": _adapt_applied,   # AC-05: видно в прогоне, что выгрузку подстроили под процесс
              "report_md": str(md), "report_xlsx": str(xlsx), "host": __import__("socket").gethostname()}
    if _reason:
        result["needs_review_reason"] = _reason
    # межустройственный слепок последнего прогона в KV — планировщик читает его,
    # т.к. вложенный прогон возвращается отложенным без task_id.
    try:
        ns = Path(work_dir).name.replace("_run", "")
        rec = {"at": datetime.now(timezone.utc).isoformat(), "status": "success",
               "total_count": tc, "total_sum": ts, "report_xlsx": str(xlsx),
               "host": __import__("socket").gethostname()}
        requests.post(api_base.rstrip("/") + "/api/kv/set", headers=headers,
                      json={"key": "lastrun:" + ns, "value": json.dumps(rec, ensure_ascii=False, default=str),
                            "description": "lastrun " + ns}, timeout=60)
    except Exception:
        pass
    return result
'''


# Детерминированная обёртка knowledge-стадии: РЕЮЗ готового kp_ask (retrieval-движок не переписываем).
# Читает вход, формирует запрос, зовёт kp_ask(pack, question) на нужном устройстве (target), кладёт
# найденные нормы в legal_context для следующей стадии. Мягкая деградация: если базы/сети нет
# (build-срез без target) — проводит вход дальше с пустым контекстом, не роняя пайплайн.
_KP_STAGE_TEMPLATE = '''$extens("include.py")
include("import urllib.request", [])

def %(NAME)s(input_path="", output_path="", target="", api_token="", api_base="https://api.extella.ai", rules_json="", fields_json="") -> dict:   # rules/fields — контракт F2 (заглушки: kp-стадия ищет нормы)
    import json
    from pathlib import Path

    def _b(v):
        return (not v) or str(v).startswith("{{")

    if _b(api_token):
        try:
            api_token = json.loads((Path.home() / "extella_wizard" / "app" / "config.json").read_text(encoding="utf-8")).get("auth_token", "")
        except Exception:
            api_token = ""

    data = {}
    try:
        if input_path and Path(input_path).exists():
            data = json.loads(Path(input_path).read_text(encoding="utf-8"))
    except Exception:
        data = {}

    q = ""
    if isinstance(data, dict):
        q = str(data.get("request") or data.get("event_type") or data.get("question") or data.get("_query") or "")
    if not q:
        q = "Найди релевантные нормы и статьи по документу: " + json.dumps(data, ensure_ascii=False)[:400]

    ctx = ""
    if api_token:
        try:
            body = {"name": "kp_ask", "params": {"name": "%(PACK)s", "question": q}, "global": True}
            if not _b(target):
                body["target"] = target
            req = urllib.request.Request(api_base.rstrip("/") + "/api/expert/run",
                                         data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                                         headers={"X-Auth-Token": api_token, "Content-Type": "application/json",
                                                  "X-Profile-Id": "default", "X-Agent-Id": "agent_extella_default"},
                                         method="POST")
            rr = json.loads(urllib.request.urlopen(req, timeout=180).read().decode("utf-8"))
            ctx = rr.get("result") or rr.get("output") or ""
            if isinstance(ctx, (dict, list)):
                ctx = json.dumps(ctx, ensure_ascii=False)
        except Exception as e:
            ctx = "[knowledge lookup unavailable: %%s]" %% str(e)[:120]

    out = dict(data) if isinstance(data, dict) else {"input": data}
    out["legal_context"] = ctx
    out["knowledge_pack"] = "%(PACK)s"
    if output_path:
        Path(output_path).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "success", "legal_context_len": len(str(ctx)), "pack": "%(PACK)s"}
'''


def _build_kp_stage(name, pack_id):
    """РЕЮЗ kp_ask: сохраняет тонкую обёртку-стадию (не кодоген). Возвращает (name|None, sv)."""
    code = _KP_STAGE_TEMPLATE % {"NAME": name, "PACK": pack_id}
    sv = api("/api/expert/save", {"name": name,
                                  "description": "Knowledge grounding stage (reuses kp_ask on pack '" + pack_id + "'): "
                                                 "finds relevant articles and adds legal_context for the next stage.",
                                  "code": code,
                                  "kwargs": {"input_path": "", "output_path": "", "target": "",
                                             "api_token": "", "api_base": "https://api.extella.ai"},
                                  "cspl": "fython", "global": True})
    ok = sv.get("status") == "success" or sv.get("id") is not None
    return (name if ok else None), sv


# ВЕБ-ОБОГАЩЕНИЕ = QWEN-ВЕДОМЫЙ ШАГ (модель в контуре), НЕ механический код (идея Анвара 20.07:
# «если скинуть эксель в чат, Qwen сам найдёт поставщиков и не перепутает колонки»). Раньше визард
# застывал задачу в тупой перебор ячеек — и терял суждение модели: искал по суммам/заголовкам →
# мусор (Гульжан). Теперь стадия детерминирована ПО СТРУКТУРЕ (шаблон, не кодоген), а ВНУТРИ зовёт
# платформенную Qwen (web_search в каноне): модель читает строку, сама берёт название из разрешённых
# текстовых колонок и ищет — как чат, только застывшее в работника. entity_fields ВШИТЫ на сборке
# (детерминированный `_entity_columns`), agent_id — платформенная Qwen (НЕ Claude-дефолт).
_WEB_STAGE_TEMPLATE = '''$extens("include.py")
include("import urllib.request", [])

def %(NAME)s(input_path="", output_path="", target="", api_token="", api_base="https://api.extella.ai", rules_json="", fields_json="") -> dict:   # rules/fields — контракт F2 (заглушки: веб-стадия обогащает)
    import json, re
    from pathlib import Path

    def _b(v):
        return (not v) or str(v).startswith("{{")

    if _b(api_token):
        try:
            api_token = json.loads((Path.home() / "extella_wizard" / "app" / "config.json").read_text(encoding="utf-8")).get("auth_token", "")
        except Exception:
            api_token = ""
    agent_id = "%(AGENT_ID)s"
    if _b(agent_id) or agent_id == "agent_extella_default":   # никогда не платный Claude-дефолт
        agent_id = "agent_extella_alibaba_default"
    entity_fields = %(ENTITY_FIELDS)s   # разрешённые колонки-названия (вшиты на сборке)

    data = {}
    try:
        if input_path and Path(input_path).exists():
            data = json.loads(Path(input_path).read_text(encoding="utf-8"))
    except Exception:
        data = {}
    reckey = None
    if isinstance(data, list):
        recs = data
    elif isinstance(data, dict):
        for k in ("records", "rows", "items"):
            if isinstance(data.get(k), list):
                recs, reckey = data[k], k
                break
        else:
            recs = []
    else:
        recs = []

    def _entity(r):
        parts = []
        for f in entity_fields:
            v = str((r or {}).get(f, "")).strip()
            if v and re.search(r"[^\\W\\d_]", v) and not re.fullmatch(r"[\\d\\s.,%%+\\-()]+", v):
                parts.append(v)
        return " ".join(parts[:2]).strip()

    idx = [i for i, r in enumerate(recs) if isinstance(r, dict) and _entity(r)]
    enriched = 0
    # Нечего искать (нет api_token или ни одной текстовой сущности) — ЧЕСТНО проводим вход дальше,
    # не выдумывая запросы (канон: немой отказ = дефект, но и мусор выдавать нельзя).
    if api_token and idx:
        BATCH = 8
        for b in range(0, len(idx), BATCH):
            chunk = idx[b:b + BATCH]
            items = [{"i": j, "name": _entity(recs[j])} for j in chunk]
            prompt = ("Ты ищешь внешних поставщиков/предложения в интернете по позициям закупки. "
                      "Для КАЖДОЙ позиции ищи ТОЛЬКО по её полю name (это товар/поставщик) — НИКОГДА по числам, "
                      "суммам, датам или служебным полям. Верни СТРОГО JSON-массив объектов без пояснений; "
                      "у каждого объекта поле i (индекс позиции как дано) и поле suppliers — список объектов "
                      "с полями name (поставщик), url (ссылка), note (кратко). Если по позиции ничего "
                      "релевантного нет — suppliers пустой список. Позиции (JSON):\\n"
                      + json.dumps(items, ensure_ascii=False))
            try:
                body = {"agent_id": agent_id, "input": prompt, "store": False}
                req = urllib.request.Request(api_base.rstrip("/") + "/api/agent/run",
                                             data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                                             headers={"X-Auth-Token": api_token, "Content-Type": "application/json",
                                                      "X-Profile-Id": "default", "X-Agent-Id": agent_id},
                                             method="POST")
                rr = json.loads(urllib.request.urlopen(req, timeout=180).read().decode("utf-8"))
                text = ""
                for it in (rr.get("output") or []):
                    if isinstance(it, dict) and it.get("type") == "message":
                        for c in it.get("content", []):
                            if isinstance(c, dict) and c.get("type") == "output_text":
                                text += c.get("text", "")
                if not text:
                    _r = rr.get("output_text") or rr.get("result") or ""
                    text = _r if isinstance(_r, str) else json.dumps(_r, ensure_ascii=False)
                m = re.search(r"\\[.*\\]", text, re.S)
                arr = json.loads(m.group(0)) if m else []
                for e in (arr if isinstance(arr, list) else []):
                    if not isinstance(e, dict):
                        continue
                    j = e.get("i")
                    if isinstance(j, int) and 0 <= j < len(recs) and isinstance(recs[j], dict):
                        recs[j]["external_suppliers"] = e.get("suppliers") or []
                        enriched += 1
            except Exception as ex:
                for j in chunk:
                    if isinstance(recs[j], dict):
                        recs[j].setdefault("external_suppliers", [])
                        recs[j]["_enrich_note"] = "обогащение недоступно: " + str(ex)[:80]

    if reckey:
        data[reckey] = recs
        out = data
    elif isinstance(data, list):
        out = recs
    else:
        out = data
    if output_path:
        Path(output_path).write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {"status": "success", "enriched": enriched, "candidates": len(idx), "total": len(recs)}
'''


def _build_websearch_stage(name, entity_fields, agent_id):
    """Qwen-ведомый шаг веб-обогащения (детерминированный шаблон, модель в контуре). Возвращает (name|None, sv)."""
    code = _WEB_STAGE_TEMPLATE % {"NAME": name,
                                  "AGENT_ID": str(agent_id or ""),
                                  "ENTITY_FIELDS": json.dumps(list(entity_fields or []), ensure_ascii=False)}
    sv = api("/api/expert/save", {"name": name,
                                  "description": "Qwen-driven web enrichment stage (model in the loop): reads each "
                                                 "record, takes the entity name from allowed text columns and finds "
                                                 "external suppliers via the platform Qwen (web_search). No mechanical "
                                                 "cell iteration — never searches by numbers/headers.",
                                  "code": code,
                                  "kwargs": {"input_path": "", "output_path": "", "target": "",
                                             "api_token": "", "api_base": "https://api.extella.ai"},
                                  "cspl": "fython", "global": True})
    ok = sv.get("status") == "success" or sv.get("id") is not None
    return (name if ok else None), sv


def _kp_install_on(pack_id, target):
    """Авто-установка базы знаний на устройство прогона (best-effort, не роняет сборку)."""
    try:
        body = {"name": "kp_install_pack", "params": {"pack_id": pack_id}, "global": True}
        if target:
            body["target"] = target
        return api("/api/expert/run", body, timeout=600)
    except Exception as e:
        return {"status": "error", "message": str(e)[:150]}


def _make_orchestrator(ns, stage_names, work_dir, session_id="", kp_stages=None, want_code=False, web_stages=None):
    """Создаёт (external save → persist) вызываемый оркестратор процесса с вшитыми стадиями.
    session_id вшивается (%(SID)s) для резолвера файла из общего стора на хостинге.
    kp_stages — имена knowledge-стадий, которым оркестратор пробрасывает target+api_token (реюз kp_ask).
    Это фактический КОМПИЛЯТОР pipeline_dsl (CSPL Studio S2): программа {stages,...} → эксперт.
    want_code=True — вернуть третьим элементом сгенерированный код (проверка детерминизма)."""
    name = ns + "_run_pipeline"
    code = _ORCH_TEMPLATE % {"NAME": name, "WORKDIR": work_dir,
                             "STAGES": json.dumps(stage_names, ensure_ascii=False),
                             "KP_STAGES": json.dumps(kp_stages or [], ensure_ascii=False),
                             "WEB_STAGES": json.dumps(web_stages or [], ensure_ascii=False),
                             "SID": session_id}
    sv = api("/api/expert/save", {"name": name,
                                  "description": "Auto-generated process orchestrator: runs the contract pipeline ("
                                                 + " -> ".join(stage_names) + ") on a source file, cleans headers, "
                                                 "returns summary and renders .md/.xlsx report. Params: source_file, work_dir, api_token, target, source_key.",
                                  "code": code, "kwargs": {"source_file": "", "work_dir": work_dir, "rules_json": "", "fields_json": "",
                                                           "api_token": "", "api_base": "https://api.extella.ai",
                                                           "target": "", "source_key": "", "placement_json": "", "adapter_json": "", "report_spec_json": ""},
                                  "cspl": "fython", "global": True})
    ok = sv.get("status") == "success" or sv.get("id") is not None
    if want_code:
        return (name if ok else None), sv, code
    return (name if ok else None), sv


_UPC_ORCH_TEMPLATE = r'''$extens("include.py")
include("import requests", ["extella-pip install requests"])

def __NAME__(source_file: str = "", output_dir: str = "", api_token: str = "", api_base: str = "https://api.extella.ai", target: str = "", source_key: str = "", rules_json: str = "", fields_json: str = "", run_id: str = "", placement_json: str = "", adapter_json: str = "", report_spec_json: str = "", approval_json: str = "") -> dict:
    """Universal Process Contract v1 DAG runner with durable per-run checkpoints."""
    import json
    import hashlib
    import shutil
    import uuid
    import requests
    from pathlib import Path
    from datetime import datetime, timezone

    GRAPH = __GRAPH__
    ERROR_MARKERS = ("[execution error]", "traceback (most recent call last)", "nameerror:",
                     "typeerror:", "valueerror:", "eoferror", "syntaxerror:", "runtimeerror:")
    DANGEROUS = ("move", "modify", "delete", "install", "send", "external_write")

    def stamp():
        return datetime.now(timezone.utc).isoformat()

    def atomic_json(path, value):
        tmp = path.with_name(path.name + ".tmp")
        tmp.write_text(json.dumps(value, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        tmp.replace(path)

    def result_error(value):
        try:
            text = json.dumps(value, ensure_ascii=False, default=str).casefold()
        except Exception:
            text = str(value).casefold()
        return next((marker for marker in ERROR_MARKERS if marker in text), "")

    if not api_token:
        cfg = Path.home() / "extella_wizard" / "app" / "config.json"
        try:
            api_token = json.loads(cfg.read_text(encoding="utf-8")).get("auth_token", "")
        except Exception:
            api_token = ""
    if not api_token:
        return {"status": "error", "message": "api_token is required"}
    if not source_file:
        return {"status": "error", "message": "source_file is required"}
    source = Path(source_file)
    if not source.exists():
        return {"status": "error", "message": "source_file does not exist: " + str(source)}
    if not run_id or run_id.startswith("{"):
        run_id = uuid.uuid4().hex[:16]
    root = Path(output_dir or __WORKDIR__) / "upc_runs" / run_id
    root.mkdir(parents=True, exist_ok=True)
    checkpoint_path = root / "checkpoint.json"
    try:
        checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
    except Exception:
        checkpoint = {"schema": "upc-run/1.0", "run_id": run_id,
                      "process_id": GRAPH.get("process_id"), "accepted": {}, "events": []}
    if checkpoint.get("run_id") != run_id or checkpoint.get("process_id") != GRAPH.get("process_id"):
        checkpoint = {"schema": "upc-run/1.0", "run_id": run_id,
                      "process_id": GRAPH.get("process_id"), "accepted": {}, "events": []}
    try:
        approvals = (json.loads(approval_json)
                     if approval_json and not approval_json.startswith("{{") else {})
    except Exception:
        approvals = {}
    accepted = checkpoint.setdefault("accepted", {})
    headers = {"X-Auth-Token": api_token, "Content-Type": "application/json",
               "X-Profile-Id": "default", "X-Agent-Id": "agent_extella_default"}
    placement = {}
    try:
        placement = (json.loads(placement_json)
                     if placement_json and not placement_json.startswith("{{") else {})
    except Exception:
        placement = {}
    steps = {str(step.get("id")): step for step in GRAPH.get("steps") or []}

    # A checkpoint is trusted only while its artifact evidence still exists.
    for sid in list(accepted):
        row = accepted.get(sid) or {}
        if int(row.get("step_version") or 0) != int((steps.get(sid) or {}).get("version") or 0):
            accepted.pop(sid, None)
            continue
        artifacts = row.get("artifacts") or []
        if artifacts and not all(Path(str(item.get("path") if isinstance(item, dict) else item)).exists()
                                 for item in artifacts):
            accepted.pop(sid, None)

    while len(accepted) < len(steps):
        ready = []
        for sid, step in steps.items():
            if sid in accepted:
                continue
            if all(str(dep) in accepted for dep in step.get("dependencies") or []):
                ready.append(step)
        if not ready:
            return {"status": "error", "message": "process graph is cyclic or blocked",
                    "run_id": run_id, "accepted_steps": sorted(accepted)}
        progressed = False
        for step in ready:
            sid = str(step.get("id"))
            version = int(step.get("version") or 1)
            mode = str((step.get("implementation") or {}).get("mode") or "")
            expert = str((step.get("implementation") or {}).get("expert_ref") or "")
            if mode in ("human", "acquire") or not expert:
                return {"status": "blocked_human", "run_id": run_id, "step_id": sid,
                        "question": ((step.get("human_gate") or {}).get("question") or
                                     "Шаг требует решения человека: " + str(step.get("title") or sid)),
                        "accepted_steps": sorted(accepted)}
            dangerous = [kind for kind in DANGEROUS if (step.get("permissions") or {}).get(kind)]
            approval_key = sid + ":v" + str(version)
            if dangerous and approvals.get(approval_key) is not True:
                return {"status": "blocked_human", "run_id": run_id, "step_id": sid,
                        "question": "Подтвердите точное действие перед исполнением шага",
                        "permission_preview": {"approval_key": approval_key, "permissions": dangerous,
                                               "targets": {k: (step.get("permissions") or {}).get(k)
                                                           for k in dangerous}},
                        "accepted_steps": sorted(accepted)}

            step_dir = root / ("step_" + sid) / ("v" + str(version))
            step_dir.mkdir(parents=True, exist_ok=True)
            deps = [str(x) for x in step.get("dependencies") or []]
            if not deps:
                step_source = source
            else:
                bundle = step_dir / "inputs"
                bundle.mkdir(parents=True, exist_ok=True)
                manifest = {"schema": "upc-dependency-bundle/1.0", "step_id": sid, "dependencies": []}
                for dep in deps:
                    dep_row = accepted.get(dep) or {}
                    dep_copy = {"step_id": dep, "output": dep_row.get("output"), "artifacts": []}
                    for index, raw in enumerate(dep_row.get("artifacts") or []):
                        item = raw if isinstance(raw, dict) else {"path": str(raw)}
                        src = Path(str(item.get("path") or ""))
                        if src.exists() and src.is_file():
                            dest = bundle / (dep + "_" + str(index) + "_" + src.name)
                            if not dest.exists():
                                shutil.copy2(str(src), str(dest))
                            dep_copy["artifacts"].append({"kind": item.get("kind") or "artifact",
                                                          "path": str(dest)})
                    manifest["dependencies"].append(dep_copy)
                atomic_json(bundle / "dependency_manifest.json", manifest)
                step_source = bundle
            params = {"source_file": str(step_source), "output_dir": str(step_dir / "outputs"),
                      "api_token": api_token, "api_base": api_base, "target": target,
                      "rules_json": rules_json, "fields_json": fields_json,
                      "run_id": run_id + ":" + sid + ":v" + str(version)}
            body = {"expert_name": expert, "params": params, "global": True}
            step_target = placement.get(sid) or placement.get(expert) or target
            if step_target:
                body["target"] = step_target
            try:
                response = requests.post(api_base.rstrip("/") + "/api/expert/run", headers=headers,
                                         json=body, timeout=900)
                envelope = response.json()
                result = envelope.get("result", envelope)
            except Exception as exc:
                result = {"status": "error", "message": "transport: " + str(exc)}
            marker = result_error(result)
            if not isinstance(result, dict) or result.get("status") != "success" or marker:
                checkpoint["events"].append({"at": stamp(), "type": "step_failed", "step_id": sid,
                                             "detail": marker or str(result)[:500]})
                atomic_json(checkpoint_path, checkpoint)
                return {"status": "error", "run_id": run_id, "failed_step": sid,
                        "step_version": version, "detail": marker or str(result)[:800],
                        "accepted_steps": sorted(accepted), "resumable": True}
            evidence = result.get("evidence") if isinstance(result.get("evidence"), dict) else {}
            checks = evidence.get("acceptance_checks") or result.get("acceptance_checks") or []
            if not checks or any(not isinstance(row, dict) or row.get("passed") is not True for row in checks):
                return {"status": "error", "run_id": run_id, "failed_step": sid,
                        "detail": "step returned no accepted evidence", "accepted_steps": sorted(accepted),
                        "resumable": True}
            artifacts = []
            for raw in result.get("artifacts") or []:
                item = raw if isinstance(raw, dict) else {"path": str(raw), "kind": "artifact"}
                path = Path(str(item.get("path") or ""))
                if path.exists() and path.is_file() and path.stat().st_size > 0:
                    artifacts.append({"kind": item.get("kind") or "artifact", "path": str(path),
                                      "bytes": path.stat().st_size})
            if not artifacts:
                return {"status": "error", "run_id": run_id, "failed_step": sid,
                        "detail": "step did not create an observable artifact", "accepted_steps": sorted(accepted),
                        "resumable": True}
            accepted[sid] = {"step_version": version, "expert_ref": expert,
                             "output": result.get("output") or result.get("summary"),
                             "artifacts": artifacts, "evidence": checks, "accepted_at": stamp()}
            checkpoint["events"].append({"at": stamp(), "type": "step_accepted", "step_id": sid,
                                         "step_version": version})
            atomic_json(checkpoint_path, checkpoint)
            progressed = True
        if not progressed:
            return {"status": "error", "message": "scheduler made no progress", "run_id": run_id}

    report = root / "process_report.md"
    lines = ["# " + str(GRAPH.get("title") or "Процесс"), "", "Run: `" + run_id + "`", ""]
    for step in GRAPH.get("steps") or []:
        row = accepted.get(str(step.get("id"))) or {}
        lines.append("- ✓ " + str(step.get("title") or step.get("id")) + " · v" +
                     str(row.get("step_version") or step.get("version") or 1))
    report.write_text("\n".join(lines), encoding="utf-8")
    return {"status": "success", "run_id": run_id,
            "summary": {"processed_files": 1, "accepted_steps": len(accepted), "total_steps": len(steps)},
            "output": {"accepted_steps": sorted(accepted)},
            "evidence": {"files_used": [source.name], "acceptance_checks": [
                {"criterion": "all UPC steps accepted", "passed": len(accepted) == len(steps),
                 "evidence": str(len(accepted)) + "/" + str(len(steps))}]},
            "artifacts": [{"kind": "process_report_md", "path": str(report)},
                          {"kind": "process_checkpoint_json", "path": str(checkpoint_path)}],
            "report_md": str(report)}
'''


def _make_upc_orchestrator(ns, graph, work_dir):
    """Compile accepted UPC steps into one resumable Extella expert without a second state model."""
    name = ns + "_run_process"
    public_graph = {
        "schema": graph.get("schema"), "process_id": graph.get("process_id"),
        "version": graph.get("version"), "title": graph.get("title"),
        "steps": [{k: step.get(k) for k in ("id", "title", "dependencies", "implementation",
                                               "permissions", "version")}
                  for step in graph.get("steps") or []],
    }
    code = (_UPC_ORCH_TEMPLATE.replace("__NAME__", name)
            .replace("__GRAPH__", json.dumps(public_graph, ensure_ascii=False, default=str))
            .replace("__WORKDIR__", json.dumps(str(work_dir), ensure_ascii=False)))
    saved = api("/api/expert/save", {
        "name": name,
        "description": "Universal Process Contract v1 DAG runner with checkpoint, per-step evidence and HITL.",
        "code": code,
        "kwargs": {"source_file": "", "output_dir": str(work_dir), "api_token": "",
                   "api_base": "https://api.extella.ai", "target": "", "source_key": "",
                   "rules_json": "", "fields_json": "", "run_id": "", "placement_json": "",
                   "adapter_json": "", "report_spec_json": "", "approval_json": ""},
        "cspl": "fython", "global": True,
    })
    ok = isinstance(saved, dict) and (saved.get("status") == "success" or saved.get("id"))
    return (name if ok else None), saved, code


def _run_build(session_id, build_id):
    """Фоновая стройка процесса: план -> сборка задач -> аудит. Прогресс в build_progress.json."""
    bdir = RUNS_DIR / build_id
    bdir.mkdir(parents=True, exist_ok=True)
    prog = {"build_id": build_id, "session_id": session_id, "status": "running", "stages": [],
            "agentic_events": []}

    def now():
        return datetime.now(timezone.utc).isoformat()

    def save():
        prog["updated_at"] = now()
        (bdir / "build_progress.json").write_text(json.dumps(prog, ensure_ascii=False, indent=2), encoding="utf-8")

    def _unlock():
        # снять указатель идущей стройки (building) с сессии — стройка завершилась или упала;
        # иначе UI при возврате бесконечно переподключался бы к уже мёртвой стройке
        try:
            _sp = SESS_DIR / (session_id + ".json")
            _s = json.loads(_sp.read_text(encoding="utf-8"))
            if _s.pop("building", None) is not None:
                _s["updated_at"] = now()
                _sp.write_text(json.dumps(_s, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _wait_owner(question, phase, detail=""):
        """Терминальная пауза без ложной ошибки: сохранить checkpoint и освободить build lock."""
        prog["status"] = "waiting_for_owner"
        prog["owner_question"] = str(question or "")[:1000]
        prog["waiting_phase"] = str(phase or "source_model")[:80]
        if detail:
            prog["waiting_detail"] = str(detail)[:1200]
        save()
        try:
            sp = SESS_DIR / (session_id + ".json")
            s = json.loads(sp.read_text(encoding="utf-8"))
            s.pop("building", None)
            s["waiting_build"] = {"build_id": build_id, "question": prog["owner_question"],
                                  "phase": prog["waiting_phase"], "detail": prog.get("waiting_detail", ""),
                                  "at": now()}
            s["updated_at"] = now()
            sp.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _save_process_pointer(graph, status=None):
        """Coordinate with server._update_session through the same cross-process lock file."""
        sp = SESS_DIR / (session_id + ".json")
        lock_handle = None
        try:
            import fcntl
            lock_handle = open(str(sp) + ".lock", "w")
            fcntl.flock(lock_handle, fcntl.LOCK_EX)
            current = json.loads(sp.read_text(encoding="utf-8"))
            current["process_contract"] = {
                "schema": graph.get("schema"), "path": str(SESS_DIR / (session_id + "_process.json")),
                "process_id": graph.get("process_id"), "active_version": graph.get("version"),
                "active_run_id": (graph.get("run") or {}).get("run_id") or "",
                "status": status or universal_process_status(graph), "updated_at": now(),
            }
            current["updated_at"] = now()
            atomic_write_json(sp, current)
        finally:
            if lock_handle is not None:
                try:
                    fcntl.flock(lock_handle, fcntl.LOCK_UN)
                    lock_handle.close()
                except Exception:
                    pass

    def stage(sid, title, status="running", **extra):
        stamp = now()
        for s in prog["stages"]:
            if s["id"] == sid:
                # Один логический этап может выполняться повторно (agentic repair loop).
                # Обновляем не только статус, но и подпись: иначе во второй/третьей
                # попытке кабинет продолжал показывать пользователю «попытка 1/3».
                s["title"] = title
                s["status"] = status
                s["updated_at"] = stamp
                s.update(extra)
                save()
                return
        prog["stages"].append({"id": sid, "title": title, "status": status,
                               "updated_at": stamp, **extra})
        save()

    # Сессия помнит Qwen, которого выбрал и проверил сам пользователь. Он приоритетнее статического
    # agent_id установки: интервью, стройка и последующая память тогда действительно принадлежат
    # одному мозгу. Внешний Qwen через любой OpenAI-compatible base_url/key остаётся равноправным путём.
    try:
        _build_session = json.loads((SESS_DIR / (session_id + ".json")).read_text(encoding="utf-8"))
    except Exception:
        _build_session = {}
    llm = {"api_key": CONFIG.get("llm_api_key", ""), "model": CONFIG.get("llm_model", ""),
           "base_url": CONFIG.get("llm_base_url", ""),
           "api_token": CONFIG.get("auth_token", ""), "api_base": BASE,
           "agent_id": _build_session.get("agent_id") or qwen_agent()}
    tok = {"api_token": CONFIG["auth_token"]}

    # namespace: короткий snake-префикс для экспертов процесса (из имени клиента)
    _s = _build_session
    _words = re.findall(r"[A-Za-z]+", _s.get("client_name", "") or "")
    if _words:
        ns = ("".join(w[0] for w in _words[:3]) if len("".join(w[0] for w in _words[:3])) >= 2 else _words[0][:3]).lower()
    else:
        ns = "p" + session_id.split("_")[-1][:3]
    ns = re.sub(r"[^a-z0-9]", "", ns)[:5] or "proc"

    try:
        # 1. План стройки
        stage("plan", "Составляю план стройки", "running")
        # ПЛАН строит ёмкая design-модель (fine-tune обрезает большой план); КОДОГЕН ниже — на fine-tune (llm).
        # план строит design-агент первым; при флапе — фолбэк по цепочке Qwen (run_llm_expert)
        r = run_llm_expert("wz_build_plan", dict(session_id=session_id, namespace=ns, **llm), wait=900,
                           agents=[design_agent()])
        if not isinstance(r, dict) or str(r.get("status") or "").lower() in (
                "error", "failed", "timeout", "timed_out", "cancelled"):
            error_code = "plan_transport_failed" if llm_transient_error(r) else "plan_failed"
            stage("plan", "Составляю план стройки", "error", error=str(r)[:300])
            prog["status"] = "error"; prog["error_struct"] = _build_error(error_code, str(r)[:200])
            prog["error"] = prog["error_struct"]["message"]; save(); _unlock(); return
        plan_path = SESS_DIR / (session_id + "_build_plan.json")
        if not plan_path.exists():
            stage("plan", "Составляю план стройки", "error", error="план не сохранился")
            prog["status"] = "error"; prog["error_struct"] = _build_error("plan_not_saved")
            prog["error"] = prog["error_struct"]["message"]; save(); _unlock(); return
        pdoc = json.loads(plan_path.read_text(encoding="utf-8"))
        plan = pdoc.get("plan", pdoc)
        tasks = plan.get("tasks", [])
        built_names = []
        stage("plan", "Составляю план стройки", "success", tasks_count=len(tasks))

        # UPC v1 is the canonical graph for every new build. Legacy blueprint/build-plan remain
        # compatible projections, while all four surfaces can now read this same sidecar.
        blueprint_path = SESS_DIR / (session_id + "_blueprint.json")
        blueprint_doc = json.loads(blueprint_path.read_text(encoding="utf-8"))
        blueprint = blueprint_doc.get("blueprint", blueprint_doc)
        process_path = SESS_DIR / (session_id + "_process.json")
        process_events_path = SESS_DIR / (session_id + "_process_events.jsonl")
        candidate_graph = process_from_blueprint(session_id, blueprint, plan, origin="wizard")
        process_graph = candidate_graph
        # A human answer or bridge restart must resume the same accepted graph, not rebuild every
        # independent predecessor. The deterministic process_id protects against reusing another plan.
        if process_path.exists():
            try:
                previous_graph = json.loads(process_path.read_text(encoding="utf-8"))
                if (previous_graph.get("schema") == candidate_graph.get("schema") and
                        previous_graph.get("process_id") == candidate_graph.get("process_id") and
                        universal_process_status(previous_graph) != "succeeded"):
                    process_graph = previous_graph
                    recovered = universal_recover_after_restart(process_graph)
                    if recovered:
                        process_checkpoint(process_graph, process_path, process_events_path, {
                            "type": "process_recovered", "steps": recovered,
                        })
            except Exception:
                process_graph = candidate_graph
        process_graph["run"]["run_id"] = build_id
        process_graph["task_contract_ref"] = {"path": str(bdir / "task_contract.json"), "sha256": ""}
        process_graph["source_model_ref"] = {"path": str(bdir / "source_model.json"), "sha256": ""}
        process_checkpoint(process_graph, process_path, process_events_path, {
            "type": "process_planned", "build_id": build_id, "process_id": process_graph["process_id"],
            "version": process_graph["version"], "steps": len(process_graph["steps"]),
        })
        _save_process_pointer(process_graph)
        prog["process_contract"] = {
            "schema": process_graph["schema"], "process_id": process_graph["process_id"],
            "version": process_graph["version"], "status": universal_process_status(process_graph),
            "steps": process_graph["steps"],
        }
        stage("process_contract", "Universal Process Contract собран", "success",
              detail="шагов: %d · неизвестные capability идут в generate/llm_worker" % len(process_graph["steps"]))

        schema_hint, sample_file = _inspect_sample(session_id)
        sample_files = _sample_files(session_id)

        # Движок исполнения ниже — строго линейный input_path -> output_path. До кодогена проверяем
        # две вещи, которые раньше молча игнорировались: несколько одновременных файлов и настоящий
        # DAG в depends_on. Иначе первый файл выбирался по алфавиту, ветки насильно склеивались,
        # частичный оркестратор получал allow (кейс сверки Excel+PDF у Гульжан, 20.07).
        data_tasks = [t for t in tasks if _is_pipeline_data_task(t)]
        other_tasks = [t for t in tasks if not _is_pipeline_data_task(t)]
        topology = _pipeline_topology(data_tasks)
        if not sample_files:
            # Non-file tasks (Downloads cleanup, API action, scheduled reasoning) still need a
            # reproducible build input. This envelope is an inert fixture, never a hidden fallback
            # to the user's home directory. Dangerous actions stay preview-only during acceptance.
            fixture_dir = bdir / "build_fixture"
            fixture_dir.mkdir(parents=True, exist_ok=True)
            fixture = fixture_dir / "task_input.json"
            fixture.write_text(json.dumps({
                "schema": "upc-build-fixture/1.0", "session_id": session_id,
                "goal": _build_session.get("questionnaire_task") or _build_session.get("goal") or
                        blueprint.get("goal") or blueprint.get("process_name") or "",
                "answers": _build_session.get("answers") or {},
                "execution_mode": "build_validation_preview",
                "constraints": ["do not mutate external state", "write only to output_dir"],
            }, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            sample_files = [fixture]
            schema_hint, sample_file = "", str(fixture)
            prog["synthetic_build_fixture"] = str(fixture)
            stage("topology", "Создан безопасный образец для нефайловой задачи", "success",
                  detail="внешние действия проверяются только как preview")
        # Один контракт и одна Source Model обязательны для ОБОИХ Builder-путей. Линейный Builder
        # остаётся на месте, но больше не кодогенит стадии из одного schema_hint в обход интервью/ТЗ.
        stage("task_contract", "Собираю единый Task Contract", "running")
        stage("source_model", "Qwen сопоставляет роли источников и операций", "running")
        def _source_progress(attempt, total, state, detail):
            suffix = "попытка %d/%d" % (attempt, total)
            message = ("Qwen исправляет ссылочную модель после проверки: " + str(detail)[:500]
                       if state == "running" and detail else
                       "проверяю физические входы, разделы и бизнес-сущности")
            stage("source_model", "Qwen сопоставляет источники · " + suffix, "running", detail=message)

        prepared = prepare_task_context(session_id, sample_files, SESS_DIR, llm,
                                        progress=_source_progress)
        task_package = prepared.get("package") or {}
        source_model = prepared.get("source_model") or {}
        prog["task_contract"] = task_package.get("task_contract") or {}
        prog["source_model"] = source_model
        prog["working_memory"] = task_package.get("working_memory") or {}
        for _name, _value in (("task_contract.json", prog["task_contract"]),
                              ("source_model.json", source_model),
                              ("working_memory.json", prog["working_memory"]),
                              ("task_package.json", task_package)):
            (bdir / _name).write_text(json.dumps(_value, ensure_ascii=False, indent=2, default=str),
                                      encoding="utf-8")
        process_graph["task_contract_ref"]["sha256"] = str(
            (task_package.get("task_contract") or {}).get("sha256") or "")
        process_graph["source_model_ref"]["sha256"] = hashlib.sha256(
            json.dumps(source_model, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")).hexdigest()
        process_checkpoint(process_graph, process_path, process_events_path, {
            "type": "task_context_ready", "build_id": build_id,
            "task_contract_sha256": process_graph["task_contract_ref"]["sha256"],
            "source_model_sha256": process_graph["source_model_ref"]["sha256"],
        })
        if not prepared.get("ok"):
            detail = str(prepared.get("why") or source_model.get("reason") or "Source Model не построена")
            stage("source_model", "Source Model не прошла проверку", "error", detail=detail)
            prog["status"] = "error"; prog["build_mode"] = "agentic"
            prog["error_struct"] = _build_error("source_model_failed", detail,
                                                 failure_kind="builder_defect",
                                                 draft_created=False, expert_ran=False)
            prog["error"] = prog["error_struct"]["message"]
            save(); _unlock(); return
        stage("task_contract", "Task Contract готов", "success",
              detail="контракт " + str((task_package.get("task_contract") or {}).get("sha256") or "")[:10])
        stage("source_model", "Source Model подтверждена", "success",
              detail="стратегия: " + str(source_model.get("strategy") or ""))
        if source_model.get("status") in ("need_human", "acquire"):
            error_code = "needs_owner_input" if source_model["status"] == "need_human" else "capability_missing"
            detail = str(source_model.get("reason") or source_model.get("missing_capability") or "")
            question = str(source_model.get("question") or "")
            prog["build_mode"] = "agentic"
            prog["error_struct"] = _build_error(error_code, detail, owner_question=question,
                                                 failure_kind=source_model["status"],
                                                 draft_created=False, expert_ran=False)
            prog["error"] = prog["error_struct"]["message"]
            if source_model["status"] == "need_human":
                first = next(iter(universal_ready_steps(process_graph)), None)
                if first:
                    universal_block_for_human(process_graph, first["id"], question or detail)
                    process_checkpoint(process_graph, process_path, process_events_path, {
                        "type": "step_blocked_human", "step_id": first["id"], "question": question,
                    })
                    prog["process_contract"]["steps"] = process_graph["steps"]
                    prog["process_contract"]["status"] = universal_process_status(process_graph)
                    _save_process_pointer(process_graph)
                stage("source_model", "Нужно подтвердить бизнес-смысл источников", "warn", detail=detail)
                _wait_owner(question, "source_model", detail)
                return
            prog["status"] = "error"
            save(); _unlock(); return
        # Линейный кодоген получает ту же авторитетную фактуру; agentic path получает готовый
        # context и не делает второй, потенциально расходящийся вызов Source Model.
        llm["task_context"] = _builder_brief(task_package)
        process_modes = {str((x.get("implementation") or {}).get("mode") or "")
                         for x in (process_graph.get("steps") or [])}
        needs_agentic = (len(sample_files) > 1 or not topology["supported"] or
                         source_model.get("strategy") == "holistic_build" or
                         (len(process_graph.get("steps") or []) == 1 and
                          bool(process_modes & {"generate", "llm_worker", "delegate"})))
        use_upc_scheduler = (len(process_graph.get("steps") or []) > 1 or
                             bool(process_modes & {"generate", "llm_worker", "acquire",
                                                   "human", "delegate"}))
        if use_upc_scheduler:
            facts = []
            if len(sample_files) > 1:
                facts.append("входов: %d" % len(sample_files))
            if topology["branches"]:
                facts.append("ветки: " + ", ".join(topology["branches"]))
            if topology["joins"]:
                facts.append("merge: " + ", ".join(topology["joins"]))
            stage("topology", "Граф передан пошаговому UPC runtime", "success",
                  detail="; ".join(facts) or "каждый шаг имеет отдельного эксперта и checkpoint",
                  build_mode="universal_process")
            built_steps = []
            all_verified_memory = []

            def sync_process(event=None):
                process_checkpoint(process_graph, process_path, process_events_path, event)
                prog["process_contract"]["steps"] = process_graph["steps"]
                prog["process_contract"]["status"] = universal_process_status(process_graph)
                prog["process_contract"]["version"] = process_graph.get("version")
                _save_process_pointer(process_graph)
                save()

            def step_progress(step, sid, title, status="running", detail=""):
                prefix = "upc_" + re.sub(r"[^a-zA-Z0-9_-]+", "_", str(step.get("id") or "step"))
                event_id = prefix + "_" + str(sid)
                extra = {"build_mode": "universal_process", "step_id": step.get("id"),
                         "step_version": step.get("version")}
                if detail:
                    extra["detail"] = str(detail)[:700]
                prog.setdefault("agentic_events", []).append({
                    "at": now(), "id": event_id, "step_id": step.get("id"),
                    "title": str(title)[:300], "status": status,
                    "detail": str(detail)[:700] if detail else "",
                })
                prog["agentic_events"] = prog["agentic_events"][-100:]
                stage(event_id, title, status, **extra)

            def input_files_for(step):
                deps = [str(x) for x in step.get("dependencies") or []]
                if not deps:
                    return list(sample_files)
                by_id = universal_step_map(process_graph)
                files = []
                dependency_doc = {"schema": "upc-dependency-bundle/1.0",
                                  "step_id": step.get("id"), "dependencies": []}
                for dep_id in deps:
                    dep = by_id.get(dep_id) or {}
                    dep_row = {"step_id": dep_id, "version": dep.get("version"),
                               "output": dep.get("output"), "artifacts": dep.get("artifact_refs") or []}
                    dependency_doc["dependencies"].append(dep_row)
                    for raw in dep.get("artifact_refs") or []:
                        path = Path(str(raw.get("path") if isinstance(raw, dict) else raw))
                        if path.exists() and path.is_file():
                            files.append(path)
                if files:
                    return list(dict.fromkeys(files))
                dep_dir = bdir / "process" / "dependency_inputs"
                dep_dir.mkdir(parents=True, exist_ok=True)
                dep_path = dep_dir / (str(step.get("id")) + "_v" + str(step.get("version")) + ".json")
                dep_path.write_text(json.dumps(dependency_doc, ensure_ascii=False, indent=2, default=str),
                                    encoding="utf-8")
                return [dep_path]

            while True:
                status_now = universal_process_status(process_graph)
                if status_now == "succeeded":
                    break
                ready = list(universal_ready_steps(process_graph))
                if not ready:
                    detail = "нет готовых шагов; состояние графа: " + status_now
                    prog["status"] = "error"; prog["build_mode"] = "universal_process"
                    prog["error_struct"] = _build_error("agentic_build_failed", detail,
                                                         failure_kind="scheduler")
                    prog["error"] = prog["error_struct"]["message"]
                    sync_process({"type": "process_stalled", "detail": detail})
                    _unlock(); return

                for step in ready:
                    step_id = str(step.get("id"))
                    mode = str((step.get("implementation") or {}).get("mode") or "generate")
                    if mode in ("human", "acquire"):
                        question = (("Подтвердите или предоставьте способность для шага: "
                                     if mode == "acquire" else "Нужен ответ для шага: ") +
                                    str(step.get("title") or step_id))
                        universal_block_for_human(process_graph, step_id, question)
                        sync_process({"type": "step_blocked_human", "step_id": step_id,
                                      "step_version": step.get("version"), "mode": mode,
                                      "question": question})
                        step_progress(step, "human", "Нужен человек · " + step.get("title", step_id),
                                      "warn", question)
                        _wait_owner(question, "upc_step:" + step_id, mode)
                        return

                    budgets = process_graph.get("budgets") or {}
                    run_usage = process_graph.get("run") or {}
                    step_limit = max(1, min(10, int((step.get("retry_policy") or {}).get(
                        "max_attempts") or budgets.get("max_step_attempts") or 4)))
                    remaining_attempts = int(budgets.get("max_total_attempts") or 0) - int(
                        run_usage.get("attempts_used") or 0)
                    remaining_calls = int(budgets.get("max_llm_calls") or 0) - int(
                        run_usage.get("llm_calls_used") or 0)
                    remaining_tokens = int(budgets.get("max_total_tokens") or 0) - int(
                        run_usage.get("tokens_used") or 0)
                    rate = max(0.0, float(budgets.get("estimated_cost_per_1k_tokens_usd") or 0))
                    remaining_cost = float(budgets.get("max_cost_usd") or 0) - float(
                        run_usage.get("estimated_cost_usd") or 0)
                    if int(budgets.get("max_total_attempts") or 0) > 0:
                        step_limit = min(step_limit, remaining_attempts)
                    if int(budgets.get("max_llm_calls") or 0) > 0:
                        step_limit = min(step_limit, max(0, (remaining_calls - 1) // 2))
                    if int(budgets.get("max_total_tokens") or 0) > 0:
                        step_limit = min(step_limit, max(0, remaining_tokens // 24000))
                    if float(budgets.get("max_cost_usd") or 0) > 0 and rate > 0:
                        step_limit = min(step_limit, max(0, int(remaining_cost // (24 * rate))))
                    generated_reserve = 1 if mode in ("generate", "llm_worker", "delegate") else 0
                    reserve = {"attempts": max(3, step_limit), "llm_calls": 1 + 2 * max(3, step_limit),
                               "tokens": 24000 * max(3, step_limit),
                               "generated_experts": generated_reserve}
                    gate = universal_budget_preflight(process_graph, reserve=reserve)
                    if step_limit < 3 or not gate.get("ok"):
                        detail = gate.get("message") or "недостаточно бюджета хотя бы для build/run/verify"
                        question = ("Лимит процесса исчерпан перед шагом «%s»: %s. "
                                    "Увеличьте ограничение или отмените шаг." %
                                    (step.get("title") or step_id, detail))
                        universal_block_for_human(process_graph, step_id, question)
                        sync_process({"type": "process_budget_exhausted", "step_id": step_id,
                                      "budget": gate, "question": question})
                        _wait_owner(question, "upc_budget:" + step_id, detail)
                        return

                    step_files = input_files_for(step)
                    universal_transition_step(process_graph, step_id, "running",
                                              "resolving and building step implementation")
                    sync_process({"type": "step_started", "step_id": step_id,
                                  "step_version": step.get("version"), "mode": mode,
                                  "dependencies": step.get("dependencies") or []})
                    step_progress(step, "resolve", "Собираю шаг · " + step.get("title", step_id),
                                  "running", "режим: " + mode)

                    def source_progress(attempt, total, state, detail, _step=step):
                        step_progress(_step, "source", "Проверяю входы шага · попытка %d/%d" %
                                      (attempt, total), "running" if state == "running" else state, detail)

                    step_context = prepare_task_context(
                        session_id, step_files, SESS_DIR, llm, progress=source_progress,
                        step_contract=step)
                    if not step_context.get("ok"):
                        source = step_context.get("source_model") or {}
                        question = str(source.get("question") or "")
                        detail = str(step_context.get("why") or source.get("reason") or
                                     "не удалось доказанно понять входы шага")
                        if question or source.get("status") in ("need_human", "acquire"):
                            universal_block_for_human(process_graph, step_id, question or detail)
                            sync_process({"type": "step_blocked_human", "step_id": step_id,
                                          "question": question or detail, "detail": detail})
                            _wait_owner(question or detail, "upc_step:" + step_id, detail)
                            return
                        universal_transition_step(process_graph, step_id, "failed", detail,
                                                  {"error": {"code": "source_model_failed",
                                                             "message": detail}})
                        sync_process({"type": "step_failed", "step_id": step_id,
                                      "code": "source_model_failed", "detail": detail})
                        prog["status"] = "error"; prog["build_mode"] = "universal_process"
                        prog["error_struct"] = _build_error("source_model_failed", detail,
                                                             failure_kind="builder_defect")
                        prog["error"] = prog["error_struct"]["message"]
                        save(); _unlock(); return

                    safe_step = re.sub(r"[^a-z0-9]+", "_", step_id.casefold()).strip("_")[:28] or "step"
                    stable_expert = (ns + "_" + safe_step + "_v" + str(step.get("version")))[:64]

                    def agentic_progress(sid, title, state="running", detail="", _step=step):
                        step_progress(_step, sid, title, state, detail)

                    solution = build_agentic_solution(
                        session_id=session_id, build_id=build_id, namespace=ns,
                        sample_files=step_files, sess_dir=SESS_DIR, runs_dir=RUNS_DIR,
                        llm=llm, progress=agentic_progress,
                        max_creation_attempts=max(1, min(4, step_limit - 2)),
                        max_run_repairs=2, max_acceptance_repairs=2,
                        max_total_attempts=step_limit,
                        max_elapsed_seconds=3600, prepared_context=step_context,
                        step_contract=step, expert_name_override=stable_expert)
                    attempt_count = len(solution.get("attempts") or [])
                    judged_count = sum(1 for row in (solution.get("attempts") or [])
                                       if isinstance(row, dict) and (row.get("validation") or {}).get("ok"))
                    llm_calls_used = max(1, attempt_count + judged_count + 1)
                    universal_record_usage(
                        process_graph, attempts=attempt_count, llm_calls=llm_calls_used,
                        tokens=llm_calls_used * 12000,
                        generated_experts=1 if solution.get("draft_created") else 0,
                        estimated=True)
                    sync_process({"type": "process_usage_recorded", "step_id": step_id,
                                  "usage": dict(process_graph.get("run") or {})})
                    if not solution.get("ok"):
                        detail = str(solution.get("detail") or "шаг не прошёл приёмку")
                        question = str(solution.get("owner_question") or "")
                        step["attempts"] = solution.get("attempts") or []
                        step["error"] = {"code": solution.get("code") or "agentic_acceptance_failed",
                                         "message": detail}
                        if question:
                            universal_block_for_human(process_graph, step_id, question)
                            event = "step_blocked_human"
                        else:
                            universal_transition_step(process_graph, step_id, "failed", detail)
                            event = "step_failed"
                        sync_process({"type": event, "step_id": step_id,
                                      "step_version": step.get("version"),
                                      "code": solution.get("code"), "detail": detail})
                        prog["attempts"] = solution.get("attempts") or []
                        prog["working_memory"] = solution.get("working_memory") or {}
                        prog["repair_budgets"] = solution.get("budgets") or {}
                        if question:
                            _wait_owner(question, "upc_step:" + step_id, detail)
                            return
                        prog["status"] = "error"; prog["build_mode"] = "universal_process"
                        error_code = solution.get("code") if solution.get("code") in BUILD_ERRORS else "agentic_acceptance_failed"
                        prog["error_struct"] = _build_error(
                            error_code, detail, attempts=len(solution.get("attempts") or []),
                            failure_kind=solution.get("failure_kind") or "",
                            draft_created=bool(solution.get("draft_created")),
                            expert_ran=bool(solution.get("expert_ran")),
                            last_applied_lesson=solution.get("last_applied_lesson") or "",
                            last_unapplied_lesson=solution.get("last_unapplied_lesson") or "")
                        prog["error"] = prog["error_struct"]["message"]
                        save(); _unlock(); return

                    step["implementation"]["expert_ref"] = solution.get("expert")
                    step["implementation"]["expert_version"] = step.get("version")
                    step["implementation"]["package_sha256"] = solution.get("package_sha256")
                    step_result = normalize_step_result(
                        solution.get("result") or {}, step_id, step.get("version"),
                        attempt=len(solution.get("attempts") or []) or 1)
                    raw_result = solution.get("result") if isinstance(solution.get("result"), dict) else {}
                    proposal = raw_result.get("proposed_steps") if raw_result.get("needs_subprocess") else None
                    expansion = None
                    if proposal:
                        try:
                            expansion = universal_expand_subgraph(
                                process_graph, step_id, proposal,
                                reason=raw_result.get("subprocess_reason") or step.get("purpose") or "",
                                delegation_depth=int(((step.get("delegation") or {}).get("depth") or 0)) + 1)
                        except Exception as exc:
                            detail = "предложенный подграф отклонён runtime: " + str(exc)
                            universal_transition_step(process_graph, step_id, "failed", detail,
                                                      {"error": {"code": "invalid_subgraph",
                                                                 "message": detail}})
                            sync_process({"type": "subgraph_rejected", "step_id": step_id,
                                          "detail": detail})
                            prog["status"] = "error"; prog["build_mode"] = "universal_process"
                            prog["error_struct"] = _build_error("agentic_acceptance_failed", detail,
                                                                 failure_kind="delegation")
                            prog["error"] = prog["error_struct"]["message"]
                            save(); _unlock(); return
                    memory = []
                    for raw in solution.get("verified_memory") or []:
                        if not isinstance(raw, dict) or not raw.get("text"):
                            continue
                        kind = raw.get("kind") if raw.get("kind") in ("concept", "rule") else "concept"
                        memory.append(universal_memory_entry(
                            kind, raw.get("text"), status="candidate",
                            scope=raw.get("scope") or "process",
                            source={"type": "llm_judge", "ref": solution.get("package_sha256") or ""},
                            evidence_refs=[step_result.get("raw_sha256") or ""],
                            confidence=raw.get("confidence", 0.8), step_id=step_id,
                            step_version=step.get("version"),
                            attempt=len(solution.get("attempts") or []) or 1))
                    accepted = universal_accept_step(
                        process_graph, step_id, step_result,
                        semantic_verdict=solution.get("judge") or {}, memory=memory)
                    if not accepted.get("ok"):
                        detail = "; ".join((accepted.get("validation") or {}).get("issues") or [])
                        sync_process({"type": "step_repairing", "step_id": step_id,
                                      "step_version": step.get("version"), "detail": detail})
                        prog["status"] = "error"; prog["build_mode"] = "universal_process"
                        prog["error_struct"] = _build_error("agentic_acceptance_failed", detail,
                                                             failure_kind="acceptance",
                                                             draft_created=True, expert_ran=True)
                        prog["error"] = prog["error_struct"]["message"]
                        save(); _unlock(); return
                    built_steps.append(solution.get("expert"))
                    all_verified_memory.extend(solution.get("verified_memory") or [])
                    sync_process({"type": "step_accepted", "step_id": step_id,
                                  "step_version": step.get("version"),
                                  "expert_ref": solution.get("expert"),
                                  "step_result_sha256": step_result.get("raw_sha256"),
                                  "subgraph": expansion})
                    step_progress(step, "accepted", "Шаг принят · " + step.get("title", step_id),
                                  "success", "эксперт: " + str(solution.get("expert")))

            stage("package", "Компилирую принятый граф в процесс Extella", "running",
                  build_mode="universal_process")
            orchestrator, orch_saved, orch_code = _make_upc_orchestrator(
                ns, process_graph, str(bdir / "process_runtime"))
            if not orchestrator:
                detail = "оркестратор не сохранён: " + str(orch_saved)[:500]
                prog["status"] = "error"; prog["build_mode"] = "universal_process"
                prog["error_struct"] = _build_error("orchestrator_failed", detail)
                prog["error"] = prog["error_struct"]["message"]
                save(); _unlock(); return
            built_names = list(dict.fromkeys(built_steps + [orchestrator]))
            aud = _audit_experts(built_names)
            manifest = {
                "manifest_version": 4, "contract": "upc/1.0", "build_mode": "universal_process",
                "process": ns, "process_id": process_graph.get("process_id"),
                "process_version": process_graph.get("version"), "orchestrator": orchestrator,
                "build_id": build_id, "session_id": session_id, "built_at": now(),
                "process_contract_path": str(process_path),
                "task_contract": task_package.get("task_contract") or {},
                "source_model": source_model,
                "steps": [{"id": row.get("id"), "title": row.get("title"),
                           "mode": (row.get("implementation") or {}).get("mode"),
                           "expert": (row.get("implementation") or {}).get("expert_ref"),
                           "version": row.get("version"), "status": row.get("status"),
                           "dependencies": row.get("dependencies") or []}
                          for row in process_graph.get("steps") or []],
                "verified_memory": all_verified_memory,
                "runtime": {"checkpoint": True, "local_repair": True, "human_interrupt": True,
                            "approval_json": True, "max_steps": 40},
            }
            (bdir / "upc_orchestrator.py.cspl").write_text(orch_code, encoding="utf-8")
            (bdir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2,
                                                            default=str), encoding="utf-8")
            prog.update({"audit": aud, "built_experts": built_names, "orchestrator": orchestrator,
                         "build_mode": "universal_process", "manifest": manifest,
                         "verified_memory": all_verified_memory})
            stage("package", "Граф упакован в возобновляемый процесс", "success",
                  expert=orchestrator, detail="шагов: %d · checkpoint: да" % len(built_steps),
                  build_mode="universal_process")
            stage("audit", "Проверяю код и полномочия перед запуском", "success",
                  verdict=aud["verdict"], issues=aud["issues"], build_mode="universal_process")

            sp = SESS_DIR / (session_id + ".json")
            lock_handle = None
            try:
                import fcntl
                lock_handle = open(str(sp) + ".lock", "w")
                fcntl.flock(lock_handle, fcntl.LOCK_EX)
                session = json.loads(sp.read_text(encoding="utf-8"))
                session["stage"] = "built"
                session.pop("building", None)
                session.pop("waiting_build", None)
                session.setdefault("builds", []).append({
                    "build_id": build_id, "at": now(), "build_mode": "universal_process",
                    "experts": built_names, "components_human": [row.get("title") for row in
                                                                    process_graph.get("steps") or []],
                    "audit": aud, "orchestrator": orchestrator, "manifest": manifest,
                    "source_file": str(sample_files[0]) if len(sample_files) == 1 else str(Path(sample_files[0]).parent),
                    "source_files": [str(x) for x in sample_files],
                    "task_contract": task_package.get("task_contract") or {},
                    "source_model": source_model, "verified_memory": all_verified_memory,
                    "process_contract": session.get("process_contract"), "agentic_contract": 2,
                    "process_contract_version": 1,
                })
                session["updated_at"] = now()
                atomic_write_json(sp, session)
            finally:
                if lock_handle is not None:
                    try:
                        fcntl.flock(lock_handle, fcntl.LOCK_UN)
                        lock_handle.close()
                    except Exception:
                        pass
            process_graph["run"]["finished_at"] = now()
            sync_process({"type": "process_built", "orchestrator": orchestrator,
                          "experts": built_names})
            prog["status"] = "built"
            save()
            return
        if needs_agentic:
            facts = []
            if len(sample_files) > 1:
                facts.append("одновременных файлов: %d (%s)" %
                             (len(sample_files), ", ".join(p.name for p in sample_files[:5])))
            if topology["branches"]:
                facts.append("разветвление после: " + ", ".join(topology["branches"]))
            if topology["joins"]:
                facts.append("объединение веток в: " + ", ".join(topology["joins"]))
            detail = "; ".join(facts) or "план не является последовательной цепочкой"
            stage("topology", "Несколько входов или ветки — Qwen решит задачу целиком", "success",
                  detail=detail, build_mode="agentic")

            # A single-step unknown task is the vertical UPC path: the generated Extella expert is
            # the implementation of that exact step, and only its accepted StepResult closes it.
            # Multi-step DAGs remain explicit in the ledger; they are not falsely marked accepted by
            # the legacy holistic fallback while the per-step scheduler is being applied below.
            upc_active_step = None
            if len(process_graph.get("steps") or []) == 1:
                upc_active_step = next(iter(universal_ready_steps(process_graph)), None)
                if upc_active_step:
                    universal_transition_step(process_graph, upc_active_step["id"], "running",
                                              "Builder started the generated implementation")
                    process_checkpoint(process_graph, process_path, process_events_path, {
                        "type": "step_started", "step_id": upc_active_step["id"],
                        "step_version": upc_active_step["version"], "mode": "generate",
                    })
                    prog["process_contract"]["steps"] = process_graph["steps"]
                    prog["process_contract"]["status"] = universal_process_status(process_graph)
                    _save_process_pointer(process_graph)

            def _agentic_progress(sid, title, status="running", detail=""):
                extra = {"build_mode": "agentic"}
                if detail:
                    extra["detail"] = str(detail)[:700]
                event = {"at": now(), "id": sid, "title": str(title)[:300], "status": status}
                if detail:
                    event["detail"] = str(detail)[:700]
                prog.setdefault("agentic_events", []).append(event)
                prog["agentic_events"] = prog["agentic_events"][-60:]
                # Память и модель пишутся агентным циклом атомарно; подтягиваем snapshot в UI,
                # чтобы пользователь видел, что именно следующая попытка уже получила.
                for _key, _name in (("working_memory", "working_memory.json"),
                                    ("task_contract", "task_contract.json"),
                                    ("source_model", "source_model.json")):
                    try:
                        _path = bdir / _name
                        if _path.exists():
                            prog[_key] = json.loads(_path.read_text(encoding="utf-8"))
                    except Exception:
                        pass
                stage(sid, title, status, **extra)

            solution = build_agentic_solution(
                session_id=session_id, build_id=build_id, namespace=ns,
                sample_files=sample_files, sess_dir=SESS_DIR, runs_dir=RUNS_DIR,
                llm=llm, progress=_agentic_progress, max_creation_attempts=4,
                max_run_repairs=2, max_acceptance_repairs=2, max_total_attempts=10,
                max_elapsed_seconds=3600, prepared_context=prepared)
            if not solution.get("ok"):
                detail = str(solution.get("detail") or "решение не прошло приёмку")
                if solution.get("owner_question"):
                    detail += " Вопрос владельцу: " + str(solution["owner_question"])
                prog["status"] = "error"
                prog["build_mode"] = "agentic"
                prog["attempts"] = solution.get("attempts") or []
                prog["task_contract"] = solution.get("task_contract") or {}
                prog["source_model"] = solution.get("source_model") or {}
                prog["working_memory"] = solution.get("working_memory") or {}
                prog["repair_budgets"] = solution.get("budgets") or {}
                prog["draft_created"] = bool(solution.get("draft_created"))
                prog["expert_ran"] = bool(solution.get("expert_ran"))
                prog["last_applied_lesson"] = solution.get("last_applied_lesson") or ""
                prog["last_unapplied_lesson"] = solution.get("last_unapplied_lesson") or ""
                if upc_active_step:
                    if solution.get("owner_question"):
                        universal_block_for_human(process_graph, upc_active_step["id"],
                                                  solution.get("owner_question"))
                        event_type = "step_blocked_human"
                    else:
                        universal_transition_step(process_graph, upc_active_step["id"], "failed", detail)
                        event_type = "step_failed"
                    process_checkpoint(process_graph, process_path, process_events_path, {
                        "type": event_type, "step_id": upc_active_step["id"],
                        "code": solution.get("code") or "agentic_acceptance_failed", "detail": detail,
                    })
                    prog["process_contract"]["steps"] = process_graph["steps"]
                    prog["process_contract"]["status"] = universal_process_status(process_graph)
                    _save_process_pointer(process_graph)
                error_code = solution.get("code") or "agentic_acceptance_failed"
                if error_code not in BUILD_ERRORS:
                    error_code = "agentic_acceptance_failed"
                prog["error_struct"] = _build_error(
                    error_code, detail,
                    owner_question=solution.get("owner_question") or "",
                    attempts=len(solution.get("attempts") or []),
                    failure_kind=solution.get("failure_kind") or "",
                    draft_created=bool(solution.get("draft_created")),
                    expert_ran=bool(solution.get("expert_ran")),
                    files_processed=solution.get("files_processed") or [],
                    last_applied_lesson=solution.get("last_applied_lesson") or "",
                    last_unapplied_lesson=solution.get("last_unapplied_lesson") or "")
                prog["error"] = prog["error_struct"]["message"]
                if error_code == "needs_owner_input":
                    stage("agentic_accept", "Нужен один ответ для продолжения приёмки", "warn",
                          detail=solution.get("owner_question") or detail)
                    _wait_owner(solution.get("owner_question") or "", "acceptance", detail)
                    return
                save(); _unlock(); return

            if upc_active_step:
                step_result = normalize_step_result(
                    solution.get("result") or {}, upc_active_step["id"], upc_active_step["version"],
                    attempt=len(solution.get("attempts") or []) or 1)
                accepted_memory = []
                for raw in solution.get("verified_memory") or []:
                    if not isinstance(raw, dict) or not raw.get("text"):
                        continue
                    kind = raw.get("kind") if raw.get("kind") in ("concept", "rule") else "concept"
                    accepted_memory.append(universal_memory_entry(
                        kind, raw.get("text"), status="candidate", scope=raw.get("scope") or "process",
                        source={"type": "llm_judge", "ref": solution.get("package_sha256") or ""},
                        evidence_refs=[step_result.get("raw_sha256") or ""],
                        confidence=raw.get("confidence", 0.8), step_id=upc_active_step["id"],
                        step_version=upc_active_step["version"],
                        attempt=len(solution.get("attempts") or []) or 1))
                upc_acceptance = universal_accept_step(
                    process_graph, upc_active_step["id"], step_result,
                    semantic_verdict=solution.get("judge") or {}, memory=accepted_memory)
                process_checkpoint(process_graph, process_path, process_events_path, {
                    "type": "step_accepted" if upc_acceptance.get("ok") else "step_repairing",
                    "step_id": upc_active_step["id"], "step_version": upc_active_step["version"],
                    "step_result_sha256": step_result.get("raw_sha256"),
                    "issues": (upc_acceptance.get("validation") or {}).get("issues") or [],
                })
                prog["process_contract"]["steps"] = process_graph["steps"]
                prog["process_contract"]["status"] = universal_process_status(process_graph)
                _save_process_pointer(process_graph)
                if not upc_acceptance.get("ok"):
                    detail = "; ".join((upc_acceptance.get("validation") or {}).get("issues") or [])
                    stage("upc_accept", "UPC не принял StepResult", "error", detail=detail)
                    prog["status"] = "error"
                    prog["error_struct"] = _build_error("agentic_acceptance_failed", detail,
                                                         failure_kind="acceptance",
                                                         draft_created=True, expert_ran=True)
                    prog["error"] = prog["error_struct"]["message"]
                    save(); _unlock(); return
                stage("upc_accept", "Шаг UPC принят и сохранён в checkpoint", "success",
                      detail=upc_active_step["title"] + " · v" + str(upc_active_step["version"]))

            expert = solution["expert"]
            stage("package", "Упаковываю доказанное решение в процесс Extella", "running",
                  build_mode="agentic")
            aud = _audit_experts([expert])
            prog["audit"] = aud
            prog["built_experts"] = [expert]
            prog["orchestrator"] = expert
            prog["slice_summary"] = solution.get("summary") or {}
            prog["build_mode"] = "agentic"
            prog["acceptance"] = solution.get("judge") or {}
            prog["source_files"] = solution.get("source_files") or []
            prog["task_contract"] = solution.get("task_contract") or {}
            prog["source_model"] = solution.get("source_model") or {}
            prog["working_memory"] = solution.get("working_memory") or {}
            prog["verified_memory"] = solution.get("verified_memory") or []
            manifest = {
                "manifest_version": 3,
                "build_mode": "agentic",
                "process": ns,
                "orchestrator": expert,
                "build_id": build_id,
                "session_id": session_id,
                "built_at": now(),
                "task_package_sha256": solution.get("package_sha256"),
                "task_contract": solution.get("task_contract") or {},
                "source_model": solution.get("source_model") or {},
                "strategy": solution.get("strategy") or "holistic_build",
                "input": {"kind": "file_bundle" if len(sample_files) > 1 else "file",
                          "files": solution.get("source_files") or []},
                "output": {"summary_keys": sorted((solution.get("summary") or {}).keys()),
                           "reports": ["md", "xlsx"]},
                "steps": [{"name": expert, "title": "Целостное решение Qwen", "mode": "agentic", "ok": True}],
                "contracts": {"params": 1, "agentic": 2, "task_contract": 1, "source_model": 1,
                              "memory": 1, "multi_input": int(len(sample_files) > 1)},
                "acceptance": solution.get("judge") or {},
                "verified_memory": solution.get("verified_memory") or [],
            }
            prog["manifest"] = manifest
            stage("package", "Решение упаковано в исполняемого эксперта", "success", expert=expert,
                  detail="сначала доказан результат, затем создан процесс", build_mode="agentic")
            stage("audit", "Проверяю код и полномочия перед запуском", "success",
                  verdict=aud["verdict"], issues=aud["issues"], build_mode="agentic")

            # Новый режим использует того же кабинета/агента/историю запусков: наружный контракт build
            # не меняется, но внутри один эксперт уже доказал всю задачу на пакете файлов.
            sp = SESS_DIR / (session_id + ".json")
            s = json.loads(sp.read_text(encoding="utf-8"))
            s["stage"] = "built"
            s.pop("building", None)
            s.setdefault("builds", []).append({
                "build_id": build_id, "at": now(), "build_mode": "agentic",
                "experts": [expert], "components_human": ["Целостное решение Qwen"],
                "audit": aud, "orchestrator": expert,
                "slice_summary": solution.get("summary") or {},
                "manifest": manifest, "acceptance": solution.get("judge") or {},
                "source_file": solution.get("source_file"),
                "source_files": solution.get("source_files") or [],
                "task_contract": solution.get("task_contract") or {},
                "source_model": solution.get("source_model") or {},
                "strategy": solution.get("strategy") or "holistic_build",
                "verified_memory": solution.get("verified_memory") or [],
                "attempts": solution.get("attempts") or [],
                "agentic_contract": 2, "params_contract": 1,
                "report_contract": 0, "adapter_contract": 0, "placement_contract": 0,
            })
            if not s.get("panel_manifest"):
                try:
                    _bpm = json.loads((SESS_DIR / (session_id + "_blueprint.json")).read_text(
                        encoding="utf-8")).get("blueprint", {})
                    _mani = gen_panel_manifest(_bpm.get("goal") or _bpm.get("summary") or "",
                                               _bpm.get("stages") or [])
                    if _mani:
                        s["panel_manifest"] = dict(_mani, generated_at=now())
                except Exception:
                    pass
            s["updated_at"] = now()
            sp.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding="utf-8")
            # Публикуем built ПОСЛЕ записи сборки в сессию. Иначе UI/автотест успевает открыть
            # кабинет между двумя write и ловит «процесс ещё не построен» при уже зелёном прогрессе.
            prog["status"] = "built"
            save()
            return
        stage("topology", "Входные данные образуют последовательную цепочку", "success")

        # KNOWLEDGE-СТАДИЯ: из blueprint берём базу знаний (knowledge_pack) и какие стадии на неё опираются.
        # Такие стадии НЕ кодогеним — реюзаем готовый kp_ask (тонкая обёртка). Базу авто-ставим на target прогона.
        kp_pack = ""
        kp_stage_ids = set()
        try:
            _bp = json.loads((SESS_DIR / (session_id + "_blueprint.json")).read_text(encoding="utf-8")).get("blueprint", {})
            kp_pack = ((_bp.get("knowledge_pack") or {}).get("pack_id")) or ""
            for _st in _bp.get("stages", []):
                if "knowledge_grounding" in (_st.get("capability_ids") or []):
                    kp_stage_ids.add(_st.get("id"))
        except Exception:
            pass
        # WZ-B03 (ТЗ v2 §16.3): тип knowledge-стадии — ТОЛЬКО из манифеста blueprint
        # (capability_ids содержит knowledge_grounding). Keyword-эвристика запрещена:
        # «норм» ловила «нормализованных» и превращала processing-стадию в knowledge.
        def is_kp_task(t):
            return bool(kp_pack) and t.get("stage_id") in kp_stage_ids

        kp_stage_names = []
        web_stage_names = []   # Qwen-ведомые шаги веб-обогащения (реюз шаблона, не кодоген)
        if kp_pack and not kp_stage_ids:
            # база объявлена, но НИ ОДНА стадия не помечена knowledge_grounding — манифест неконсистентен;
            # честно показываем (раньше keyword-эвристика маскировала это, угадывая стадию по словам)
            stage("kp_manifest", "База " + kp_pack + " объявлена, но стадии-потребители не помечены в blueprint — "
                  "knowledge-реюз пропущен", "success", warning=True)
        if kp_pack:
            stage("kp_install", "Ставлю базу знаний: " + kp_pack, "running")
            _ki = _kp_install_on(kp_pack, HOST_TARGET)
            _kis = _ki.get("status", "?") if isinstance(_ki, dict) else "?"
            stage("kp_install", "База знаний " + kp_pack + " (" + str(_kis) + ")", "success", pack=kp_pack)  # best-effort

        # ДАТА-СТАДИИ конвейера (парсинг/анализ/отчёт) — строим ВСЕ заново под единый контракт
        # (реюз старых экспертов не по контракту рвёт цепочку). Не-дата задачи (расписание) — вне среза.
        for t in other_tasks:
            tid = t.get("id", "x")
            stage("task_" + tid, "После сборки: " + _human_title(t, ns), "warn", skipped=True,
                  runtime_setup=True, detail="Это настройка запуска, а не обработка данных; "
                                                   "источник и режим 24/7 задаются в кабинете процесса")

        # 2. Сборка МОСТОМ по единому контракту + вертикальный срез на реальном файле:
        #    каждая дата-стадия принимает выход предыдущей (первая — исходный файл клиента).
        current_input = sample_file
        slice_ok = bool(sample_file)
        stage_doubts = []   # шаги, собранные структурно, но сомнительные по смыслу (требуют доводки)
        failed_stage = None
        failed_index = None
        _entity_cols, _all_cols = _entity_columns(session_id)   # колонки для веб-поиска (детерминированно)
        _skip_ids = set()   # веб-поиск, которому не по чему искать — исключаем из каркаса ПРАВИЛЬНО, а не строим мусор
        for idx, t in enumerate(data_tasks):
            tid = t.get("id", "t%d" % (idx + 1))
            title = _human_title(t, ns)
            nm = t.get("expert_name") or (ns + "_" + tid)
            # ГЕЙТ «каркас правильно»: веб-поиск обязан искать по колонке-СУЩНОСТИ. Нет текстовой колонки
            # в источнике → шаг не строим (честно пропускаем). Есть → это Qwen-ВЕДОМЫЙ шаг (модель в
            # контуре), а не механический кодоген — см. _build_websearch_stage (идея Анвара 20.07).
            _is_web = _is_websearch_task(t)
            if _is_web and not _entity_cols:
                _skip_ids.add(tid)
                stage_doubts.append({"expert": nm, "title": title,
                                     "why": "в источнике нет текстовой колонки-названия для веб-поиска — шаг пропущен"})
                stage("task_" + tid, "Пропущен (искать не по чему): " + title, "warn",
                      expert=nm, detail="в источнике нет колонки-названия для внешнего поиска; "
                                        "добавьте её или уберите этот шаг в чате доводки", needs_review=True)
                continue
            stage("task_" + tid, "Собираю и проверяю: " + title, "running")
            if not current_input:
                _no_input = _build_error("no_sample_file",
                                         "шаг «" + str(title) + "» остался без входных данных")
                stage("task_" + tid, "Ошибка: " + title, "error", expert=nm,
                      detail=_no_input["detail"], code=_no_input["code"], remedy=_no_input["remedy"])
                slice_ok = False
                failed_index = idx
                failed_stage = {"id": tid, "expert": nm, "title": title,
                                "detail": "нет входа для стадии"}
                break
            if _is_web:
                # QWEN-ВЕДОМЫЙ ШАГ: детерминированный шаблон, внутри — платформенная Qwen (web_search).
                # На build-срезе passthrough (реальное обогащение — на прогоне, как kp): не гоняем веб на
                # каждой сборке. Модель сама читает строку и ищет по названию — не по суммам/заголовкам.
                web_stage_names.append(nm)
                nm2, sv = _build_websearch_stage(nm, _entity_cols, llm.get("agent_id"))
                outp = str(bdir / (nm + "_out.json"))
                if not nm2:
                    ok, detail = False, "веб-стадия не сохранилась: " + str(sv)[:120]
                else:
                    try:
                        import shutil as _sh
                        _sh.copy(current_input, outp)   # passthrough: обогащение выполнится на прогоне
                        ok, detail = True, "Qwen-ведомое веб-обогащение по колонкам " + \
                            json.dumps(_entity_cols, ensure_ascii=False) + " (модель в контуре; выполнится на прогоне)"
                    except Exception as _e:
                        ok, detail = False, "passthrough веб-стадии не удался: " + str(_e)[:120]
                doubt = None
            elif is_kp_task(t):
                # РЕЮЗ kp_ask: сохраняем тонкую обёртку и прогоняем на входе. На build-срезе без target
                # обёртка мягко деградирует (пустой legal_context) — на run-time оркестратор даёт target.
                kp_stage_names.append(nm)
                nm2, sv = _build_kp_stage(nm, kp_pack)
                outp = str(bdir / (nm + "_out.json"))
                if not nm2:
                    ok, detail = False, "kp-обёртка не сохранилась: " + str(sv)[:120]
                else:
                    try:
                        api("/api/expert/run", {"name": nm, "params": {"input_path": current_input, "output_path": outp},
                                                "global": True}, timeout=200)
                    except Exception:
                        pass
                    if not Path(outp).exists():
                        try:
                            import shutil as _sh
                            _sh.copy(current_input, outp)  # passthrough: срез продолжается, retrieval — на run-time
                        except Exception:
                            outp = current_input
                        ok, detail = True, "реюз kp_ask(" + kp_pack + "): на build-срезе passthrough — retrieval выполнится на run-time"   # #28: честная метка вместо success-строки
                    else:
                        ok, detail = True, "реюз kp_ask(" + kp_pack + ")"
                doubt = None
            else:
                ok, outp, detail, doubt = _build_one(nm, t, schema_hint, is_first=(idx == 0),
                                                     is_last=(idx == len(data_tasks) - 1),
                                                     accept_input=current_input, llm=llm)
            if ok:
                built_names.append(nm)
                current_input = outp  # выход стадии = вход следующей (это и есть срез)
            else:
                slice_ok = False
            # ТРИ состояния, не два: собрано / собрано-но-требует-доводки / ошибка. «Требует доводки» —
            # структурно цело, а смысл сомнителен (см. _stage_sanity): показываем ЧЕСТНО, не зелёным.
            if ok and doubt:
                stage_doubts.append({"expert": nm, "title": title, "why": doubt})
                stage("task_" + tid, "Собрано, требует доводки: " + title, "warn",
                      expert=nm, detail=str(doubt)[:200], needs_review=True)
            else:
                stage("task_" + tid, ("Собрано+прогнано: " if ok else "Ошибка: ") + title,
                      "success" if ok else "error", expert=nm, detail=str(detail)[:200])
            if not ok:
                failed_index = idx
                failed_stage = {"id": tid, "expert": nm, "title": title,
                                "detail": str(detail)[:300]}
                break

        # Первый реальный провал — корневая причина. Всё ниже по цепочке не «не собрано», а вообще
        # не запускалось. Частичный оркестратор и audit=allow здесь были ложным сигналом готовности.
        if failed_stage:
            downstream = []
            for j, t in enumerate(data_tasks[(failed_index or 0) + 1:], start=(failed_index or 0) + 1):
                tid = t.get("id", "t%d" % (j + 1))
                title = _human_title(t, ns)
                downstream.append(title)
                stage("task_" + tid, "Не запускался: " + title, "blocked",
                      detail="предыдущий шаг «%s» не прошёл проверку" % failed_stage["title"])
            stage("orchestrator", "Оркестратор не собирался", "blocked",
                  detail="сначала должен успешно пройти каждый шаг данных")
            stage("audit", "Проверка допуска не проводилась", "blocked",
                  detail="неполный процесс нельзя допускать к запуску")
            prog["audit"] = {"verdict": "not_run", "issues": ["процесс собран не полностью"]}
            prog["built_experts"] = [n for n in built_names if n]
            prog["failed_stage"] = failed_stage
            prog["status"] = "error"
            prog["error_struct"] = _build_error(
                "stage_failed",
                "шаг «%s»: %s" % (failed_stage["title"], failed_stage["detail"]),
                failed_stage=failed_stage, downstream=downstream)
            prog["error"] = prog["error_struct"]["message"]
            save(); _unlock(); return

        # итог среза: последний output = сводка
        slice_summary = None
        if slice_ok and current_input and current_input != sample_file and Path(current_input).exists():
            try:
                sdata = json.loads(Path(current_input).read_text(encoding="utf-8"))
                slice_summary = sdata.get("summary") if isinstance(sdata, dict) else {"records": len(sdata)}
                prog["slice_output"] = current_input
                prog["slice_summary"] = slice_summary
            except Exception:
                pass

        built_ok = [n for n in built_names if n]
        had_build_tasks = any(str(t.get("action", "build")).lower() != "reuse" for t in tasks)
        if had_build_tasks and not built_ok:
            prog["status"] = "error"
            prog["error_struct"] = _build_error("no_components_built")
            prog["error"] = prog["error_struct"]["message"]
            save(); _unlock(); return

        # 3. Автосоздание вызываемого оркестратора процесса (стадии — построенные дата-эксперты)
        orchestrator = None
        stage_experts = [t.get("expert_name") or (ns + "_" + t.get("id", "")) for t in data_tasks]
        stage_experts = [n for n in stage_experts if n in built_ok]
        if stage_experts:
            stage("orchestrator", "Собираю оркестратор процесса", "running")
            orchestrator, _sv = _make_orchestrator(ns, stage_experts, "/tmp/" + ns + "_run", session_id,
                                                   kp_stages=[n for n in kp_stage_names if n in built_ok],
                                                   web_stages=[n for n in web_stage_names if n in built_ok])
            stage("orchestrator", "Оркестратор процесса: " + (orchestrator or "ошибка"),
                  "success" if orchestrator else "error", expert=orchestrator)
            if orchestrator:
                prog["orchestrator"] = orchestrator

        # 4. Аудит перед запуском
        stage("audit", "Проверяю процесс перед запуском", "running")
        aud = _audit_experts([n for n in built_names if n])
        prog["audit"] = aud
        prog["built_experts"] = [n for n in built_names if n]
        stage("audit", "Проверяю процесс перед запуском", "success",
              verdict=aud["verdict"], issues=aud["issues"])

        # WZ-B01 (ТЗ v2 §25): built — ТОЛЬКО по полному обязательному конвейеру.
        # Любая упавшая/несобранная дата-стадия, сорванный срез или несобранный оркестратор
        # блокируют built (раньше статус ставился безусловно и маскировал провал стадии).
        missing = [(t.get("title") or t.get("expert_name") or t.get("id") or "?")
                   for i, t in enumerate(data_tasks)
                   if t.get("id", "t%d" % (i + 1)) not in _skip_ids   # намеренно пропущенный веб-поиск ≠ провал сборки
                   and (t.get("expert_name") or (ns + "_" + t.get("id", "t%d" % (i + 1)))) not in built_ok]
        if missing or not slice_ok or (stage_experts and not orchestrator):
            if missing:
                prog["error_struct"] = _build_error("stages_missing",
                                                    "не собраны шаги: " + ", ".join(map(str, missing)),
                                                    missing=[str(m) for m in missing])
            elif stage_experts and not orchestrator:
                prog["error_struct"] = _build_error("orchestrator_failed")
            else:
                prog["error_struct"] = _build_error("slice_failed")
            prog["status"] = "error"
            prog["error"] = prog["error_struct"]["message"]
            save(); _unlock(); return

        prog["manifest"] = _process_manifest(ns, orchestrator, data_tasks, built_ok, sample_file,
                                             prog.get("slice_summary"), kp_pack, build_id, session_id,
                                             skip_ids=_skip_ids)
        prog["manifest"].update({"task_contract": task_package.get("task_contract") or {},
                                 "source_model": source_model,
                                 "strategy": source_model.get("strategy") or "compose",
                                 "verified_memory": []})
        # Сомнительные по смыслу шаги — в итог сборки, чтобы «готово» не скрывало доводку.
        # Процесс собран (built), но честно сказано, какие шаги стоит проверить/поправить словами.
        if stage_doubts:
            prog["needs_review"] = stage_doubts
        prog["status"] = "built"
        save()
        # отметка в сессии
        try:
            sp = SESS_DIR / (session_id + ".json")
            s = json.loads(sp.read_text(encoding="utf-8"))
            s["stage"] = "built"
            s.pop("building", None)   # стройка успешно завершилась — снять указатель идущей стройки
            s.setdefault("builds", []).append({"build_id": build_id, "at": now(),
                                               "experts": prog["built_experts"], "audit": aud,
                                               "report_contract": 1,   # оформитель PDF по спеке вида
                                               "adapter_contract": 1,   # AC-05: оркестратор умеет применять адаптер источника
                                               "params_contract": 1,   # F2: оркестратор+стадии принимают rules_json/fields_json
                                               "placement_contract": 1,   # A1: оркестратор понимает карту размещения (стадия→устройство)
                                               "orchestrator": orchestrator,
                                               "slice_summary": prog.get("slice_summary"),
                                               "manifest": prog.get("manifest"),   # AC-06: контракт собранного процесса
                                               "task_contract": task_package.get("task_contract") or {},
                                               "source_model": source_model,
                                               "strategy": source_model.get("strategy") or "compose",
                                               "verified_memory": [],
                                               "needs_review": stage_doubts,   # шаги, требующие смысловой доводки
                                               "source_file": sample_file})
            # §7bis ступень 3: автопанель — новая автоматизация выходит из Строителя с готовой формой
            # настроек (best-effort; если Qwen моргнул — владелец соберёт кнопкой в кабинете).
            if not s.get("panel_manifest"):
                try:
                    _bpf = SESS_DIR / (session_id + "_blueprint.json")
                    _bpm = json.loads(_bpf.read_text(encoding="utf-8")).get("blueprint", {}) if _bpf.exists() else {}
                    _mani = gen_panel_manifest(_bpm.get("goal") or _bpm.get("summary") or "", _bpm.get("stages") or [])
                    if _mani:
                        s["panel_manifest"] = dict(_mani, generated_at=now())
                except Exception:
                    pass
            s["updated_at"] = now()
            sp.write_text(json.dumps(s, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
    except Exception as e:
        prog["status"] = "error"
        prog["error_struct"] = _build_error("crashed", str(e)[:200])
        prog["error"] = prog["error_struct"]["message"]; save(); _unlock()
