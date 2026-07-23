"""Агентная стройка Визарда: решить задачу целиком -> прогнать -> исправить -> доказать.

Этот слой не назначает тип эксперта заранее. Qwen сначала получает Task Contract и фактические профили,
выбирает task-first маршрут (reuse/build/llm_worker/split/replan), и лишь затем харнесс материализует
исполняемого эксперта, независимо запускает его и проверяет факты результата.
"""
import ast
import hashlib
import importlib.util
import base64
import copy
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
import unicodedata
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from wz_platform import BASE, CONFIG, api, qwen_agent, run_expert
from wz_llm import design_agent
try:
    from wz_process import (adaptive_failure_decision, canonical_artifact_requirements,
                            step_can_partition)
except ModuleNotFoundError:  # direct-file QA loaders do not add ui/ to sys.path
    _process_spec = importlib.util.spec_from_file_location(
        "wz_process", Path(__file__).with_name("wz_process.py"))
    _process_module = importlib.util.module_from_spec(_process_spec)
    _process_spec.loader.exec_module(_process_module)
    adaptive_failure_decision = _process_module.adaptive_failure_decision
    canonical_artifact_requirements = _process_module.canonical_artifact_requirements
    step_can_partition = _process_module.step_can_partition


EXPERT_KWARGS = {
    "source_file": "", "output_dir": "", "api_token": "", "api_base": "https://api.extella.ai",
    "target": "", "source_key": "", "rules_json": "", "fields_json": "", "run_id": "",
    "placement_json": "", "adapter_json": "", "report_spec_json": "",
}
REPORT_KEYS = ("report_md", "report_xlsx", "report_pdf", "report_docx", "report_pptx")
SOURCE_STATUSES = ("ready", "need_human", "acquire")
BUILD_STRATEGIES = ("reuse", "compose", "build", "holistic_build", "acquire", "need_human")
EXECUTION_ROUTES = ("reuse", "build", "llm_worker", "split_step", "replan", "need_human", "acquire")
MEMORY_KINDS = ("evidence", "lesson", "concept", "rule", "artifact")
MEMORY_STATUSES = ("candidate", "verified", "rejected", "superseded")


def _clip(value, limit=400):
    s = str(value if value is not None else "").replace("\x00", "").strip()
    return s if len(s) <= limit else s[:limit] + "…"


def _nfc(value):
    """Одинаковое представление Unicode для macOS NFD и обычных NFC-строк."""
    return unicodedata.normalize("NFC", str(value or ""))


def _ref_key(value, basename=False):
    """Ключ ссылки Source Model: Unicode NFC + casefold; при необходимости только basename."""
    text = _nfc(value).strip()
    if basename:
        text = Path(text).name
    return text.casefold()


def _compact(value, depth=0):
    """Ограниченный JSON-контекст: сохраняет структуру ТЗ, не отправляет модели бесконечные логи."""
    if depth > 6:
        return _clip(value, 300)
    if isinstance(value, dict):
        return {str(k)[:80]: _compact(v, depth + 1) for k, v in list(value.items())[:80]}
    if isinstance(value, list):
        return [_compact(v, depth + 1) for v in value[:60]]
    if isinstance(value, str):
        return _clip(value, 1800)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    return _clip(value, 500)


def _sha256(path):
    h = hashlib.sha256()
    with Path(path).open("rb") as fh:
        while True:
            chunk = fh.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _pdf_text(path):
    """Текст первых страниц; для скана best-effort OCR. Отсутствие движка — факт профиля, не успех."""
    text, source = "", ""
    for mod in ("pypdf", "PyPDF2"):
        try:
            pkg = __import__(mod)
            reader = pkg.PdfReader(str(path))
            text = "\n".join((p.extract_text() or "") for p in reader.pages[:5])
            if text.strip():
                source = mod
                break
        except Exception:
            continue
    if not text.strip() and shutil.which("pdftotext"):
        try:
            cp = subprocess.run(["pdftotext", "-f", "1", "-l", "5", str(path), "-"],
                                capture_output=True, text=True, timeout=35)
            text = cp.stdout or ""
            if text.strip():
                source = "pdftotext"
        except Exception:
            pass
    if not text.strip() and shutil.which("pdftoppm") and shutil.which("tesseract"):
        try:
            with tempfile.TemporaryDirectory(prefix="wz_pdf_") as td:
                pref = str(Path(td) / "page")
                subprocess.run(["pdftoppm", "-f", "1", "-l", "2", "-r", "150", "-png", str(path), pref],
                               capture_output=True, timeout=60)
                parts = []
                for image in sorted(Path(td).glob("page-*.png")):
                    cp = subprocess.run(["tesseract", str(image), "stdout", "-l", "rus+eng"],
                                        capture_output=True, text=True, timeout=60)
                    parts.append(cp.stdout or "")
                text = "\n".join(parts)
                if text.strip():
                    source = "tesseract"
        except Exception:
            pass
    return _clip(text, 9000), source or "unavailable_or_scan"


def _xlsx_profile_ooxml(path):
    """Read-only XLSX profile without openpyxl, for clean Macs and QA installers.

    The Builder only needs sheet identity plus a bounded cell sample.  OOXML stores those facts in
    workbook relationships, shared strings and worksheet XML, so installing a spreadsheet runtime
    into the bridge is unnecessary.  Full business processing remains the generated expert's job.
    """
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
          "p": "http://schemas.openxmlformats.org/package/2006/relationships"}

    def column_index(cell_ref):
        letters = re.match(r"[A-Za-z]+", str(cell_ref or ""))
        value = 0
        for char in (letters.group(0).upper() if letters else ""):
            value = value * 26 + ord(char) - 64
        return max(0, value - 1)

    def texts(node):
        return "".join(str(x.text or "") for x in node.findall(".//m:t", ns))

    with zipfile.ZipFile(str(path)) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relations = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {str(row.attrib.get("Id") or ""): str(row.attrib.get("Target") or "")
                   for row in relations.findall("p:Relationship", ns)}
        shared = []
        if "xl/sharedStrings.xml" in archive.namelist():
            strings = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = [texts(item) for item in strings.findall("m:si", ns)]
        sheets = []
        for sheet in workbook.findall("m:sheets/m:sheet", ns)[:12]:
            rid = str(sheet.attrib.get("{%s}id" % ns["r"]) or "")
            target = targets.get(rid, "")
            if target.startswith("/"):
                target = target.lstrip("/")
            elif not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            target = str(Path(target))
            if target not in archive.namelist():
                continue
            worksheet = ET.fromstring(archive.read(target))
            row_values, max_row, max_column = [], 0, 0
            for row in worksheet.findall(".//m:sheetData/m:row", ns)[:30]:
                values = {}
                row_number = int(row.attrib.get("r") or (len(row_values) + 1))
                max_row = max(max_row, row_number)
                for cell in row.findall("m:c", ns):
                    index = column_index(cell.attrib.get("r"))
                    if index >= 30:
                        continue
                    kind = str(cell.attrib.get("t") or "")
                    raw_node = cell.find("m:v", ns)
                    raw = str(raw_node.text or "") if raw_node is not None else ""
                    if kind == "s":
                        try:
                            value = shared[int(raw)]
                        except Exception:
                            value = raw
                    elif kind == "inlineStr":
                        value = texts(cell)
                    elif kind == "b":
                        value = raw == "1"
                    elif kind in ("str", "e"):
                        value = raw
                    else:
                        try:
                            value = int(raw) if re.fullmatch(r"[-+]?\d+", raw) else float(raw)
                        except Exception:
                            value = raw
                    values[index] = _clip(value, 180)
                    max_column = max(max_column, index + 1)
                width = min(30, max(values.keys(), default=-1) + 1)
                row_values.append([values.get(index, "") for index in range(width)])
            header_row, best = 0, -1
            for index, row in enumerate(row_values):
                filled = sum(1 for value in row if str(value or "").strip())
                textual = sum(1 for value in row if str(value or "").strip() and
                              not re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", str(value).strip()))
                if filled + textual > best:
                    best, header_row = filled + textual, index
            sheets.append({"title": _nfc(sheet.attrib.get("name")),
                           "max_row": max_row or len(row_values), "max_column": max_column,
                           "header_row": header_row + 1,
                           "columns": [_nfc(v) for v in
                                       (row_values[header_row] if row_values else []) if str(v).strip()],
                           "sample_rows": row_values})
        return sheets


def profile_file(path):
    """Фактический, ограниченный профиль входа для рассуждения Qwen и независимой приёмки."""
    p = Path(path)
    display_name = _nfc(p.name)
    out = {"name": display_name, "extension": p.suffix.lower(), "bytes": p.stat().st_size,
           "sha256": _sha256(p)}
    if display_name != p.name:
        # Физическое имя нужно только харнессу; Qwen и Source Model работают со stable source_id.
        out["filesystem_name"] = p.name
    ext = p.suffix.lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                sheets = []
                # Порядок листов не является бизнес-смыслом. Профилируем достаточно широкий набор,
                # чтобы Строитель не принял первый попавшийся лист за нужный реестр (кейс 20.07),
                # но ограничиваем строки/колонки для предсказуемого контекста Qwen.
                for ws in wb.worksheets[:12]:
                    rows = []
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 30), values_only=True):
                        rows.append([_clip(v, 180) for v in list(row)[:30]])
                    header_row, best = 0, -1
                    for index, row in enumerate(rows):
                        filled = sum(1 for value in row if str(value or "").strip())
                        textual = sum(1 for value in row if str(value or "").strip() and
                                      not re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", str(value).strip()))
                        score = filled + textual
                        if score > best:
                            best, header_row = score, index
                    sheets.append({"title": _nfc(ws.title), "max_row": ws.max_row,
                                   "max_column": ws.max_column, "header_row": header_row + 1,
                                   "columns": [_nfc(v) for v in
                                               (rows[header_row] if rows else []) if str(v).strip()],
                                   "sample_rows": rows})
                wb.close()
                out["workbook"] = sheets
            except (ImportError, ModuleNotFoundError):
                out["workbook"] = _xlsx_profile_ooxml(p)
        elif ext == ".csv":
            import csv
            with p.open("r", encoding="utf-8", errors="replace", newline="") as fh:
                rows = []
                for i, row in enumerate(csv.reader(fh)):
                    rows.append([_clip(v, 180) for v in row[:30]])
                    if i >= 29:
                        break
            out["sample_rows"] = rows
            out["columns"] = [_nfc(v) for v in rows[0]] if rows else []
        elif ext == ".pdf":
            text, source = _pdf_text(p)
            out["text_sample"], out["text_source"] = text, source
        elif ext == ".docx":
            # DOCX — zip с document.xml. Для Source Model достаточно текста без оформления;
            # выполнение рабочего процесса по-прежнему выбирает нужную библиотеку само.
            import zipfile
            with zipfile.ZipFile(str(p)) as zf:
                raw = zf.read("word/document.xml").decode("utf-8", errors="replace")
            out["text_sample"] = _clip(re.sub(r"<[^>]+>", " ", raw), 9000)
        elif ext in (".txt", ".md", ".json", ".xml", ".html"):
            out["text_sample"] = _clip(p.read_text(encoding="utf-8", errors="replace"), 9000)
    except Exception as exc:
        out["profile_error"] = _clip(exc, 300)
    return out


def _read_json(path, fallback=None):
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return fallback


def _answer_rows(answers):
    """Адаптивное интервью: id вопроса не несёт бизнес-смысла, важны все показанные пары."""
    out = []
    for qid, raw in (answers or {}).items():
        if isinstance(raw, dict):
            question = str(raw.get("question") or qid).strip()
            answer = str(raw.get("answer") or "").strip()
        else:
            question, answer = str(qid).replace("_", " "), str(raw or "").strip()
        if answer:
            out.append({"id": str(qid), "question": question, "answer": answer})
    return out


def apply_owner_clarification(session, question, answer, build_id="", answered_at=""):
    """Записать один checkpoint-ответ как авторитетный факт следующего Task Contract."""
    session = session if isinstance(session, dict) else {}
    question = str(question or "").strip()[:1000]
    answer = str(answer or "").strip()[:6000]
    if not question or not answer:
        return ""
    stamp = answered_at or datetime.now(timezone.utc).isoformat()
    answer_id = "builder_clarification_" + hashlib.sha256(question.encode("utf-8")).hexdigest()[:12]
    session.setdefault("answers", {})[answer_id] = {
        "question": question, "answer": answer, "source": "builder_checkpoint", "answered_at": stamp}
    decisions = [x for x in (session.get("decisions") or []) if not (
        isinstance(x, dict) and x.get("type") == "builder_clarification" and
        str(x.get("question") or "") == question)]
    decisions.append({"type": "builder_clarification", "question": question, "answer": answer,
                      "build_id": str(build_id or ""), "at": stamp})
    session["decisions"] = decisions[-120:]
    session.pop("waiting_build", None)
    session.pop("building", None)
    return answer_id


def _as_text_list(value, limit=30):
    if isinstance(value, list):
        return [_clip(v, 500) for v in value[:limit] if str(v or "").strip()]
    if isinstance(value, dict):
        return [_clip({"name": k, "value": v}, 500) for k, v in list(value.items())[:limit]]
    return [_clip(value, 500)] if str(value or "").strip() else []


def _contract_profile(profile):
    """Достаточная схема входа для Task Contract без дублирования всей выборки в каждом промпте."""
    out = {"profile_error": profile.get("profile_error") or ""}
    if profile.get("workbook"):
        out["sections"] = [{"name": x.get("title"), "rows": x.get("max_row"),
                            "columns_count": x.get("max_column"), "header_row": x.get("header_row"),
                            "columns": x.get("columns") or [], "sample_rows": (x.get("sample_rows") or [])[:3]}
                           for x in (profile.get("workbook") or [])[:12]]
    elif profile.get("columns") or profile.get("sample_rows"):
        out["sections"] = [{"name": "data", "columns": profile.get("columns") or [],
                            "sample_rows": (profile.get("sample_rows") or [])[:3]}]
    elif profile.get("text_sample") is not None:
        out["sections"] = [{"name": "document", "text_sample": _clip(profile.get("text_sample"), 1800),
                            "text_source": profile.get("text_source") or ""}]
    return _compact(out)


def make_task_contract(session, blueprint, spec, profiles):
    """Единый договор задачи для интервью → стройки → приёмки → эксплуатации.

    Поля намеренно общие: никакой отраслевой логики, названий листов или ключей здесь нет.
    """
    session = session if isinstance(session, dict) else {}
    blueprint = blueprint if isinstance(blueprint, dict) else {}
    test_plan = blueprint.get("sample_test_plan") if isinstance(blueprint.get("sample_test_plan"), dict) else {}
    stages = [x for x in (blueprint.get("stages") or []) if isinstance(x, dict)]
    original = str(session.get("questionnaire_task") or session.get("goal") or "").strip()
    goal = str(blueprint.get("goal") or original).strip()
    source = session.get("source") if isinstance(session.get("source"), dict) else {}
    schedule = session.get("schedule") if isinstance(session.get("schedule"), dict) else {}
    permissions = session.get("permissions") or session.get("authority") or {}
    recipients = session.get("recipients") or []
    contract = {
        "version": 1,
        "original_request": original,
        "business_goal": goal,
        "process_name": str(blueprint.get("process_name") or session.get("client_name") or "").strip(),
        "required_result": {
            "success_criteria": _as_text_list(test_plan.get("success_criteria")),
            "planned_steps": _as_text_list(test_plan.get("steps")),
            "stage_outputs": [x.get("output") or x.get("outputs") for x in stages
                              if x.get("output") or x.get("outputs")],
            "required_artifacts": ["report.md", "report.xlsx"],
        },
        "inputs": [{"source_id": "source_%03d" % index,
                    "name": _nfc(p.get("name")), "format": p.get("extension"), "bytes": p.get("bytes"),
                    "sha256": p.get("sha256"), "profile": _contract_profile(p)}
                   for index, p in enumerate((profiles or []), 1)],
        "source_configuration": source,
        "interview": _answer_rows(session.get("answers") or {}),
        "owner_comments": session.get("comments") or [],
        "owner_decisions": session.get("decisions") or [],
        "owner_rules": session.get("rules") or [],
        "structured_rules": session.get("rules_struct") or [],
        "fields": session.get("fields") or {},
        "exceptions": session.get("exceptions") or blueprint.get("exceptions") or [],
        "permissions": permissions,
        "constraints": {
            "blueprint_gaps": blueprint.get("gaps") or [],
            "open_questions": blueprint.get("open_questions") or [],
            "confidentiality": session.get("confidentiality") or session.get("privacy") or {},
            "no_external_writes_during_acceptance": True,
        },
        "operation": {
            "schedule": schedule,
            "recipients": recipients,
            "delivery": session.get("delivery") or schedule.get("deliver") or {},
            "trigger": session.get("trigger") or schedule.get("period") or "manual",
        },
        "human_decisions": blueprint.get("open_questions") or [],
        "data_check": session.get("data_check") or {},
        "approved_blueprint": blueprint,
        "project_spec": _clip(spec, 12000),
        "authority_order": [
            "owner: прямые ответы, решения, правила и полномочия",
            "data: детерминированные факты профиля источников",
            "accepted_memory: знания, доказанные полной приёмкой",
            "qwen_hypotheses: рабочие гипотезы, которые ещё нужно проверить",
            "rejected_memory: только отрицательные уроки, не факты",
        ],
    }
    raw = json.dumps(contract, ensure_ascii=False, sort_keys=True, default=str)
    contract["sha256"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return contract


def _memory_id(entry):
    raw = "|".join(str(entry.get(k) or "") for k in ("kind", "text", "source", "attempt"))
    return "mem_" + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:14]


def _normalize_memory_entry(raw, kind="concept", status="candidate", source="qwen",
                            scope="session", attempt=0):
    raw = raw if isinstance(raw, dict) else {"text": str(raw or "")}
    text = str(raw.get("text") or raw.get("value") or raw.get("rule") or "").strip()
    if not text:
        return None
    item = {
        "text": _clip(text, 600),
        "kind": str(raw.get("kind") or kind).lower(),
        "status": str(raw.get("status") or status).lower(),
        "source": str(raw.get("source") or source)[:80],
        "scope": str(raw.get("scope") or scope).lower(),
        "evidence": _compact(raw.get("evidence") or ""),
        "confidence": raw.get("confidence", 0.5),
        "attempt": int(raw.get("attempt") or attempt or 0),
        "created_at": str(raw.get("created_at") or datetime.now(timezone.utc).isoformat()),
        "supersedes": str(raw.get("supersedes") or "")[:80],
        "rejection_reason": _clip(raw.get("rejection_reason") or "", 300),
    }
    if item["kind"] not in MEMORY_KINDS:
        item["kind"] = kind if kind in MEMORY_KINDS else "concept"
    if item["status"] not in MEMORY_STATUSES:
        item["status"] = status if status in MEMORY_STATUSES else "candidate"
    if item["scope"] not in ("attempt", "session", "process", "agent"):
        item["scope"] = scope if scope in ("attempt", "session", "process", "agent") else "session"
    try:
        item["confidence"] = max(0.0, min(1.0, float(item["confidence"])))
    except Exception:
        item["confidence"] = 0.5
    item["id"] = str(raw.get("id") or _memory_id(item))[:80]
    return item


def _merge_memory(memory, additions):
    memory = memory if isinstance(memory, dict) else {}
    entries = [x for x in (memory.get("entries") or []) if isinstance(x, dict)]
    by_id = {str(x.get("id") or _memory_id(x)): dict(x) for x in entries}
    for raw in additions or []:
        item = raw if isinstance(raw, dict) and raw.get("id") else _normalize_memory_entry(raw)
        if not item:
            continue
        old = by_id.get(item["id"])
        if old:
            old.update({k: v for k, v in item.items() if v not in (None, "", [], {})})
        else:
            by_id[item["id"]] = item
    memory["version"] = 1
    memory["entries"] = list(by_id.values())[-120:]
    memory["updated_at"] = datetime.now(timezone.utc).isoformat()
    return memory


def _initial_working_memory(session, contract_sha):
    previous = session.get("agentic_memory") if isinstance(session.get("agentic_memory"), dict) else {}
    memory = {"version": 1, "task_contract_sha256": contract_sha, "entries": []}
    if previous.get("task_contract_sha256") == contract_sha:
        memory["entries"] = [x for x in (previous.get("entries") or []) if isinstance(x, dict)][-120:]
    owner = []
    for rule in session.get("rules") or []:
        owner.append(_normalize_memory_entry(
            {"text": str(rule), "status": "verified", "source": "owner", "scope": "process",
             "confidence": 1.0, "evidence": "прямое правило владельца"}, kind="rule"))
    return _merge_memory(memory, [x for x in owner if x])


def make_task_package(session_id, sample_files, sess_dir, llm=None):
    """Единый источник контекста: интервью + blueprint + ТЗ + решения + правила + все образцы."""
    root = Path(sess_dir)
    session = _read_json(root / (session_id + ".json"), {}) or {}
    bpdoc = _read_json(root / (session_id + "_blueprint.json"), {}) or {}
    pdoc = _read_json(root / (session_id + "_build_plan.json"), {}) or {}
    chat = _read_json(root / (session_id + "_chat.json"), {}) or {}
    spec_path = root / (session_id + "_spec.md")
    spec = spec_path.read_text(encoding="utf-8", errors="replace")[:24000] if spec_path.exists() else ""
    profiles = [profile_file(p) for p in sample_files]
    blueprint = bpdoc.get("blueprint", bpdoc) if isinstance(bpdoc, dict) else {}
    task_contract = make_task_contract(session, blueprint, spec, profiles)
    llm = llm if isinstance(llm, dict) else {}
    task_contract["selected_qwen"] = {
        "family": "Qwen", "mode": "openai_compatible" if llm.get("api_key") else "extella_agent",
        "model": str(llm.get("model") or "")[:160],
        "agent_id": str(llm.get("agent_id") or "")[:160],
        "endpoint_configured": bool(llm.get("base_url")), "user_key_configured": bool(llm.get("api_key")),
    }
    task_contract["available_connectors"] = {
        "source": session.get("source") or {}, "delivery": session.get("delivery") or {},
        "schedule": session.get("schedule") or {},
    }
    task_contract["sha_scope"] = "task, owner contract, inputs, profiles, selected Qwen; runtime catalog excluded"
    contract_raw = json.dumps({k: v for k, v in task_contract.items() if k != "sha256"},
                              ensure_ascii=False, sort_keys=True, default=str)
    task_contract["sha256"] = hashlib.sha256(contract_raw.encode("utf-8")).hexdigest()
    working_memory = _initial_working_memory(session, task_contract.get("sha256", ""))
    package = {
        "contract_version": 2,
        "session_id": session_id,
        "client_name": session.get("client_name"),
        "language": session.get("language") or "ru",
        "original_request": session.get("questionnaire_task") or session.get("goal") or "",
        "interview_answers": session.get("answers") or {},
        "owner_comments": session.get("comments") or [],
        "decisions": session.get("decisions") or [],
        "rules": session.get("rules") or [],
        "fields": session.get("fields") or {},
        "permissions": session.get("permissions") or session.get("authority") or {},
        "source": session.get("source") or {},
        "schedule": session.get("schedule") or {},
        "recipients": session.get("recipients") or [],
        "data_check": session.get("data_check") or {},
        "blueprint": blueprint,
        "project_spec": spec,
        "build_plan": pdoc.get("plan", pdoc),
        "assistant_context": chat,
        "inputs": profiles,
        "task_contract": task_contract,
        "working_memory": working_memory,
        "runtime_input_contract": {
            "source_file": "путь к одному файлу ИЛИ папке-пакету; для нескольких входов обрабатывается вся папка",
            "output_dir": "отдельная папка для результатов; входные файлы нельзя изменять",
        },
        "acceptance_contract": {
            "must_use_every_sample": True,
            "must_run_on_real_samples": True,
            "must_return": ["status=success", "summary", "evidence.files_used", "evidence.acceptance_checks",
                            "report_md", "report_xlsx"],
            "no_external_writes": True,
        },
    }
    package = _compact(package)
    # Эти части уже ограничены профилировщиком/чтением файла. Общий _compact режет любую строку до
    # 1800 символов, но для Builder это как раз самая важная фактура: полный текст ТЗ и PDF-фрагмент.
    package["project_spec"] = _clip(spec, 12000)
    package["inputs"] = profiles
    package["task_contract"] = task_contract
    package["working_memory"] = working_memory
    raw = json.dumps(package, ensure_ascii=False, sort_keys=True, default=str)
    package["package_sha256"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return package


def _available_capabilities():
    """Единый реестр Extella: эксперты, плагины, модели, CSPL, устройства и блоки Композитора."""
    out = []
    try:
        got = api("/api/kv/get", {"key": "capability:registry"}, timeout=60)
        meta = json.loads(got.get("value") or "{}") if isinstance(got, dict) else {}
        if isinstance(meta, dict) and meta.get("chunks"):
            buf = ""
            for index in range(min(int(meta.get("chunks") or 0), 200)):
                chunk = api("/api/kv/get", {"key": "capability:registry:" + str(index)}, timeout=60)
                buf += str((chunk or {}).get("value") or "")
            if meta.get("enc") == "b64" and buf:
                buf = base64.b64decode(buf).decode("utf-8")
            registry = json.loads(buf) if buf else {}
        else:
            registry = meta
        for cap in (registry.get("capabilities") or [])[:500]:
            if not isinstance(cap, dict) or not cap.get("capability_id"):
                continue
            out.append({"expert": str(cap.get("capability_id")), "kind": cap.get("type"),
                        "purpose": _clip((str(cap.get("title") or "") + ": " +
                                          str(cap.get("description") or "")).strip(": "), 280),
                        "source": cap.get("source"), "surfaces": cap.get("surfaces") or [],
                        "params": _compact({k: v for k, v in cap.items() if k not in
                                            ("capability_id", "title", "description", "type")})})
    except Exception:
        pass
    # Старые установки могли ещё не собрать unified registry. Composer остаётся безопасным fallback.
    try:
        got = api("/api/kv/get", {"key": "composer:catalog", "global": True}, timeout=60)
        raw = got.get("value") if isinstance(got, dict) else None
        catalog = json.loads(raw) if raw else {}
    except Exception:
        catalog = {}
    for block in (catalog.get("blocks") or catalog.get("items") or [])[:160]:
        if not isinstance(block, dict):
            continue
        name = str(block.get("id") or block.get("expert") or block.get("name") or "").strip()
        if not name:
            continue
        if not any(x.get("expert") == name for x in out):
            out.append({"expert": name, "kind": block.get("kind") or "composer_block",
                        "purpose": _clip(block.get("what") or block.get("description") or block.get("title"), 280),
                        "source": "composer:catalog", "surfaces": ["composer", "chat", "wizard"],
                        "params": _compact(block.get("params") or block.get("defaults") or {})})
    return out


def _select_capabilities(task_package, capabilities, limit=6):
    """Передать Строителю только способности, которые похожи на текущую задачу.

    Полный каталог (десятки несвязанных экспертов) раньше занимал четверть промпта и подталкивал
    Qwen к случайному реюзу. Это не запрет возможностей: каталог остаётся источником кандидатов,
    но в конкретный BUILD_BRIEF попадают только релевантные входам и смыслу задачи записи.
    """
    package = task_package or {}
    context = json.dumps({
        "request": package.get("original_request"),
        "answers": package.get("interview_answers"),
        "rules": package.get("rules"),
        "fields": package.get("fields"),
        "blueprint": package.get("blueprint"),
        "spec": _clip(package.get("project_spec"), 5000),
    }, ensure_ascii=False, default=str).casefold()
    extensions = {str(p.get("extension") or "").lower() for p in (package.get("inputs") or [])}
    aliases = {
        ".pdf": ("pdf", "ocr", "tesseract", "document", "документ", "скан"),
        ".xlsx": ("xlsx", "excel", "spreadsheet", "таблиц"),
        ".xlsm": ("xlsm", "xlsx", "excel", "spreadsheet", "таблиц"),
        ".csv": ("csv", "spreadsheet", "таблиц"),
        ".docx": ("docx", "word", "document", "документ"),
        ".json": ("json",),
        ".xml": ("xml",),
    }
    stop = {"этот", "этого", "данные", "файл", "файла", "процесс", "задача", "нужно", "будет",
            "with", "from", "that", "this", "data", "file", "process", "task", "expert", "plugin",
            "request", "answers", "rules", "fields", "blueprint", "spec"}
    words = {w for w in re.findall(r"[a-zа-яё0-9_\-]{4,}", context) if w not in stop}
    ranked = []
    for index, cap in enumerate(capabilities or []):
        if not isinstance(cap, dict):
            continue
        hay = json.dumps(cap, ensure_ascii=False, default=str).casefold()
        score = 0
        for ext in extensions:
            if any(alias in hay for alias in aliases.get(ext, (ext.lstrip("."),))):
                score += 12
        score += min(10, sum(1 for word in words if word in hay))
        if score:
            ranked.append((score, -index, cap))
    ranked.sort(reverse=True, key=lambda item: (item[0], item[1]))
    return [cap for _, _, cap in ranked[:max(1, int(limit))]]


def _attach_capabilities(package):
    package = dict(package or {})
    selected = _select_capabilities(package, _available_capabilities())
    package["available_plugins_and_experts"] = selected
    if isinstance(package.get("task_contract"), dict):
        package["task_contract"] = dict(package["task_contract"])
        package["task_contract"]["relevant_capabilities"] = selected
    raw = json.dumps({k: v for k, v in package.items() if k != "package_sha256"},
                     ensure_ascii=False, sort_keys=True, default=str)
    package["package_sha256"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return package


def _builder_inputs(profiles):
    """Фактура файлов без лишнего объёма: схема и репрезентативная выборка, не весь профайл."""
    out = []
    for raw in profiles or []:
        item = dict(raw) if isinstance(raw, dict) else {"profile": _clip(raw, 1000)}
        if isinstance(item.get("text_sample"), str):
            item["text_sample"] = _clip(item["text_sample"], 6000)
        sheets = []
        for sheet in item.get("workbook") or []:
            sh = dict(sheet)
            sh["sample_rows"] = (sh.get("sample_rows") or [])[:12]
            sheets.append(sh)
        if sheets:
            item["workbook"] = sheets
        if isinstance(item.get("sample_rows"), list):
            item["sample_rows"] = item["sample_rows"][:12]
        out.append(item)
    return out


def _builder_brief(task_package):
    """Авторитетный контекст для кодогенерации без старого compiler-plan и стенограммы чата."""
    p = task_package or {}
    brief = {
        "brief_version": 3,
        "language": p.get("language") or "ru",
        "task_contract": p.get("task_contract"),
        "upc_step": p.get("upc_step"),
        "source_model": p.get("source_model"),
        "execution_decision": p.get("execution_decision"),
        "process_source_context": {
            "manifest": p.get("global_source_manifest"),
            "accepted_source_model": p.get("global_source_model"),
            "boundary": ("Это справочная модель всего процесса. Физически обрабатывай только "
                         "input_profiles текущего шага; соседние источники принадлежат runtime."),
        },
        "working_memory": p.get("working_memory"),
        "goal": {"process_name": p.get("client_name"), "original_request": p.get("original_request")},
        "owner_contract": {
            "interview_answers": p.get("interview_answers"),
            "owner_comments": p.get("owner_comments"),
            "decisions": p.get("decisions"),
            "rules": p.get("rules"),
            "fields": p.get("fields"),
            "permissions": p.get("permissions"),
            "source": p.get("source"),
            "schedule": p.get("schedule"),
            "recipients": p.get("recipients"),
            "data_check": p.get("data_check"),
        },
        "approved_design": {
            "blueprint": p.get("blueprint"),
            "project_spec": _clip(p.get("project_spec"), 7000),
        },
        "input_profiles": _builder_inputs(p.get("inputs") or []),
        "relevant_capabilities": _select_capabilities(
            p, p.get("available_plugins_and_experts") or [], limit=6),
        "runtime_contract": p.get("runtime_input_contract"),
        "acceptance_contract": p.get("acceptance_contract"),
        "authority_order": [
            "owner_contract: прямые ответы, решения, правила и полномочия владельца",
            "approved_design: согласованный blueprint и ТЗ",
            "input_profiles: только структура/формат фактических данных, не источник бизнес-правил",
        ],
    }

    def prune(value):
        if isinstance(value, dict):
            return {k: v for k, raw in value.items() if (v := prune(raw)) not in (None, "", [], {})}
        if isinstance(value, list):
            return [v for raw in value if (v := prune(raw)) not in (None, "", [], {})]
        return value

    clean = prune(brief)
    raw = json.dumps(clean, ensure_ascii=False, sort_keys=True, default=str)
    clean["brief_sha256"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return clean


def _source_inventory(profiles):
    """Детерминированная фактура источников. Семантику поверх неё определяет Qwen, не порядок листов."""
    out = []
    for source_index, profile in enumerate(profiles or [], 1):
        source_id = "source_%03d" % source_index
        item = {"id": source_id, "name": _nfc(profile.get("name")), "format": profile.get("extension"),
                "bytes": profile.get("bytes"), "sha256": profile.get("sha256")}
        if profile.get("workbook"):
            item["sections"] = [{"id": "%s_section_%03d" % (source_id, section_index),
                                 "name": _nfc(sh.get("title")), "rows": sh.get("max_row"),
                                 "columns_count": sh.get("max_column"),
                                 "header_row": sh.get("header_row"),
                                 "columns": sh.get("columns") or [],
                                 "sample_rows": (sh.get("sample_rows") or [])[:12]}
                                for section_index, sh in enumerate(profile.get("workbook") or [], 1)]
        elif profile.get("columns") or profile.get("sample_rows"):
            item["sections"] = [{"id": source_id + "_section_001", "name": "data",
                                 "columns": profile.get("columns") or [],
                                 "sample_rows": (profile.get("sample_rows") or [])[:12]}]
        elif profile.get("text_sample"):
            item["sections"] = [{"id": source_id + "_section_001", "name": "document",
                                 "text_sample": _clip(profile.get("text_sample"), 6000),
                                 "text_source": profile.get("text_source")}]
        if profile.get("profile_error"):
            item["profile_error"] = profile.get("profile_error")
        out.append(item)
    return out


def _source_model_prompt(package, previous_error=""):
    contract = package.get("task_contract") or {}
    inventory = _source_inventory(package.get("inputs") or [])
    caps = package.get("available_plugins_and_experts") or []
    correction = ("\n\nПРЕДЫДУЩАЯ МОДЕЛЬ НЕ ПРОШЛА ДЕТЕРМИНИРОВАННУЮ ПРОВЕРКУ:\n" +
                  _clip(previous_error, 1200)) if previous_error else ""
    return ("Ты — архитектор источников универсального Builder Extella. До написания кода построй "
            "явную модель фактических входов и выбери способ выполнения САМОЙ ЗАДАЧИ. На этом этапе "
            "технического имени будущего эксперта ещё нет: не проектируй эксперта по заранее заданному "
            "типу, сначала реши, какой маршрут нужен задаче. Не угадывай роль файла/листа по позиции "
            "или одному названию: опирайся на колонки, выборку, Task Contract и приводи evidence. "
            "Идентичность задаёт харнесс: для КАЖДОГО физического входа верни ровно его source_id, "
            "для раздела — section_id, а operations.inputs заполняй ТОЛЬКО этими ID. Не объединяй "
            "несколько файлов под придуманным логическим именем и не создавай новые имена входов. "
            "Если несколько трактовок меняют бизнес-результат, status=need_human и ОДИН конкретный вопрос. "
            "VERIFIED PROCESS MEMORY и owner_answer_for_step уже подтверждены владельцем: применяй их "
            "как обязательные правила и не задавай тот же вопрос повторно. "
            "Нормализацию идентификаторов предлагай только с доказательством; сомнительную помечай "
            "requires_owner=true. reuse/compose допустимы только для явно подходящей capability из списка. "
            "acquire означает лишь предложение недостающей способности, не установку. "
            "execution_route выбирай так: reuse — есть точная проверенная capability; build — задача "
            "детерминирована и решается кодом; llm_worker — смысловое решение нужно принимать при каждом "
            "запуске; split_step — в одном шаге несколько независимо проверяемых действий; replan — неверна "
            "сама граница/последовательность шагов; need_human — только реальная бизнес-неоднозначность или "
            "недостающий доступ; acquire — отсутствует внешняя способность. Не выбирай need_human из-за "
            "технической трудности.\n\n"
            "Верни ТОЛЬКО JSON:\n"
            '{"status":"ready|need_human|acquire","strategy":"reuse|compose|build|holistic_build|acquire|need_human",'
            '"execution_route":"reuse|build|llm_worker|split_step|replan|need_human|acquire",'
            '"decision_record":{"goal":"что должно быть достигнуто","basis":["факт из ТЗ/входов"],'
            '"next_action":"что материализовать после решения"},'
            '"reason":"...","sources":[{"source_id":"source_001","name":"точное имя входа","role":"...","entities":["..."],'
            '"sections":[{"section_id":"source_001_section_001","name":"точное имя листа/раздела","role":"...","entities":["..."],'
            '"identifier_fields":["..."],"date_fields":["..."],"numeric_fields":["..."],'
            '"evidence":["колонка/фрагмент"]}],"limitations":["..."],"evidence":["..."]}],'
            '"operations":[{"name":"...","inputs":["source_001","source_002_section_001"],"join_keys":["..."],'
            '"normalizations":[{"field":"...","method":"...","evidence":"...","requires_owner":false}],'
            '"evidence":["..."]}],"acceptance_criteria":["наблюдаемый критерий"],'
            '"selected_capabilities":["expert id"],"missing_capability":"","question":""}.\n\n'
            "TASK CONTRACT:\n" + json.dumps(_compact(contract), ensure_ascii=False, default=str) +
            "\n\nSOURCE INVENTORY:\n" + json.dumps(inventory, ensure_ascii=False, default=str) +
            "\n\nRELEVANT CAPABILITIES:\n" + json.dumps(_compact(caps), ensure_ascii=False, default=str) +
            correction)


def _normalize_source_model(raw, package):
    """Fail-closed контракт Source Model: только фактические файлы/листы и доказанные операции."""
    if not isinstance(raw, dict):
        return {"ok": False, "why": "Qwen не вернула объект Source Model"}
    raw = dict(raw)
    contract_repairs = []
    # Some Qwen-compatible endpoints wrap the requested object even with JSON mode enabled. The
    # nested object still goes through every deterministic identity/evidence/operation check below.
    if not any(key in raw for key in ("sources", "operations", "strategy", "question")):
        for wrapper in ("source_model", "model", "result"):
            nested = raw.get(wrapper)
            if isinstance(nested, dict) and any(
                    key in nested for key in ("sources", "operations", "strategy", "question")):
                raw = dict(nested)
                contract_repairs.append("unwrapped:" + wrapper)
                break
    status = str(raw.get("status") or "").lower()
    strategy = str(raw.get("strategy") or "").lower()
    execution_route = str(raw.get("execution_route") or raw.get("route") or "").lower()
    question_hint = str(raw.get("question") or "").strip()
    missing_hint = str(raw.get("missing_capability") or "").strip()
    has_model_body = (isinstance(raw.get("sources"), list) and bool(raw.get("sources")) and
                      isinstance(raw.get("operations"), list) and bool(raw.get("operations")))
    if status in ("ok", "success", "complete", "completed") and has_model_body:
        status = "ready"
        contract_repairs.append("status:" + str(raw.get("status")) + "→ready")
    if not status:
        if strategy == "acquire" or (missing_hint and question_hint):
            status = "acquire"
        elif strategy == "need_human" or (question_hint and not has_model_body):
            status = "need_human"
        elif has_model_body:
            status = "ready"
        if status:
            contract_repairs.append("status:<missing>→" + status)
    if strategy in ("generate", "create", "code"):
        strategy = "build"
        contract_repairs.append("strategy:" + str(raw.get("strategy")) + "→build")
    if not strategy:
        if status in ("need_human", "acquire"):
            strategy = status
        elif status == "ready" and has_model_body:
            strategy = "build"
        if strategy:
            contract_repairs.append("strategy:<missing>→" + strategy)
    if status not in SOURCE_STATUSES:
        return {"ok": False, "why": "неизвестный status Source Model: " + (status or "<missing>")}
    if strategy not in BUILD_STRATEGIES:
        return {"ok": False, "why": "неизвестная strategy Source Model: " + (strategy or "<missing>")}
    if status == "ready" and strategy in ("acquire", "need_human"):
        return {"ok": False, "why": "ready несовместим со strategy=" + strategy}
    if status == "need_human" and strategy != "need_human":
        return {"ok": False, "why": "need_human требует одноимённую strategy"}
    if status == "acquire" and strategy != "acquire":
        return {"ok": False, "why": "acquire требует одноимённую strategy"}
    route_aliases = {
        "generate": "build", "create": "build", "code": "build", "deterministic": "build",
        "reason": "llm_worker", "reasoning": "llm_worker", "semantic": "llm_worker",
        "human": "need_human", "split": "split_step", "compose": "reuse",
        "holistic_build": "build",
    }
    execution_route = route_aliases.get(execution_route, execution_route)
    if not execution_route:
        if status in ("need_human", "acquire"):
            execution_route = status
        elif strategy in ("reuse", "compose"):
            execution_route = "reuse"
        else:
            execution_route = "build"
        contract_repairs.append("execution_route:<missing>→" + execution_route)
    if execution_route not in EXECUTION_ROUTES:
        return {"ok": False, "why": "неизвестный execution_route Source Model: " + execution_route}
    if status in ("need_human", "acquire") and execution_route != status:
        return {"ok": False, "why": "status=%s несовместим с execution_route=%s" %
                                     (status, execution_route)}
    if status == "ready" and execution_route in ("need_human", "acquire"):
        return {"ok": False, "why": "ready несовместим с execution_route=" + execution_route}
    question = str(raw.get("question") or "").strip()
    if status == "need_human" and not question:
        return {"ok": False, "why": "need_human без конкретного вопроса владельцу"}
    missing_capability = str(raw.get("missing_capability") or "").strip()
    if status == "acquire" and not (missing_capability and question):
        return {"ok": False, "why": "acquire без названия способности и вопроса владельцу"}

    inventory = _source_inventory(package.get("inputs") or [])
    actual_by_id = {_ref_key(x.get("id")): x for x in inventory if x.get("id")}
    actual_by_name = {_ref_key(x.get("name"), basename=True): x for x in inventory if x.get("name")}
    clean_sources, covered = [], set()
    for source in raw.get("sources") if isinstance(raw.get("sources"), list) else []:
        if not isinstance(source, dict):
            continue
        source_ref = source.get("source_id") or source.get("id")
        fact = actual_by_id.get(_ref_key(source_ref)) if source_ref else None
        if not fact:
            fact = actual_by_name.get(_ref_key(source.get("name"), basename=True))
        if not fact:
            bad_ref = source_ref or source.get("name")
            return {"ok": False, "why": "Source Model ссылается на неизвестный вход: " + str(bad_ref)}
        source_id = str(fact.get("id") or "")
        covered.add(source_id)
        actual_sections = fact.get("sections") or []
        sections_by_id = {_ref_key(x.get("id")): x for x in actual_sections if x.get("id")}
        sections_by_name = {_ref_key(x.get("name")): x for x in actual_sections if x.get("name")}
        sections = []
        for section in source.get("sections") if isinstance(source.get("sections"), list) else []:
            if not isinstance(section, dict):
                continue
            section_ref = section.get("section_id") or section.get("id")
            section_fact = sections_by_id.get(_ref_key(section_ref)) if section_ref else None
            if not section_fact:
                section_fact = sections_by_name.get(_ref_key(section.get("name")))
            if actual_sections and not section_fact:
                bad_ref = section_ref or section.get("name")
                return {"ok": False, "why": "Source Model ссылается на неизвестный лист/раздел: " + str(bad_ref)}
            sname = _nfc((section_fact or {}).get("name") or section.get("name")).strip()
            sections.append({
                "section_id": str((section_fact or {}).get("id") or section_ref or ""),
                "name": sname,
                "role": _clip(section.get("role"), 300),
                "entities": _as_text_list(section.get("entities"), 20),
                "identifier_fields": _as_text_list(section.get("identifier_fields"), 30),
                "date_fields": _as_text_list(section.get("date_fields"), 30),
                "numeric_fields": _as_text_list(section.get("numeric_fields"), 30),
                "evidence": _as_text_list(section.get("evidence"), 20),
            })
        evidence = _as_text_list(source.get("evidence"), 20)
        if status == "ready" and not str(source.get("role") or "").strip():
            return {"ok": False, "why": "для входа %s не определена роль" % fact.get("name")}
        covered_sections = {str(x.get("section_id") or "") for x in sections if x.get("section_id")}
        required_sections = {str(x.get("id") or "") for x in actual_sections if x.get("id")}
        if status == "ready" and required_sections - covered_sections:
            return {"ok": False, "why": "для входа %s не описаны разделы: %s" %
                    (fact.get("name"), ", ".join(
                        str(x.get("name")) for x in actual_sections if x.get("id") in required_sections - covered_sections))}
        if status == "ready" and not evidence and not any(x.get("evidence") for x in sections):
            return {"ok": False, "why": "для входа %s нет evidence роли" % fact.get("name")}
        source_entities = _as_text_list(source.get("entities"), 30)
        if status == "ready" and not source_entities and not any(x.get("entities") for x in sections):
            return {"ok": False, "why": "для входа %s не определены бизнес-сущности" % fact.get("name")}
        clean_sources.append({"source_id": source_id, "name": _nfc(fact.get("name")),
                              "role": _clip(source.get("role"), 300),
                              "entities": source_entities,
                              "sections": sections, "limitations": _as_text_list(source.get("limitations"), 20),
                              "evidence": evidence})
    if status == "ready" and {str(x.get("id")) for x in inventory} - covered:
        return {"ok": False, "why": "Source Model не описала обязательные входы: " +
                ", ".join(str(x.get("name")) for x in inventory if x.get("id") not in covered)}

    ref_targets = {}

    def add_ref(ref, canonical_id, owner_id):
        key = _ref_key(ref)
        if key:
            ref_targets.setdefault(key, []).append((canonical_id, owner_id))

    for source in clean_sources:
        source_id = str(source.get("source_id") or "")
        source_name = str(source.get("name") or "")
        add_ref(source_id, source_id, source_id)
        add_ref(source_name, source_id, source_id)
        for section in source.get("sections") or []:
            section_id = str(section.get("section_id") or "")
            section_name = str(section.get("name") or "")
            add_ref(section_id, section_id, source_id)
            add_ref(section_name, section_id, source_id)
            add_ref(source_id + "/" + section_id, section_id, source_id)
            add_ref(source_name + "/" + section_name, section_id, source_id)
            add_ref(source_name + ":" + section_name, section_id, source_id)
    operations, operation_sources = [], set()
    for operation in raw.get("operations") if isinstance(raw.get("operations"), list) else []:
        if not isinstance(operation, dict) or not str(operation.get("name") or "").strip():
            continue
        normalizations = []
        for norm in operation.get("normalizations") if isinstance(operation.get("normalizations"), list) else []:
            if not isinstance(norm, dict):
                continue
            normalized = {"field": _clip(norm.get("field"), 200),
                          "method": _clip(norm.get("method"), 300),
                          "evidence": _clip(norm.get("evidence"), 400),
                          "requires_owner": norm.get("requires_owner") is True}
            if status == "ready" and (not normalized["field"] or not normalized["method"] or
                                       not normalized["evidence"]):
                return {"ok": False, "why": "нормализация без field/method/evidence"}
            normalizations.append(normalized)
        raw_inputs = _as_text_list(operation.get("inputs"), 30)
        op_inputs, unknown_refs, ambiguous_refs = [], [], []
        for item in raw_inputs:
            candidates = list(dict.fromkeys(ref_targets.get(_ref_key(item), [])))
            if not candidates:
                unknown_refs.append(item)
            elif len(candidates) > 1:
                ambiguous_refs.append(item)
            else:
                canonical_id, owner_id = candidates[0]
                op_inputs.append(canonical_id)
                operation_sources.add(owner_id)
        if status == "ready" and unknown_refs:
            return {"ok": False, "why": "операция ссылается на неизвестные входы/разделы: " +
                    ", ".join(str(x) for x in unknown_refs)}
        if status == "ready" and ambiguous_refs:
            return {"ok": False, "why": "операция использует неоднозначное имя раздела; укажите section_id: " +
                    ", ".join(str(x) for x in ambiguous_refs)}
        operations.append({"name": _clip(operation.get("name"), 300),
                           "inputs": op_inputs,
                           "join_keys": _as_text_list(operation.get("join_keys"), 30),
                           "normalizations": normalizations,
                           "evidence": _as_text_list(operation.get("evidence"), 20)})
    if status == "ready" and not operations:
        return {"ok": False, "why": "Source Model ready, но не описала ни одной операции"}
    if status == "ready" and {str(x.get("source_id") or "") for x in clean_sources} - operation_sources:
        return {"ok": False, "why": "операции Source Model используют не все обязательные источники"}
    if status == "ready" and any(not x.get("evidence") for x in operations):
        return {"ok": False, "why": "операция Source Model не имеет evidence из Task Contract/данных"}
    ambiguous = [n for op in operations for n in op.get("normalizations") or [] if n.get("requires_owner")]
    if status == "ready" and ambiguous:
        return {"ok": False, "why": "неоднозначная нормализация требует need_human"}
    criteria = _as_text_list(raw.get("acceptance_criteria"), 30)
    if status == "ready" and not criteria:
        return {"ok": False, "why": "Source Model ready без наблюдаемых критериев приёмки"}

    available = {str(x.get("expert") or "") for x in (package.get("available_plugins_and_experts") or [])
                 if isinstance(x, dict) and str(x.get("expert") or "")}
    selected = _as_text_list(raw.get("selected_capabilities"), 12)
    unknown = [x for x in selected if x not in available]
    if unknown:
        return {"ok": False, "why": "выбраны неизвестные capabilities: " + ", ".join(unknown)}
    if status == "ready" and strategy in ("reuse", "compose") and not selected:
        return {"ok": False, "why": strategy + " требует хотя бы одну доступную capability"}
    if status == "ready" and execution_route == "reuse" and not selected:
        return {"ok": False, "why": "execution_route=reuse требует выбранную проверенную capability"}
    decision_raw = raw.get("decision_record") if isinstance(raw.get("decision_record"), dict) else {}
    decision_record = {
        "goal": _clip(decision_raw.get("goal") or
                      ((package.get("task_contract") or {}).get("business_goal") if
                       isinstance(package.get("task_contract"), dict) else "") or
                      package.get("original_request"), 500),
        "route": execution_route,
        "basis": _as_text_list(decision_raw.get("basis"), 12) or
                 _as_text_list(raw.get("reason"), 1),
        "next_action": _clip(decision_raw.get("next_action") or {
            "reuse": "материализовать единый wrapper над выбранной capability",
            "build": "создать детерминированного исполняемого эксперта",
            "llm_worker": "создать исполняемого эксперта с Qwen внутри runtime",
            "split_step": "разделить шаг на независимо проверяемые подшаги",
            "replan": "перестроить границы графа до создания эксперта",
            "need_human": "получить один ответ владельца",
            "acquire": "предложить недостающую способность владельцу",
        }.get(execution_route, "остановиться безопасно"), 500),
    }
    model = {"version": 1, "status": status, "strategy": strategy,
             "execution_route": execution_route, "decision_record": decision_record,
             "reason": _clip(raw.get("reason"), 600), "sources": clean_sources,
             "operations": operations,
             "acceptance_criteria": criteria,
             "selected_capabilities": selected,
             "missing_capability": missing_capability, "question": question,
             "contract_repairs": contract_repairs}
    body = json.dumps(model, ensure_ascii=False, sort_keys=True, default=str)
    model["sha256"] = hashlib.sha256(body.encode("utf-8")).hexdigest()
    return {"ok": True, "model": model}


def _neutral_source_model(package, previous_error=""):
    """Identity-only recovery when Qwen returned no parseable Source Model at all.

    This is deliberately narrower than semantic recovery: exact source/section ids come from the
    profiler, every physical input stays mandatory, and no join key or normalization is invented.
    The downstream holistic Qwen worker still receives the full Task Contract and acceptance must
    prove the business result.  A *substantive but invalid* Qwen model never enters this fallback.
    """
    inventory = _source_inventory((package or {}).get("inputs") or [])
    if not inventory:
        return {"ok": False, "why": previous_error or "нет фактических входов для Source Model"}
    sources = []
    for item in inventory:
        sections = []
        for section in item.get("sections") or []:
            columns = [str(x) for x in (section.get("columns") or []) if str(x)]
            section_evidence = (["фактические колонки: " + ", ".join(columns[:20])] if columns else
                                ["фактический раздел профиля: " + str(section.get("id") or section.get("name"))])
            sections.append({
                "section_id": section.get("id"), "name": section.get("name"),
                "role": "фактический раздел обязательного входа; бизнес-смысл берётся из Task Contract",
                "entities": ["записи обязательного входа"], "identifier_fields": [],
                "date_fields": [], "numeric_fields": [], "evidence": section_evidence,
            })
        sources.append({
            "source_id": item.get("id"), "name": item.get("name"),
            "role": "обязательный физический вход из утверждённого Task Contract",
            "entities": ["данные обязательного входа"], "sections": sections,
            "limitations": ["семантическая роль не угадана; её определяет целостный Qwen-исполнитель, а проверяет приёмка"],
            "evidence": ["идентичность зафиксирована профилем: format=%s sha256=%s" %
                         (item.get("format") or "unknown", item.get("sha256") or "unknown")],
        })
    contract = (package or {}).get("task_contract") if isinstance((package or {}).get("task_contract"), dict) else {}
    required = contract.get("required_result") if isinstance(contract.get("required_result"), dict) else {}
    criteria = _as_text_list(required.get("success_criteria"), 20)
    if not criteria:
        criteria = ["все фактические входы прочитаны и перечислены в evidence результата",
                    "итоговый результат соответствует утверждённому Task Contract"]
    raw = {
        "status": "ready", "strategy": "holistic_build",
        "reason": "Qwen не вернула JSON Source Model; харнесс зафиксировал только доказанные идентичности входов без смысловых догадок",
        "sources": sources,
        "operations": [{"name": "Выполнить утверждённый Task Contract целиком",
                        "inputs": [str(x.get("id")) for x in inventory if x.get("id")],
                        "join_keys": [], "normalizations": [],
                        "evidence": ["утверждённый Task Contract и фактический профиль всех входов"]}],
        "acceptance_criteria": criteria, "selected_capabilities": [],
        "missing_capability": "", "question": "",
    }
    checked = _normalize_source_model(raw, package)
    if checked.get("ok"):
        model = checked["model"]
        model.setdefault("contract_repairs", []).append("fallback:deterministic_identity_after_empty_qwen")
        model["fallback_reason"] = _clip(previous_error, 500)
        body = json.dumps({k: v for k, v in model.items() if k != "sha256"},
                          ensure_ascii=False, sort_keys=True, default=str)
        model["sha256"] = hashlib.sha256(body.encode("utf-8")).hexdigest()
    return checked


def build_source_model(package, llm, max_tries=2, progress=None):
    """Qwen строит Source Model, детерминированный валидатор не позволяет ей выдумать вход/лист."""
    last_error = ""
    tries = max(1, min(int(max_tries or 1), 3))
    for attempt in range(1, tries + 1):
        if progress:
            progress(attempt, tries, "running", last_error)
        raw = _llm_json(llm, _source_model_prompt(package, last_error), max_tokens=3800, timeout=260)
        checked = _normalize_source_model(raw, package)
        if checked.get("ok"):
            if progress:
                progress(attempt, tries, "success", "")
            return checked
        last_error = checked.get("why") or "невалидная Source Model"
        # No object at all means the provider used its response budget for reasoning and emitted no
        # final JSON. Repeating the same long prompt costs minutes and usually reproduces the same
        # shape. Identity-only recovery is safe here because it invents no semantic mapping.
        if not raw:
            fallback = _neutral_source_model(package, last_error)
            if fallback.get("ok"):
                if progress:
                    progress(attempt, tries, "success", "использована доказанная identity-only Source Model")
                return fallback
        if progress:
            progress(attempt, tries, "retry" if attempt < tries else "failed", last_error)
    return {"ok": False, "why": last_error or "Source Model не построена"}


def _source_memory(model):
    additions = []
    for source in (model or {}).get("sources") or []:
        role = str(source.get("role") or "").strip()
        if role:
            additions.append(_normalize_memory_entry({
                "text": "Источник %s: %s" % (source.get("name"), role),
                "status": "candidate", "source": "source_model", "scope": "process",
                "evidence": source.get("evidence") or [x.get("evidence") for x in source.get("sections") or []],
                "confidence": 0.65}, kind="concept"))
    for op in (model or {}).get("operations") or []:
        additions.append(_normalize_memory_entry({
            "text": "Операция процесса: %s; входы: %s" %
                    (op.get("name"), ", ".join(str(x) for x in (op.get("inputs") or []))),
            "status": "candidate", "source": "source_model", "scope": "process",
            "evidence": op.get("evidence"), "confidence": 0.6}, kind="concept"))
        for norm in op.get("normalizations") or []:
            additions.append(_normalize_memory_entry({
                "text": "Для операции «%s» проверить преобразование %s: %s" %
                        (op.get("name"), norm.get("field"), norm.get("method")),
                "status": "candidate", "source": "source_model", "scope": "process",
                "evidence": norm.get("evidence"), "confidence": 0.55}, kind="rule"))
    for capability in ((model or {}).get("selected_capabilities") or []) if (model or {}).get("strategy") in ("reuse", "compose") else []:
        additions.append(_normalize_memory_entry({
            "text": "Проверенная capability процесса: " + str(capability),
            "status": "candidate", "source": "source_model", "scope": "agent",
            "evidence": "должна быть подтверждена полным контрольным прогоном",
            "confidence": 0.5}, kind="concept"))
    return [x for x in additions if x]


def _apply_upc_step(package, step_contract):
    """Narrow the full owner contract to one graph step without losing source/permission facts."""
    if not isinstance(step_contract, dict):
        return package
    package["upc_step"] = _compact(step_contract)
    contract = copy.deepcopy(package.get("task_contract") or {})
    input_contract = (step_contract.get("input_contract")
                      if isinstance(step_contract.get("input_contract"), dict) else {})
    scope = str(input_contract.get("scope") or "atomic_step")
    global_inputs = copy.deepcopy(contract.get("inputs") or [])
    local_inputs = [{"source_id": "source_%03d" % index,
                     "name": _nfc(profile.get("name")), "format": profile.get("extension"),
                     "bytes": profile.get("bytes"), "sha256": profile.get("sha256"),
                     "profile": _contract_profile(profile)}
                    for index, profile in enumerate(package.get("inputs") or [], 1)]
    contract["business_goal"] = str(step_contract.get("purpose") or step_contract.get("title") or "")
    contract["process_business_goal"] = str(
        input_contract.get("parent_business_goal") or contract.get("business_goal") or "")
    contract["upc_scope"] = scope
    contract["process_inputs"] = global_inputs
    contract["inputs"] = local_inputs
    contract["process_source_manifest"] = copy.deepcopy(
        input_contract.get("process_source_manifest") or package.get("global_source_manifest") or [])
    acceptance = step_contract.get("acceptance") if isinstance(step_contract.get("acceptance"), dict) else {}
    output_contract = (step_contract.get("output_contract")
                       if isinstance(step_contract.get("output_contract"), dict) else {})
    required_artifacts = canonical_artifact_requirements(
        acceptance.get("required_artifacts"), output_contract.get("artifacts"),
        ensure_step_result=True)
    contract["required_result"] = {
        "success_criteria": ((acceptance.get("deterministic_checks") or []) +
                             (acceptance.get("semantic_criteria") or [])),
        "planned_steps": [step_contract.get("purpose") or step_contract.get("title")],
        "stage_outputs": output_contract.get("artifacts") or output_contract.get("postconditions") or [],
        "required_artifacts": required_artifacts,
    }
    contract["upc_step_id"] = step_contract.get("id")
    contract["upc_step_version"] = step_contract.get("version")
    contract["permissions"] = step_contract.get("permissions") or {}
    contract["verified_process_memory"] = step_contract.get("process_memory") or []
    human_gate = step_contract.get("human_gate") if isinstance(step_contract.get("human_gate"), dict) else {}
    if human_gate.get("answer"):
        contract["owner_answer_for_step"] = human_gate.get("answer")
    package["runtime_input_contract"] = {
        "source_file": "file or dependency bundle selected by the UPC runtime",
        "output_dir": "isolated directory for this step version; inputs are immutable",
        "build_validation": "dangerous effects are preview-only while the expert is being accepted",
        "scope": scope,
        "siblings_managed_by_runtime": scope == "map_partition",
    }
    package["acceptance_contract"] = {
        "must_use_every_sample": not bool(step_contract.get("dependencies")),
        "must_run_on_real_samples": True,
        "must_return": ["status=success", "summary", "evidence.acceptance_checks", "artifacts"],
        "required_artifacts": required_artifacts,
        "no_external_writes": True,
        "step_result_artifact": True,
    }
    contract_raw = json.dumps({k: v for k, v in contract.items() if k != "sha256"},
                              ensure_ascii=False, sort_keys=True, default=str)
    contract["sha256"] = hashlib.sha256(contract_raw.encode("utf-8")).hexdigest()
    package["task_contract"] = contract
    return _refresh_package(package)


def _profile_for_local_path(path, root_profiles):
    """Reuse the expensive root profile whenever content is already known."""
    path = Path(str(path))
    sha = _sha256(path) if path.is_file() else ""
    by_sha = {str(x.get("sha256") or ""): x for x in root_profiles or [] if x.get("sha256")}
    by_name = {_ref_key(x.get("name"), basename=True): x for x in root_profiles or [] if x.get("name")}
    found = by_sha.get(sha) or by_name.get(_ref_key(path.name, basename=True))
    return copy.deepcopy(found) if found else profile_file(path)


def derive_step_context(prepared_context, sample_files, step_contract):
    """Derive a per-step context from the one authoritative process Source Model.

    Profiling and Source Model reasoning used to run again for every step.  Besides being slow, that
    let a scoped PDF child reinterpret the whole Excel/PDF business task and ask the owner for files
    assigned to sibling nodes.  A step now receives only its physical/dependency inputs, while the
    global manifest remains visible as runtime-owned context.  No LLM call is needed for this view.
    """
    prepared_context = prepared_context if isinstance(prepared_context, dict) else {}
    if not prepared_context.get("ok"):
        return copy.deepcopy(prepared_context)
    root_package = copy.deepcopy(prepared_context.get("package") or {})
    root_profiles = list(root_package.get("inputs") or [])
    global_inventory = _source_inventory(root_profiles)
    root_package["global_source_manifest"] = [
        {"source_id": row.get("id"), "name": row.get("name"), "format": row.get("format"),
         "bytes": row.get("bytes"), "sha256": row.get("sha256")}
        for row in global_inventory]
    root_package["global_source_model"] = copy.deepcopy(prepared_context.get("source_model") or {})
    try:
        local_profiles = [_profile_for_local_path(path, root_profiles) for path in (sample_files or [])]
    except Exception as exc:
        return {"ok": False, "why": "не удалось профилировать вход шага: " + _clip(exc, 400),
                "package": root_package,
                "source_model": {"status": "error", "reason": _clip(exc, 400)},
                "source_memory_ids": []}
    root_package["inputs"] = local_profiles
    root_package = _apply_upc_step(root_package, step_contract)
    checked = _neutral_source_model(root_package, "derived from accepted process source model")
    if not checked.get("ok"):
        return {"ok": False, "why": checked.get("why") or "локальная Source Model не построена",
                "package": root_package,
                "source_model": {"status": "error", "reason": checked.get("why") or ""},
                "source_memory_ids": []}
    model = checked["model"]
    scope = str((((step_contract or {}).get("input_contract") or {}).get("scope")) or "atomic_step")
    model["strategy"] = "build"
    model["reason"] = ("Локальный вид выведен из принятой глобальной Source Model; соседние входы "
                       "принадлежат другим узлам runtime и не являются отсутствующими данными.")
    implementation = ((step_contract or {}).get("implementation")
                      if isinstance((step_contract or {}).get("implementation"), dict) else {})
    mode = str(implementation.get("mode") or "generate").casefold()
    execution_role = str((step_contract or {}).get("execution_role") or "").casefold()
    capability_ref = str(implementation.get("capability_ref") or "")
    available_ids = {str(x.get("expert") or "") for x in
                     (root_package.get("available_plugins_and_experts") or []) if isinstance(x, dict)}
    if mode == "reuse" and capability_ref and capability_ref in available_ids:
        route, selected = "reuse", [capability_ref]
    elif mode == "llm_worker" or execution_role == "reasoning":
        route, selected = "llm_worker", []
    else:
        route, selected = "build", []
    model["execution_route"] = route
    model["selected_capabilities"] = selected
    model["decision_record"] = {
        "goal": str((step_contract or {}).get("purpose") or (step_contract or {}).get("title") or ""),
        "route": route,
        "basis": ["UPC implementation.mode=%s" % mode,
                  "execution_role=%s" % (execution_role or "data")],
        "next_action": ({
            "reuse": "материализовать wrapper над " + capability_ref,
            "llm_worker": "создать runtime Qwen worker для смыслового решения",
            "build": "создать детерминированного code-expert",
        })[route],
    }
    model.setdefault("contract_repairs", []).append("derived:process_source_model→" + scope)
    if scope == "map_partition":
        ids = [str(row.get("source_id")) for row in model.get("sources") or [] if row.get("source_id")]
        model["operations"] = [{
            "name": "Извлечь и нормализовать канонические факты объявленной партии",
            "inputs": ids, "join_keys": [], "normalizations": [],
            "evidence": ["scoped source_refs и принятая глобальная Source Model"]}]
        model["acceptance_criteria"] = [
            "каждый локальный вход представлен фактами и provenance",
            "глобальное сопоставление и итоговый вывод оставлены merge-шагу"]
    body = json.dumps({k: v for k, v in model.items() if k != "sha256"},
                      ensure_ascii=False, sort_keys=True, default=str)
    model["sha256"] = hashlib.sha256(body.encode("utf-8")).hexdigest()
    root_package["source_model"] = model
    items = _source_memory(model)
    root_package["working_memory"] = _merge_memory(root_package.get("working_memory"), items)
    _refresh_package(root_package)
    return {"ok": True, "package": root_package, "source_model": model,
            "source_memory_ids": [x.get("id") for x in items if x.get("id")],
            "derived": True, "llm_calls": 0}


def owner_question_gate(task_package, question):
    """Allow owner interruption only for a genuine business choice, not a partition artefact."""
    question = str(question or "").strip()
    step = (task_package or {}).get("upc_step") if isinstance((task_package or {}).get("upc_step"), dict) else {}
    contract = step.get("input_contract") if isinstance(step.get("input_contract"), dict) else {}
    scope = str(contract.get("scope") or "")
    role = str(step.get("execution_role") or "")
    text = question.casefold()
    asks_for_input = any(marker in text for marker in (
        "provide", "attach", "upload", "missing file", "missing input", "which file",
        "предостав", "прилож", "загруз", "не хватает файл", "отсутствует файл",
        "какой файл", "где файл", "оба файла", "two excel", "два excel"))
    if question and asks_for_input and scope == "map_partition":
        return {"ask": False, "classification": "partition_contract_defect",
                "reason": "соседние источники существуют в process manifest и принадлежат другим узлам"}
    if question and asks_for_input and role in ("runtime_setup", "delivery", "control"):
        return {"ask": False, "classification": "runtime_context_defect",
                "reason": "служебный шаг не должен запрашивать исходные бизнес-файлы"}
    return {"ask": bool(question), "classification": "business_ambiguity" if question else "none",
            "reason": ""}


def prepare_task_context(session_id, sample_files, sess_dir, llm, progress=None, step_contract=None):
    """Общий Task Contract + Source Model для линейного и агентного Builder до любого кодогена."""
    package = _attach_capabilities(make_task_package(session_id, sample_files, sess_dir, llm=llm))
    package = _apply_upc_step(package, step_contract)
    checked = (build_source_model(package, llm, max_tries=2, progress=progress) if progress else
               build_source_model(package, llm, max_tries=2))
    if not checked.get("ok"):
        return {"ok": False, "why": checked.get("why") or "Source Model не построена",
                "package": package, "source_model": {"status": "error",
                                                       "reason": checked.get("why") or "не построена"},
                "source_memory_ids": []}
    model = checked["model"]
    package["source_model"] = model
    items = _source_memory(model)
    package["working_memory"] = _merge_memory(package.get("working_memory"), items)
    _refresh_package(package)
    return {"ok": True, "package": package, "source_model": model,
            "source_memory_ids": [x.get("id") for x in items if x.get("id")]}


def _refresh_package(package):
    body = {k: v for k, v in (package or {}).items() if k != "package_sha256"}
    raw = json.dumps(body, ensure_ascii=False, sort_keys=True, default=str)
    package["package_sha256"] = hashlib.sha256(raw.encode("utf-8")).hexdigest()
    return package


def _memory_from_judge(judge, attempt, passed=False):
    """Структурировать вывод приёмщика. Провал никогда не превращает гипотезу в знание."""
    judge = judge if isinstance(judge, dict) else {}
    raw_memory = judge.get("memory") if isinstance(judge.get("memory"), dict) else {}
    additions = []
    for kind, key in (("concept", "concepts"), ("rule", "rules")):
        for raw in raw_memory.get(key) if isinstance(raw_memory.get(key), list) else []:
            item = _normalize_memory_entry(
                raw, kind=kind, status="verified" if passed else "candidate",
                source="acceptance_gate", scope="process", attempt=attempt)
            if item:
                # Даже если модель сама написала verified, право подтвердить это даёт только PASS.
                item["status"] = "verified" if passed else "candidate"
                additions.append(item)
    for raw in judge.get("rejected_hypotheses") if isinstance(judge.get("rejected_hypotheses"), list) else []:
        item = _normalize_memory_entry(
            raw, kind="rule", status="rejected", source="acceptance_gate",
            scope="session", attempt=attempt)
        if item:
            item["status"] = "rejected"
            if not item.get("rejection_reason"):
                item["rejection_reason"] = "гипотеза отклонена независимой приёмкой"
            additions.append(item)
    return additions


def _failure_lesson(attempt, issues, code_sha="", source="harness", no_progress=False):
    detail = "; ".join(str(x) for x in (issues or []) if str(x).strip())[:520]
    if not detail:
        detail = "попытка не прошла проверку"
    prefix = "Не повторять неизменённое решение" if no_progress else "Не считать решением"
    return _normalize_memory_entry({
        "text": "%s: %s" % (prefix, detail), "status": "rejected", "source": source,
        "scope": "session", "attempt": attempt, "confidence": 1.0,
        "evidence": {"code_sha256": code_sha, "issues": list(issues or [])[:8]},
        "rejection_reason": detail}, kind="rule", status="rejected", source=source,
        scope="session", attempt=attempt)


def _promote_memory(memory, ids, attempt):
    wanted = {str(x) for x in ids if str(x)}
    for item in (memory or {}).get("entries") or []:
        if str(item.get("id")) in wanted and item.get("status") == "candidate":
            item["status"] = "verified"
            item["confidence"] = max(0.8, float(item.get("confidence") or 0))
            item["evidence"] = {"accepted_attempt": attempt, "previous": item.get("evidence")}
    return memory


def _verified_memory(memory):
    """Только подтверждённые знания процесса/агента разрешено переносить в production brain."""
    out = []
    for item in (memory or {}).get("entries") or []:
        if (item.get("status") == "verified" and item.get("scope") in ("process", "agent") and
                item.get("source") != "owner"):
            out.append(dict(item))
    return out


def _persist_agentic_state(sess_dir, session_id, bdir, package, source_model, status, build_id):
    """Переживает перезапуск UI, но память жёстко привязана к хэшу Task Contract."""
    root = Path(sess_dir)
    bdir = Path(bdir)
    memory = package.get("working_memory") or {"version": 1, "entries": []}
    memory["task_contract_sha256"] = (package.get("task_contract") or {}).get("sha256", "")
    memory["last_build_id"] = build_id
    memory["last_status"] = status
    memory["updated_at"] = datetime.now(timezone.utc).isoformat()
    (bdir / "task_contract.json").write_text(
        json.dumps(package.get("task_contract") or {}, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (bdir / "source_model.json").write_text(
        json.dumps(source_model or {}, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (bdir / "working_memory.json").write_text(
        json.dumps(memory, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    session_path = root / (session_id + ".json")
    lock_handle = None
    temp_path = None
    try:
        import fcntl
        lock_handle = open(str(session_path) + ".lock", "w")
        fcntl.flock(lock_handle, fcntl.LOCK_EX)
        session = _read_json(session_path, {}) or {}
        session["agentic_memory"] = memory
        fd, temp_name = tempfile.mkstemp(prefix=session_path.name + ".agentic.",
                                         suffix=".tmp", dir=str(root))
        temp_path = Path(temp_name)
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(session, handle, ensure_ascii=False, indent=2, default=str)
            handle.flush()
            os.fsync(handle.fileno())
        temp_path.replace(session_path)
        temp_path = None
    finally:
        if temp_path is not None:
            try:
                temp_path.unlink()
            except OSError:
                pass
        if lock_handle is not None:
            try:
                fcntl.flock(lock_handle, fcntl.LOCK_UN)
                lock_handle.close()
            except Exception:
                pass
    return memory


def _repair_context(feedback, task_package):
    """Короткая диагностическая обратная связь без готовых ответов контрольного набора."""
    if not feedback:
        return ""
    if not isinstance(feedback, str):
        feedback = json.dumps(_compact(feedback), ensure_ascii=False, default=str)
    text = str(feedback)
    for literal in _sample_literals(task_package or {}):
        if literal:
            text = re.sub(re.escape(literal), "<sample_value>", text, flags=re.I)
    # Абсолютный путь не помогает ремонту и может случайно стать частью сгенерированного решения.
    text = re.sub(r"(?:/[^\s\"']+){2,}/([^/\s\"']+)", r"<path>/\1", text)
    return _clip(text, 3200)


def _agent_headers(agent_id):
    return {"X-Auth-Token": CONFIG.get("auth_token", ""), "Content-Type": "application/json",
            "X-Profile-Id": "default", "X-Agent-Id": agent_id or qwen_agent()}


def _post_agent(agent_id, payload, timeout=700):
    req = urllib.request.Request(BASE.rstrip("/") + "/api/agent/run",
                                 data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                                 headers=_agent_headers(agent_id), method="POST")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        return {"status": "error", "message": _clip(exc, 400)}


def _agent_text(response):
    """Текст Responses API: финальный message приоритетнее внутреннего reasoning.

    Некоторые Qwen-compatible провайдеры возвращают только ``reasoning_text`` даже при
    ``status=completed``. Это не доказательство выполненного действия, но такой текст нужен
    JSON-fallback-у и для короткой диагностики вместо печати всего envelope в интерфейс.
    """
    final, reasoning = [], []
    for item in (response or {}).get("output", []):
        if not isinstance(item, dict):
            continue
        bucket = final if item.get("type") == "message" else reasoning
        for content in item.get("content", []):
            if not isinstance(content, dict):
                continue
            if content.get("type") in ("output_text", "text") and item.get("type") == "message":
                bucket.append(str(content.get("text") or ""))
            elif content.get("type") == "reasoning_text":
                reasoning.append(str(content.get("text") or ""))
    direct = (response or {}).get("output_text")
    if isinstance(direct, str) and direct.strip():
        final.append(direct)
    return "".join(final).strip() or "".join(reasoning).strip()


def _json_object(text):
    """Извлечь один JSON-объект из ответа модели; пояснения и markdown не считаются контрактом."""
    if isinstance(text, dict):
        return text
    raw = str(text or "").strip()
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        pass
    decoder = json.JSONDecoder()
    for index, char in enumerate(raw):
        if char != "{":
            continue
        try:
            parsed, _ = decoder.raw_decode(raw[index:])
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            continue
    return {}


def _llm_json(llm, prompt, max_tokens=3800, timeout=260):
    """Один Qwen JSON-вызов: пользовательский OpenAI-compatible endpoint или платформа Extella."""
    llm = llm if isinstance(llm, dict) else {}
    if llm.get("api_key"):
        try:
            body = {"model": llm.get("model"), "temperature": 0,
                    "response_format": {"type": "json_object"}, "max_tokens": int(max_tokens),
                    "messages": [{"role": "system", "content": "Верни только один валидный JSON-объект."},
                                 {"role": "user", "content": prompt}]}
            req = urllib.request.Request(llm.get("base_url", "").rstrip("/") + "/chat/completions",
                                         data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                                         headers={"Authorization": "Bearer " + str(llm["api_key"]),
                                                  "Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=timeout) as response:
                envelope = json.loads(response.read().decode("utf-8"))
            content = envelope["choices"][0]["message"]["content"]
            return _json_object(content)
        except Exception:
            return {}
    agent_id = llm.get("agent_id") or qwen_agent()
    response = _post_agent(agent_id, {"agent_id": agent_id, "input": prompt,
                                      "run_timeout": max(60, int(timeout) - 20), "store": False,
                                      "temperature": 0, "max_output_tokens": int(max_tokens)}, timeout=timeout)
    return _json_object(_agent_text(response))


PDF_LIBRARY_MARKERS = ("pypdf", "pypdf2", "pdfminer", "pdfplumber", "pymupdf", "fitz")


def _task_has_pdf_input(task_package):
    """True only when the current executable step physically receives a PDF."""
    for profile in (task_package or {}).get("inputs") or []:
        if not isinstance(profile, dict):
            continue
        extension = str(profile.get("extension") or profile.get("format") or "").strip().casefold()
        name = str(profile.get("name") or "").strip().casefold()
        if extension in (".pdf", "pdf") or name.endswith(".pdf"):
            return True
    return False


def _pdf_uses_verified_reuse(task_package):
    """A verified PDF capability owns parsing on a REUSE route; generated BUILD code does not."""
    decision = ((task_package or {}).get("execution_decision")
                if isinstance((task_package or {}).get("execution_decision"), dict) else {})
    if str(decision.get("route") or "").casefold() != "reuse":
        return False
    selected = json.dumps(decision.get("selected_capabilities") or [],
                          ensure_ascii=False, default=str).casefold()
    return any(marker in selected for marker in PDF_LIBRARY_MARKERS + ("ocr",))


def _pdf_library_include_present(code):
    """Fython loads dependencies only from module-level includes after the include extension."""
    source = str(code or "")
    first_line = next((line.strip() for line in source.splitlines() if line.strip()), "")
    if first_line != '$extens("include.py")':
        return False
    top_function = re.search(r"^def\s+[A-Za-z_][A-Za-z0-9_]*\s*\(", source, flags=re.M)
    prefix = source[:top_function.start()] if top_function else source
    for match in re.finditer(r"\binclude\s*\((.*?)\)", prefix, flags=re.I | re.S):
        body = match.group(1).casefold()
        if "extella-pip install" in body and any(marker in body for marker in PDF_LIBRARY_MARKERS):
            return True
    return False


def _pdf_extraction_repair_required(feedback, task_package):
    """Escalate a rejected empty/raw PDF attempt to a different extraction method."""
    if not feedback or not _task_has_pdf_input(task_package):
        return False
    text = (feedback if isinstance(feedback, str) else
            json.dumps(_compact(feedback), ensure_ascii=False, default=str))
    folded = str(text).casefold()
    method_markers = (
        "basic_regex", "raw pdf", "raw byte", "сырым байт", "сырых байт",
        "flate", "pdf-вход требует библиотеч", "pdf input requires",
        "pdf library", "pdf-библиотек",
    )
    if any(marker in folded for marker in method_markers):
        return True
    zero_records = re.search(
        r"(?:\b0\s+(?:records?|rows?|entries?)\b|\b0\s+(?:запис|строк|объект)[а-яё]*\b)",
        folded)
    extraction_context = any(marker in folded for marker in (
        "extract", "извлеч", "прочитан", "обработ", "processed_files", "files_confirmed"))
    return bool(zero_records and extraction_context)


def _build_prompt(expert_name, task_package, feedback, agent_id):
    brief_json = json.dumps(_builder_brief(task_package), ensure_ascii=False, indent=2, default=str)
    decision = (task_package.get("execution_decision")
                if isinstance(task_package.get("execution_decision"), dict) else {})
    route = str(decision.get("route") or "build")
    selected_caps = [str(x) for x in (decision.get("selected_capabilities") or []) if str(x)]
    route_directives = {
        "reuse": ("Маршрут уже выбран: REUSE. Не изобретай замену выбранной способности. Создай одну "
                  "стабильную точку входа, которая вызывает только выбранные verified capabilities " +
                  json.dumps(selected_caps, ensure_ascii=False) +
                  " и приводит их результат к acceptance_contract."),
        "build": ("Маршрут уже выбран: BUILD. Создай детерминированный кодовый эксперт. Не вызывай Qwen "
                  "во время runtime, если Task Contract не требует смыслового решения."),
        "llm_worker": ("Маршрут уже выбран: LLM_WORKER. Создай исполняемого эксперта, который вызывает "
                       "платформенную Qwen при каждом запуске для смысловой части, но детерминированно "
                       "проверяет JSON-схему, обязательные поля и acceptance evidence."),
    }
    route_directive = route_directives.get(route, "Маршрут выполнения зафиксирован: " + route + ".")
    pdf_contract = ""
    if _task_has_pdf_input(task_package):
        if _pdf_uses_verified_reuse(task_package):
            pdf_contract = (
                "\n\nPDF INPUT CONTRACT. Разбор PDF принадлежит выбранной verified PDF/OCR capability: "
                "вызови именно её и зафиксируй capability evidence. Даже на REUSE-маршруте запрещено "
                "искать текст регулярным выражением в сырых байтах PDF: content streams могут быть сжаты.")
        else:
            pdf_contract = (
                "\n\nPDF INPUT CONTRACT — ОБЯЗАТЕЛЬНО. Физический .pdf извлекай настоящим PDF-парсером, "
                "объявленным через каноническую зависимость Fython. ПЕРВАЯ непустая строка исходника должна "
                'быть $extens("include.py"); затем на уровне модуля ДО def размести include. '
                "Никогда не вызывай include внутри функции — в runtime это даст name 'include' is not defined. "
                "Например: "
                'include("from pypdf import PdfReader", ["extella-pip install pypdf"]) или '
                'include("import pdfplumber", ["extella-pip install pdfplumber"]). '
                "Bare import без include не гарантирует библиотеку в чистом runtime. ЗАПРЕЩЕНО читать PDF "
                "как сырые bytes и искать текст/записи regex или decode: PDF content streams часто сжаты "
                "(FlateDecode), поэтому такой метод ложно возвращает 0 записей. Если библиотечное извлечение "
                "вернуло пустой текст и профиль доказывает скан, используй библиотечный renderer/OCR либо "
                "верни точную техническую ошибку; не выдавай пустой разбор за success.")
    upc_step = task_package.get("upc_step") if isinstance(task_package.get("upc_step"), dict) else None
    step_directive = ""
    if upc_step:
        required = ((task_package.get("acceptance_contract") or {}).get("required_artifacts") or
                    ["step_result_json"])
        input_contract = (upc_step.get("input_contract")
                          if isinstance(upc_step.get("input_contract"), dict) else {})
        scope = str(input_contract.get("scope") or "atomic_step")
        role = str(upc_step.get("execution_role") or "data")
        step_directive = ("\n\nЭТО ОДИН ШАГ UNIVERSAL PROCESS CONTRACT. Реализуй ТОЛЬКО upc_step из BUILD_BRIEF, "
                          "не всю конечную задачу и не соседние шаги. Выполни его input/output contract, "
                          "не расширяй permissions и верни наблюдаемые доказательства именно его acceptance. "
                          "Зависимости уже приняты runtime-ом и передаются через source_file. "
                          "При контрольной сборке любые move/modify/delete/install/send/external_write выполняй "
                          "только как preview: создай manifest точных предполагаемых действий, но не меняй внешний мир. "
                          "Всегда запиши в output_dir файл step_result.json с summary, output и evidence; верни его в "
                          "artifacts как kind=step_result_json. Дополнительные обязательные артефакты: " +
                          json.dumps(required, ensure_ascii=False) + ".")
        if scope == "map_partition":
            step_directive += (
                "\nЭТО MAP-ПАРТИЦИЯ. Её единственная задача — прочитать локальные source_refs и вернуть "
                "канонические факты с provenance. Отсутствие других файлов процесса ожидаемо: они существуют "
                "в process_source_manifest, назначены соседним узлам и будут сведены merge-шагом. Не сравнивай "
                "локальный PDF с отсутствующим здесь Excel, не делай глобальный бизнес-вывод и не проси "
                "владельца приложить соседние файлы.")
        if role in ("runtime_setup", "delivery", "control"):
            step_directive += (
                "\nЭТО СЛУЖЕБНЫЙ ШАГ. source_file содержит конфигурационный/dependency envelope, а не набор "
                "сырых бизнес-данных. Проверяй только настройку/preview и не ищи PDF/XLSX исходного процесса.")
    repair = ""
    if feedback:
        action = str(feedback.get("required_next_action") or "") if isinstance(feedback, dict) else ""
        previous_code = str(feedback.get("previous_expert_code") or "") \
            if isinstance(feedback, dict) else ""
        replace_pdf_method = _pdf_extraction_repair_required(feedback, task_package)
        pdf_repair = (
            "\nСМЕНА PDF-МЕТОДА ОБЯЗАТЕЛЬНА: предыдущая попытка не извлекла записи или использовала "
            "сырой byte/regex-подход. Не сохраняй эту вычислительную часть, даже если контроллер ранее "
            "назвал ремонт выходным. Замени извлечение на PDF-библиотеку через include(...), а для "
            "доказанного скана — на renderer/OCR; затем заново посчитай evidence."
            if replace_pdf_method else "")
        if action == "repair_output_contract" and previous_code and not replace_pdf_method:
            repair = ("\n\nРЕМОНТ ТОЛЬКО ВЫХОДНОГО КОНТРАКТА. Вычислительная логика прошлого эксперта "
                      "уже отработала: не меняй выбор входов, чтение, нормализацию, сопоставление и расчёты. "
                      "Сделай минимальное изменение упаковки summary/output/evidence/artifacts и записи "
                      "обязательных файлов в output_dir. Верни полный код с сохранённой вычислительной частью. "
                      "Диагностика харнесса:\n" + _repair_context(feedback, task_package) +
                      "\n\nКОД, ВЫЧИСЛИТЕЛЬНУЮ ЧАСТЬ КОТОРОГО НУЖНО СОХРАНИТЬ:\n" +
                      _clip(previous_code, 12000))
        else:
            repair = ("\n\nПРЕДЫДУЩИЙ РЕАЛЬНЫЙ ПРОГОН НЕ ПРОШЁЛ. Исправь первопричину, не подгоняя решение "
                      "под конкретные строки образца. Обнови и снова сохрани того же эксперта. "
                      "Диагностика харнесса:\n" + _repair_context(feedback, task_package) + pdf_repair)
    return f"""Ты — Строитель Extella. Task-first архитектор уже изучил задачу БЕЗ технического имени
эксперта и сохранил execution_decision. Теперь не выбирай другой тип решения: материализуй выбранный
маршрут в одном реально исполняемом эксперте `{expert_name}` действием платформы.

{route_directive}

	Сначала молча составь для себя цепочку «требование владельца → реализация → наблюдаемое доказательство».
	Решай ТЗ и ВСЕ профили файлов как одну бизнес-задачу. Внутри выбранного маршрута сам выбери алгоритм,
	библиотеки и внутренние этапы.
	Не переноси в код догадки из образца и не восстанавливай старый план стадий: источник истины — BUILD_BRIEF
	в указанном там порядке авторитетности. Если требования действительно противоречат друг другу, верни
	status=error с точным противоречием вместо выдуманного правила.
	Source Model уже сопоставила реальные входы, разделы, сущности и операции. Следуй её strategy, но не
	считай недоказанную нормализацию фактом. В working_memory: verified можно применять, candidate нужно
	проверять реальным прогоном, rejected — это отрицательные уроки, их нельзя повторять как решение.

В relevant_capabilities уже отобраны только вероятно полезные проверенные способности. Реюз необязателен:
используй способность через Extella API лишь когда она непосредственно улучшает решение. Финальная точка
запуска и ответственности всё равно одна — `{expert_name}`.

Обязательный контракт кода:
- CSPL=fython, ровно одна top-level функция; любые helpers только внутри неё;
- сигнатура РОВНО:
  def {expert_name}(source_file="", output_dir="", api_token="", api_base="https://api.extella.ai", target="", source_key="", rules_json="", fields_json="", run_id="", placement_json="", adapter_json="", report_spec_json="")
- source_file бывает файлом или папкой. Если это папка, найди и используй ВСЕ относящиеся к задаче файлы;
- искать входы можно ТОЛЬКО внутри source_file. Никаких fallback-поисков в cwd, home, /tmp, /data
  и других каталогах: если нужного файла нет в source_file, верни точную ошибку;
- входы не изменять; результаты писать только в output_dir;
- никаких абсолютных путей, секретов, shell-команд, отправки писем и записей во внешние системы;
- клиентские правила брать из rules_json/fields_json, а не зашивать значениями образца;
- ЗАПРЕЩЕНО помещать в код значения строк из образцов/контрольных кейсов (идентификаторы, суммы, даты,
  ожидаемые A/B/C). Образцы нужны только для проверки; алгоритм обязан работать на следующих файлах;
- для PDF сначала извлеки текст; OCR применяй только если профиль указывает на скан или текста нет;
- детерминированные файловые задачи (парсинг, сверка, расчёт, отчёт) решай ЧИСТЫМ КОДОМ без вызова
  моделей изнутри: весь интеллект закладывается сейчас, при генерации;
- если смысловой шаг ДЕЙСТВИТЕЛЬНО нужен — разрешена ТОЛЬКО платформенная Qwen ровно так:
  POST <api_base>/api/agent/run, заголовки X-Auth-Token: api_token, X-Profile-Id: "default",
  X-Agent-Id: "{agent_id}", тело {{"agent_id":"{agent_id}","input":<текст>,"store":false}}.
  Endpoint'ов вида /v1/chat/completions или openai-style на платформе НЕТ — такой вызов упадёт 404
  и код будет отклонён проверкой ещё до запуска;
- на этапе создания ТОЛЬКО сохрани/обнови эксперта. Не запускай его сам и не делай пробных вызовов:
  Визард отдельно прогонит сохранённый код с явными source_file и output_dir;
- для целостной legacy-сборки создай report.md и report.xlsx; для шага UPC создай перечисленные в
  acceptance_contract артефакты и обязательный step_result.json;
	- верни dict:
	  {{"status":"success","summary":{{...,"processed_files":N}},"output":{{...}},
	    "evidence":{{"files_used":["basename",...],"capabilities_used":[{{"id":"...","evidence":"факт вызова"}}],"acceptance_checks":[
      {{"criterion":"...","passed":true,"evidence":"конкретный факт"}}]}},
    "artifacts":[{{"kind":"step_result_json","path":"/absolute/generated/step_result.json"}}]}}
- acceptance_checks доказывают, что алгоритм выполнил требование и честно отразил результат. Найденное
  расхождение, исключение, пустая категория или отрицательный бизнес-вердикт — это passed=true, если они
  корректно вычислены и показаны; passed=false означает только сбой/неполную обработку. Не зашивай в
  критерии ожидаемые идентификаторы или количества конкретного образца;
- status=success только если бизнес-цель выполнена на данных; при невозможности верни status=error и точную причину.

Исходный код помещай ТОЛЬКО в действие создания/обновления эксперта, не печатай код в ответе.
После действия ответь кратко, но источником истины будет сохранённый эксперт и его реальный прогон.

BUILD_BRIEF:
{brief_json}{pdf_contract}{step_directive}{repair}"""


def _code_artifact_prompt(build_prompt, expert_name):
    """Детерминированный основной контракт: модель проектирует, харнесс сохраняет.

    Код возвращается харнессу, проходит тот же валидатор и сохраняется только через
    контролируемый ``/api/expert/save``. Нативное tool-action используется лишь как fallback.
    """
    normalized = build_prompt.replace(
        "Исходный код помещай ТОЛЬКО в действие создания/обновления эксперта, не печатай код в ответе.",
        "Исходный код верни ТОЛЬКО в поле code итогового JSON.")
    return ("РЕЖИМ CODE-ARTIFACT. Не вызывай действия платформы и не объясняй ход мысли. "
            "На основании полной спецификации ниже "
            "верни РОВНО один JSON-объект без markdown: "
            '{"code":"<полный исходный код>","description":"<назначение эксперта>"}. '
            "Поле code обязано содержать ровно одну top-level функцию с именем " + expert_name +
            ". Не сокращай код и не используй многоточия. Фразы полной спецификации о действии сохранения "
            "заменены этим контрактом: сохранять будет харнесс.\n\nПОЛНАЯ СПЕЦИФИКАЦИЯ:\n" + normalized)


def _sample_literals(task_package):
    """Отличительные значения СТРОК образца, которым запрещено просачиваться в исходник решения."""
    values = set()
    declared = json.dumps(
        {key: value for key, value in (task_package or {}).items() if key != "inputs"},
        ensure_ascii=False, default=str).casefold()

    def is_declared_schema_literal(value):
        """Allow output enum labels that are merely humanized spellings of the contract.

        A downstream report legitimately maps ``only_excel`` to ``Only Excel``.  It must not,
        however, embed a concrete sample identifier such as ``A-101`` merely because the owner's
        acceptance example also mentions it.
        """
        text = str(value or "").strip()
        if text.casefold().startswith(("extella.", "upc-")):
            return True
        if re.search(r"\d", text):
            return False
        snake = re.sub(r"\s+", "_", text.casefold())
        return len(snake) >= 4 and re.search(
            r"(?<![a-z0-9])" + re.escape(snake) + r"(?![a-z0-9])", declared) is not None

    for profile in (task_package or {}).get("inputs") or []:
        for sheet in profile.get("workbook") or []:
            rows = sheet.get("sample_rows") or []
            for row in rows[1:]:  # первая строка — схема; имена колонок коду знать разрешено
                for cell in row:
                    s = str(cell or "").strip()
                    if (4 <= len(s) <= 100 and
                            (re.search(r"\d", s) or len(s.split()) >= 2) and
                            not is_declared_schema_literal(s)):
                        values.add(s)
        rows = profile.get("sample_rows") or []
        for row in rows[1:]:
            for cell in row:
                s = str(cell or "").strip()
                if (4 <= len(s) <= 100 and
                        (re.search(r"\d", s) or len(s.split()) >= 2) and
                        not is_declared_schema_literal(s)):
                    values.add(s)
        text = str(profile.get("text_sample") or "")
        for match in re.finditer(
                r"(?<!\w)[A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9_.\-/]{3,}(?!\w)", text):
            token = match.group(0)
            # Dependency envelopes and JSON fixtures contain technical schema keys such as
            # ``sha256``.  A generated expert must be free to implement that schema; only values
            # (for example a real seal number) are sample literals.  Treating a key as client data
            # exhausted the t4 creation budget after the PDF steps had already passed live E2E.
            if re.match(r"""["']?\s*:""", text[match.end():]):
                continue
            if re.search(r"\d", token):
                values.add(token)
    return sorted(values, key=lambda x: (-len(x), x))[:120]


def _validate_code(expert_name, code, task_package=None):
    issues = []
    tops = re.findall(r"^def\s+([A-Za-z_][A-Za-z0-9_]*)\s*\(", code or "", flags=re.M)
    if tops != [expert_name]:
        issues.append("нужна одна top-level функция %s, найдено %s" % (expert_name, tops))
    sig = re.search(r"^def\s+" + re.escape(expert_name) + r"\s*\((.*?)\)\s*(?:->[^:]+)?\s*:",
                    code or "", flags=re.M | re.S)
    if not sig or "source_file" not in sig.group(1) or "output_dir" not in sig.group(1):
        issues.append("сигнатура не содержит source_file/output_dir")
    for marker in ("/Users/", "/home/", "os.system(", "shell=True",
                   "Path.cwd(", "Path.home(", "os.getcwd(", "Path('/tmp')", 'Path("/tmp")',
                   "Path('/data')", 'Path("/data")'):
        if marker in (code or ""):
            issues.append("запрещённый маркер в коде: " + marker)
    leaked = [v for v in _sample_literals(task_package or {}) if v.casefold() in (code or "").casefold()]
    if leaked:
        issues.append("в код зашиты значения образца: " + ", ".join(leaked[:8]))
    if (_task_has_pdf_input(task_package) and not _pdf_uses_verified_reuse(task_package) and
            not _pdf_library_include_present(code)):
        issues.append(
            'PDF-вход требует $extens("include.py") первой строкой и библиотечный парсер в module-level '
            "include(... pypdf/PyPDF2/pdfminer/pdfplumber/PyMuPDF ...) до def; include внутри функции и "
            "regex/decode сырых PDF-байтов запрещены")
    # Класс отказа D (23.07): у «семантически пахнущих» шагов Qwen жульничает — генерит код, который
    # ИЗНУТРИ зовёт LLM по выдуманному по аналогии endpoint'у (api.extella.ai/v1/chat/completions →
    # 404 на прогоне, попытка сгорает). Детерминированный build-шаг обязан решать задачу кодом: весь
    # интеллект — сейчас, при генерации. Ловим СТАТИЧЕСКИ, до запуска — run-попытка не тратится, урок
    # ремонта адресный. llm_worker-маршрут не трогаем (там вызов модели легален).
    _route = str(((task_package or {}).get("execution_decision") or {}).get("route")
                 if isinstance((task_package or {}).get("execution_decision"), dict) else "build")
    if (_route or "build").casefold() == "build":
        low = (code or "").casefold()
        for marker in ("chat/completions", "v1/completions", "import openai", "openai.chatcompletion",
                       "api.openai.com", "api.anthropic.com", "dashscope"):
            if marker in low:
                issues.append("детерминированный шаг не должен вызывать LLM изнутри кода (найден маркер: "
                              + marker + ") — реши задачу самим кодом; вся семантика закладывается при генерации")
                break
    return issues


def _strategy_fingerprint(code):
    """Hash control/data-flow vocabulary while ignoring superficial literal/name changes."""
    try:
        tree = ast.parse(code or "")
    except Exception:
        return hashlib.sha256(re.sub(r"\s+", " ", str(code or "")).encode("utf-8")).hexdigest()
    features = []
    for node in ast.walk(tree):
        if isinstance(node, (ast.If, ast.For, ast.While, ast.Try, ast.With, ast.ListComp,
                             ast.DictComp, ast.Return, ast.Raise)):
            features.append(type(node).__name__)
        elif isinstance(node, ast.Import):
            features.extend("import:" + x.name.split(".")[0] for x in node.names)
        elif isinstance(node, ast.ImportFrom):
            features.append("from:" + str(node.module or "").split(".")[0])
        elif isinstance(node, ast.Call):
            fn = node.func
            if isinstance(fn, ast.Name):
                features.append("call:" + fn.id)
            elif isinstance(fn, ast.Attribute):
                features.append("call:" + fn.attr)
        elif isinstance(node, ast.Constant) and isinstance(node.value, str):
            value = node.value.casefold()
            for marker in ("status", "success", "summary", "evidence", "artifacts", "report_",
                           "read_excel", "read_csv", "json", "xlsx", "pdf", "csv"):
                if marker in value:
                    features.append("str:" + marker)
    return hashlib.sha256("\n".join(sorted(features)).encode("utf-8")).hexdigest()


def _get_scoped_expert(expert_name, agent_id):
    req = urllib.request.Request(BASE.rstrip("/") + "/api/expert/get",
                                 data=json.dumps({"name": expert_name}).encode("utf-8"),
                                 headers=_agent_headers(agent_id), method="POST")
    try:
        with urllib.request.urlopen(req, timeout=70) as response:
            return json.loads(response.read().decode("utf-8"))
    except Exception:
        return {}


def _create_or_update(expert_name, task_package, feedback, llm):
    """Create one expert with the shortest platform-appropriate path.

    BYOC endpoints return a code artifact which the harness validates and saves.  The Extella Qwen
    already has the native expert action used by Chat, so asking it for a long JSON code artifact
    first only to fall back to the action doubled the wall time.  For that path the native action is
    primary; deterministic code-artifact generation remains one bounded fallback.
    """
    started = time.monotonic()
    agent_id = llm.get("agent_id") or qwen_agent()
    prompt = _build_prompt(expert_name, task_package, feedback, agent_id)
    generation_path = "code_artifact"
    if llm.get("api_key"):
        try:
            body = {"model": llm.get("model"), "temperature": 0,
                    "response_format": {"type": "json_object"}, "max_tokens": 8000,
                    "messages": [{"role": "system", "content": "Верни JSON {code,description}."},
                                 {"role": "user", "content": prompt.replace(
                                     "Исходный код помещай ТОЛЬКО в действие создания/обновления эксперта, не печатай код в ответе.",
                                     "Верни исходный код в поле code JSON.")} ]}
            req = urllib.request.Request(llm.get("base_url", "").rstrip("/") + "/chat/completions",
                                         data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
                                         headers={"Authorization": "Bearer " + llm["api_key"],
                                                  "Content-Type": "application/json"}, method="POST")
            with urllib.request.urlopen(req, timeout=300) as response:
                raw = json.loads(response.read().decode("utf-8"))["choices"][0]["message"]["content"]
            spec = json.loads(raw)
            found = {"expert_code": spec.get("code", ""), "description": spec.get("description", "")}
        except Exception as exc:
            return {"ok": False, "why": "LLM не создала код: " + _clip(exc, 300)}
    else:
        generation_path = "native_action"
        run = _post_agent(agent_id, {"agent_id": agent_id, "input": prompt, "run_timeout": 260,
                                     "store": True, "max_output_tokens": 2200}, timeout=300)
        found = _get_scoped_expert(expert_name, agent_id)
        if not found or not (found.get("expert_code") or found.get("code")):
            found = api("/api/expert/get", {"name": expert_name, "global": True}, timeout=70)
        native_code = str((found or {}).get("expert_code") or (found or {}).get("code") or "")
        native_issues = _validate_code(expert_name, native_code, task_package) if native_code else [
            "нативное действие не сохранило code-artifact"]
        if native_issues:
            generation_path = "code_artifact_fallback"
            spec = _llm_json(llm, _code_artifact_prompt(prompt, expert_name),
                             max_tokens=9000, timeout=330)
            found = {"expert_code": str((spec or {}).get("code") or ""),
                     "description": str((spec or {}).get("description") or "")}
            if not found.get("expert_code"):
                response_kind = "reasoning без действия" if _agent_text(run) else \
                                str((run or {}).get("status") or "пустой ответ")
                # ФАКТИЧЕСКАЯ ошибка вызова обязана доехать до журнала: без неё «(error)» неотличим
                # от транзиента, дизайна или бага — два прогона E2E 22.07 расследовались вслепую.
                err_detail = str((run or {}).get("message") or (run or {}).get("error") or "")[:220]
                if err_detail:
                    response_kind += ": " + err_detail
                return {"ok": False, "why": ("Qwen не создала эксперта нативным действием (" +
                                               response_kind + ") и не вернула code-artifact"),
                        "generation_path": generation_path,
                        "generation_seconds": round(time.monotonic() - started, 1)}
    code = str(found.get("expert_code") or found.get("code") or "")
    issues = _validate_code(expert_name, code, task_package)
    if issues:
        return {"ok": False, "why": "; ".join(issues)}
    description = str(found.get("description") or "Целостное решение бизнес-задачи, проверенное на образцах")[:900]
    saved = api("/api/expert/save", {"name": expert_name, "description": description, "code": code,
                                      "kwargs": dict(EXPERT_KWARGS), "cspl": "fython", "global": True}, timeout=180)
    ok = isinstance(saved, dict) and (saved.get("status") == "success" or saved.get("id"))
    return {"ok": bool(ok), "why": "" if ok else "global save: " + _clip(saved, 400),
            "code": code,
            "code_sha256": hashlib.sha256(code.encode("utf-8")).hexdigest(),
            "strategy_sha256": _strategy_fingerprint(code),
            "generation_path": generation_path,
            "generation_seconds": round(time.monotonic() - started, 1)}


def _delete_draft(expert_name, agent_id=""):
    """Убрать временного эксперта из global и из скоупа Строителя (best effort)."""
    try:
        api("/api/expert/delete", {"name": expert_name, "global": True}, timeout=40)
    except Exception:
        pass
    if agent_id:
        try:
            req = urllib.request.Request(BASE.rstrip("/") + "/api/expert/delete",
                                         data=json.dumps({"name": expert_name}).encode("utf-8"),
                                         headers=_agent_headers(agent_id), method="POST")
            urllib.request.urlopen(req, timeout=40).read()
        except Exception:
            pass


def _promote_expert(draft_name, stable_name, task_package):
    """Атомарно опубликовать принятую draft-версию под стабильным именем процесса."""
    found = api("/api/expert/get", {"name": draft_name, "global": True}, timeout=70)
    code = str((found or {}).get("expert_code") or (found or {}).get("code") or "")
    if not code:
        return {"ok": False, "why": "принятый draft не найден перед публикацией"}
    code = re.sub(r"\b" + re.escape(draft_name) + r"\b", stable_name, code)
    issues = _validate_code(stable_name, code, task_package)
    if issues:
        return {"ok": False, "why": "; ".join(issues)}
    saved = api("/api/expert/save", {"name": stable_name,
                                      "description": str((found or {}).get("description") or
                                                         "Целостное решение бизнес-задачи, принятое Визардом")[:900],
                                      "code": code, "kwargs": dict(EXPERT_KWARGS),
                                      "cspl": "fython", "global": True}, timeout=180)
    ok = isinstance(saved, dict) and (saved.get("status") == "success" or saved.get("id"))
    return {"ok": bool(ok), "why": "" if ok else "stable save: " + _clip(saved, 400),
            "code_sha256": hashlib.sha256(code.encode("utf-8")).hexdigest()}


def _artifact_paths(result):
    out = []
    if not isinstance(result, dict):
        return out
    for key in REPORT_KEYS:
        if isinstance(result.get(key), str):
            out.append((key, result[key]))
    for item in result.get("artifacts") or []:
        if isinstance(item, str):
            out.append(("artifact", item))
        elif isinstance(item, dict) and isinstance(item.get("path"), str):
            out.append((str(item.get("kind") or "artifact"), item["path"]))
    seen, clean = set(), []
    for key, path in out:
        if path not in seen:
            seen.add(path); clean.append((key, path))
    return clean


def materialize_step_result(result, output_dir):
    """Persist one stable result envelope for every real expert execution.

    The model owns the business result; the harness only normalizes its transport shape and records
    it.  In particular, this function never invents used files, successful checks or business
    values.  That keeps deterministic and semantic acceptance fail-closed while giving every step
    the same inspectable artifact regardless of which Qwen/provider produced it.
    """
    raw = result
    if isinstance(raw, str):
        raw = _json_object(raw)
    if not isinstance(raw, dict):
        raw = {"status": "error", "message": "expert returned a non-object result",
               "raw_result": _clip(result, 1200)}
    else:
        raw = copy.deepcopy(raw)

    # Some executors wrap the expert payload once. Unwrap only a recognisable result object; a
    # provider response or task receipt is evidence of failure, not proof of business success.
    nested = raw.get("result")
    if (isinstance(nested, dict) and str(nested.get("status") or "") in ("success", "error") and
            not any(key in raw for key in ("summary", "evidence", "output", "structured_data"))):
        raw = copy.deepcopy(nested)

    evidence = raw.get("evidence")
    if isinstance(evidence, list):
        evidence = {"acceptance_checks": copy.deepcopy(evidence)}
    elif not isinstance(evidence, dict):
        evidence = {}
    raw["evidence"] = evidence

    files_used = evidence.get("files_used") or raw.get("files_used") or []
    if isinstance(files_used, str):
        files_used = [files_used]
    elif not isinstance(files_used, (list, tuple, set)):
        files_used = []
    files_used = [_nfc(str(item)) for item in files_used if str(item).strip()]
    if files_used:
        evidence["files_used"] = files_used
    summary = raw.get("summary") if isinstance(raw.get("summary"), dict) else {}
    # A count may be derived from the expert's own file evidence; expected inputs are deliberately
    # not used here because that would allow a skipped input to pass acceptance.
    if files_used and summary.get("processed_files") in (None, ""):
        summary["processed_files"] = len({_ref_key(Path(item).name) for item in files_used})
    raw["summary"] = summary

    structured = raw.get("structured_data")
    if structured is None:
        structured = raw.get("output") if isinstance(raw.get("output"), (dict, list)) else {}
    raw["structured_data"] = structured
    if "output" not in raw and structured:
        raw["output"] = copy.deepcopy(structured)

    errors = raw.get("errors") if isinstance(raw.get("errors"), list) else []
    if isinstance(raw.get("errors"), str) and raw.get("errors").strip():
        errors = [raw.get("errors").strip()]
    if raw.get("status") != "success" and not errors:
        why = raw.get("message") or raw.get("error") or raw.get("status") or "expert failed"
        errors = [_clip(why, 1000)]
    raw["errors"] = errors

    artifacts = raw.get("artifacts") if isinstance(raw.get("artifacts"), list) else []
    if isinstance(raw.get("artifacts"), str) and raw.get("artifacts").strip():
        artifacts = [raw.get("artifacts").strip()]
    artifacts = copy.deepcopy(artifacts)
    known_paths = {str(item.get("path")) for item in artifacts if isinstance(item, dict)} | {
        str(item) for item in artifacts if isinstance(item, str)}
    for key in REPORT_KEYS:
        path = raw.get(key)
        if isinstance(path, str) and path.strip() and path not in known_paths:
            artifacts.append({"kind": key, "path": path})
            known_paths.add(path)
    raw["artifacts"] = artifacts
    target = Path(output_dir).resolve() / "step_result.json"
    envelope = {
        "schema": "extella.step_result.v1",
        "status": raw.get("status"),
        "summary": summary,
        "processed_files": summary.get("processed_files"),
        "structured_data": structured,
        "artifacts": artifacts,
        "errors": errors,
        "evidence": evidence,
    }
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_suffix(".json.tmp")
        temporary.write_text(json.dumps(envelope, ensure_ascii=False, indent=2, default=str),
                             encoding="utf-8")
        temporary.replace(target)
        raw["artifacts"].append({"kind": "step_result_json", "path": str(target)})
    except Exception as exc:
        raw["status"] = "error"
        raw["message"] = "canonical step_result.json was not persisted: " + _clip(exc, 500)
        raw["errors"].append(raw["message"])
    return raw


def validate_result(result, sample_files, task_package=None):
    """Детерминированная часть приёмки: модель не может позеленить отсутствующие входы/отчёты."""
    issues = []
    if not isinstance(result, dict):
        return {"ok": False, "issues": ["результат не является объектом"]}
    if result.get("status") != "success":
        issues.append("status не success: " + _clip(result.get("message") or result.get("status"), 300))
    summary = result.get("summary")
    if not isinstance(summary, dict) or not summary:
        issues.append("нет содержательной summary")
    acceptance_contract = ((task_package or {}).get("acceptance_contract")
                           if isinstance(task_package, dict) else {}) or {}
    evidence = result.get("evidence") if isinstance(result.get("evidence"), dict) else {}
    used = evidence.get("files_used") or result.get("files_used") or []
    used = {_ref_key(Path(str(x)).name) for x in used if str(x).strip()}
    expected_names = {_ref_key(Path(str(x)).name): _nfc(Path(str(x)).name) for x in sample_files}
    expected = set(expected_names)
    missing = [expected_names[key] for key in sorted(expected - used)]
    if acceptance_contract.get("must_use_every_sample", True) and missing:
        issues.append("нет доказательства обработки файлов: " + ", ".join(missing))
    processed = (summary or {}).get("processed_files") if isinstance(summary, dict) else None
    if acceptance_contract.get("must_use_every_sample", True) and expected:
        try:
            if int(processed) < len(expected):
                issues.append("processed_files меньше числа входов")
        except Exception:
            issues.append("summary.processed_files не задан")
    checks = evidence.get("acceptance_checks") or result.get("acceptance_checks") or []
    if not isinstance(checks, list) or not checks:
        issues.append("нет evidence.acceptance_checks")
    elif any(not isinstance(c, dict) or c.get("passed") is not True or not str(c.get("evidence") or "").strip()
             for c in checks):
        issues.append("не все критерии приёмки подтверждены конкретными фактами")
    source_model = (task_package or {}).get("source_model") if isinstance(task_package, dict) else {}
    source_model = source_model if isinstance(source_model, dict) else {}
    selected = {str(x) for x in (source_model.get("selected_capabilities") or []) if str(x)}
    if source_model.get("strategy") in ("reuse", "compose") and selected:
        cap_rows = evidence.get("capabilities_used") if isinstance(evidence.get("capabilities_used"), list) else []
        used_caps = {str(x.get("id") or "") for x in cap_rows if isinstance(x, dict) and
                     str(x.get("evidence") or "").strip()}
        missing_caps = sorted(selected - used_caps)
        if missing_caps:
            issues.append("нет доказательства контрольного вызова capabilities: " + ", ".join(missing_caps))
    artifacts = _artifact_paths(result)
    live = []
    for key, path in artifacts:
        try:
            p = Path(path)
            if p.is_file() and p.stat().st_size >= 80:
                live.append((key, str(p), p.stat().st_size))
        except Exception:
            continue
    required = acceptance_contract.get("required_artifacts")
    if isinstance(required, list):
        required = canonical_artifact_requirements(required, ensure_step_result=bool(required))
    else:
        required = ["report_md", "report_xlsx"]
    for need in [str(x).casefold() for x in required if str(x).strip()]:
        matches = [(kind, path, size) for kind, path, size in live
                   if need in (str(kind) + " " + Path(path).name + " " + path).casefold()]
        if not matches:
            issues.append("не создан обязательный артефакт: " + need)
    return {"ok": not issues, "issues": issues, "artifacts": live,
            "files_used": sorted(used), "summary": summary if isinstance(summary, dict) else {}}


def _result_preview(result, validation):
    preview = {"result": _compact(result), "artifact_previews": []}
    for key, path, size in (validation.get("artifacts") or [])[:5]:
        item = {"kind": key, "name": Path(path).name, "bytes": size}
        try:
            if Path(path).suffix.lower() in (".md", ".txt", ".html"):
                item["text"] = _clip(Path(path).read_text(encoding="utf-8", errors="replace"), 7000)
            elif Path(path).suffix.lower() in (".xlsx", ".xlsm"):
                item["profile"] = profile_file(path).get("workbook")
        except Exception as exc:
            item["preview_error"] = _clip(exc, 200)
        preview["artifact_previews"].append(item)
    return _compact(preview)


def _repair_feedback(result, validation, judge, controller=None):
    """Ровно то, что нужно следующей попытке: тип провала и агрегаты, без таблицы правильных ответов."""
    result = result if isinstance(result, dict) else {}
    validation = validation if isinstance(validation, dict) else {}
    judge = judge if isinstance(judge, dict) else {}
    controller = controller if isinstance(controller, dict) else {}
    return {
        "phase": "acceptance",
        "returned_status": result.get("status"),
        "returned_summary": _compact(result.get("summary") or {}),
        "files_confirmed": [Path(str(x)).name for x in (validation.get("files_used") or [])[:20]],
        "artifacts_created": [str(x[0]) for x in (validation.get("artifacts") or [])[:10]
                              if isinstance(x, (list, tuple)) and x],
        "structural_issues": validation.get("issues") or [],
        "business_verdict": judge.get("verdict"),
        "business_issues": judge.get("issues") or [],
        "owner_question": judge.get("owner_question") or "",
        "failure_controller": {k: controller.get(k) for k in
                               ("failure_class", "action", "evidence", "confidence")},
        "required_next_action": controller.get("action") or "repair_code",
    }


def judge_result(task_package, result, validation):
    """Независимый смысловой гейт: статус эксперта сам по себе не доказывает бизнес-результат."""
    agent_id = design_agent() or qwen_agent()
    prompt = ("Ты независимый приёмщик результата автоматизации. Не доверяй полю status и самооценке автора. "
              "Сопоставь Task Contract, Source Model, фактические профили входов, summary, evidence и превью отчётов. "
              "PASS только если основная бизнес-цель выполнена на всех образцах и это доказано конкретными данными. "
              "Нулевой результат допустим только с наблюдаемым доказательством; если Task Contract связывает "
              "источники, неожиданное отсутствие любых совпадений требует диагностики Source Model и ключей. "
              "Если в ТЗ не хватает критического бизнес-выбора, FAIL и сформулируй один точный вопрос владельцу. "
              "Такой вопрос обязан назвать найденные варианты, наблюдаемое доказательство, предполагаемый "
              "вариант и то, на какой результат повлияет ответ. Техническую ошибку человеку не задавай. "
              "Извлеки только переносимые знания: concept — доказанный факт о структуре/семантике процесса; "
              "rule — доказанное обобщаемое правило обработки. Не записывай конкретные строки контрольного набора. "
              "В rejected_hypotheses перечисли ошибочные подходы, которые следующая попытка не должна повторять. "
              "Верни ТОЛЬКО JSON: "
              '{"verdict":"pass|fail","confidence":0.0,"issues":["..."],"owner_question":"",'
              '"memory":{"concepts":[{"text":"...","evidence":"...","confidence":0.0,"scope":"process"}],'
              '"rules":[{"text":"...","evidence":"...","confidence":0.0,"scope":"process"}]},'
              '"rejected_hypotheses":[{"text":"...","evidence":"...","rejection_reason":"..."}]}.\n\n'
              "AUTHORITATIVE BUILD BRIEF:\n" + json.dumps(_builder_brief(task_package), ensure_ascii=False, default=str) +
              "\n\nSTRUCTURAL CHECK:\n" + json.dumps(_compact(validation), ensure_ascii=False, default=str) +
              "\n\nACTUAL RESULT:\n" + json.dumps(_result_preview(result, validation), ensure_ascii=False, default=str))
    response = _post_agent(agent_id, {"agent_id": agent_id, "input": prompt, "run_timeout": 180,
                                      "store": False, "temperature": 0, "max_output_tokens": 2200}, timeout=210)
    verdict = _json_object(_agent_text(response))
    if not verdict:
        return {"verdict": "fail", "confidence": 0, "issues": ["приёмщик не вернул валидный JSON"],
                "owner_question": "", "memory": {"concepts": [], "rules": []},
                "rejected_hypotheses": []}
    if verdict.get("verdict") not in ("pass", "fail"):
        verdict["verdict"] = "fail"
    try:
        verdict["confidence"] = max(0.0, min(1.0, float(verdict.get("confidence", 0))))
    except Exception:
        verdict["confidence"] = 0.0
    verdict["issues"] = [str(x)[:300] for x in (verdict.get("issues") or [])[:8]]
    verdict["owner_question"] = str(verdict.get("owner_question") or "")[:500]
    memory = verdict.get("memory") if isinstance(verdict.get("memory"), dict) else {}
    verdict["memory"] = {
        "concepts": [x for x in (memory.get("concepts") or [])[:20] if isinstance(x, (dict, str))],
        "rules": [x for x in (memory.get("rules") or [])[:20] if isinstance(x, (dict, str))],
    }
    verdict["rejected_hypotheses"] = [x for x in (verdict.get("rejected_hypotheses") or [])[:20]
                                                if isinstance(x, (dict, str))]
    return verdict


def build_agentic_solution(session_id, build_id, namespace, sample_files, sess_dir, runs_dir, llm,
                            progress=None, max_attempts=None, max_creation_attempts=4,
                            max_run_repairs=2, max_acceptance_repairs=2, max_total_attempts=10,
                            max_elapsed_seconds=3600, prepared_context=None,
                            step_contract=None, expert_name_override=""):
    """Настоящий draft → run_expert → два гейта → bounded repair → stable.

    Бюджеты генерации, ремонта технического прогона и ремонта смысловой приёмки независимы. Поэтому
    ошибки первых генераций не отнимают у первого исполняемого варианта его repair budget.
    """
    progress = progress or (lambda *args, **kwargs: None)
    llm = llm if isinstance(llm, dict) else {}
    if max_attempts is not None:  # обратная совместимость старого caller-а: это только create budget
        max_creation_attempts = max_attempts
    create_budget = max(1, int(max_creation_attempts or 1))
    run_budget = max(2, int(max_run_repairs or 2))
    accept_budget = max(2, int(max_acceptance_repairs or 2))
    total_budget = max(create_budget + 2, int(max_total_attempts or 1))
    time_budget = max(300, int(max_elapsed_seconds or 3600))
    started_monotonic = time.monotonic()
    step_contract = step_contract if isinstance(step_contract, dict) else None
    step_acceptance = ((step_contract or {}).get("acceptance")
                       if isinstance((step_contract or {}).get("acceptance"), dict) else {}) or {}
    semantic_criteria = [str(x) for x in (step_acceptance.get("semantic_criteria") or [])
                         if str(x).strip()]
    try:
        semantic_threshold = max(0.0, min(1.0, float(step_acceptance.get("minimum_confidence", 0.7))))
    except Exception:
        semantic_threshold = 0.7
    if step_contract:
        step_dir = re.sub(r"[^A-Za-z0-9_-]+", "_", str(step_contract.get("id") or "step"))[:80]
        step_version = max(1, int(step_contract.get("version") or 1))
        bdir = Path(runs_dir) / build_id / "process" / "steps" / step_dir / ("v%d" % step_version)
    else:
        bdir = Path(runs_dir) / build_id
    bdir.mkdir(parents=True, exist_ok=True)
    progress("agentic_context", "Собираю Task Contract и изучаю все файлы", "running",
             ("шаг: " + str(step_contract.get("title") or step_contract.get("id"))) if step_contract else "")
    context = prepared_context if isinstance(prepared_context, dict) else prepare_task_context(
        session_id, sample_files, sess_dir, llm, step_contract=step_contract)
    package = copy.deepcopy(context.get("package") or {})
    if step_contract:
        package = _apply_upc_step(package, step_contract)
    progress("agentic_context", "Task Contract собран", "success",
             "образцов: %d · контракт %s" %
             (len(sample_files), (package.get("task_contract") or {}).get("sha256", "")[:10]))
    progress("agentic_source", "Проверяю Source Model фактических входов", "running",
             "файлы, разделы, сущности, поля и операции сопоставлены с Task Contract")
    if not context.get("ok"):
        source_model = context.get("source_model") or {"status": "error",
                                                        "reason": context.get("why") or "не построена"}
        _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                               "source_model_failed", build_id)
        progress("agentic_source", "Source Model не прошла проверку", "error", source_model["reason"])
        return {"ok": False, "code": "source_model_failed", "failure_kind": "builder_defect",
                "detail": source_model["reason"], "owner_question": "", "attempts": [],
                "task_contract": package.get("task_contract"), "source_model": source_model,
                "working_memory": package.get("working_memory"), "verified_memory": [], "draft_created": False,
                "expert_ran": False, "package_sha256": package.get("package_sha256")}
    source_model = context["source_model"]
    source_ids = context.get("source_memory_ids") or []
    (bdir / "task_package.json").write_text(
        json.dumps(package, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                           "source_model_" + source_model.get("status", "unknown"), build_id)
    progress("agentic_source", "Source Model построена", "success",
             "стратегия: %s · источников: %d" %
             (source_model.get("strategy"), len(source_model.get("sources") or [])))
    progress("agentic_memory", "Рабочая память подготовлена", "success",
             "концептов/правил: %d · rejected: %d" % (
                 len((package.get("working_memory") or {}).get("entries") or []),
                 sum(1 for x in (package.get("working_memory") or {}).get("entries") or []
                     if x.get("status") == "rejected")))
    execution_route = str(source_model.get("execution_route") or "build")
    execution_decision = dict(source_model.get("decision_record") or {})
    execution_decision.update({
        "route": execution_route,
        "reason": str(source_model.get("reason") or ""),
        "selected_capabilities": list(source_model.get("selected_capabilities") or []),
        "decided_before_expert_name": True,
        "source": "task_first_source_model",
    })
    package["execution_decision"] = execution_decision
    _refresh_package(package)
    (bdir / "task_package.json").write_text(
        json.dumps(package, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    if source_model.get("status") in ("need_human", "acquire"):
        code = "needs_owner_input" if source_model["status"] == "need_human" else "capability_missing"
        progress("agentic_source",
                 "Нужен ответ владельца" if source_model["status"] == "need_human" else
                 "Нужна внешняя capability — установка не выполнялась",
                 "warn", source_model.get("question") or source_model.get("reason") or "")
        return {"ok": False, "code": code, "failure_kind": source_model["status"],
                "detail": source_model.get("reason") or source_model.get("missing_capability") or "",
                "owner_question": source_model.get("question") or "", "attempts": [],
                "task_contract": package.get("task_contract"), "source_model": source_model,
                "execution_decision": execution_decision,
                "working_memory": package.get("working_memory"), "verified_memory": [], "draft_created": False,
                "expert_ran": False, "package_sha256": package.get("package_sha256")}
    _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                           "execution_route_" + execution_route, build_id)
    route_titles = {
        "reuse": "Использовать проверенную способность",
        "build": "Создать детерминированного эксперта",
        "llm_worker": "Создать эксперта с Qwen внутри",
        "split_step": "Разделить задачу на проверяемые шаги",
        "replan": "Перестроить границы процесса",
    }
    progress("agentic_route", "Qwen выбрала способ выполнения: " +
             route_titles.get(execution_route, execution_route), "success",
             str(execution_decision.get("reason") or
                 "; ".join(execution_decision.get("basis") or []))[:600])
    if execution_route in ("split_step", "replan"):
        action = "split_step" if execution_route == "split_step" else "replan"
        detail = str(execution_decision.get("reason") or execution_decision.get("next_action") or
                     "граница текущего шага требует изменения до создания эксперта")
        return {"ok": False, "code": "execution_" + action + "_required",
                "failure_kind": "decomposition", "detail": detail, "owner_question": "",
                "attempts": [], "control_action": action,
                "failure_decision": {"failure_class": "task_boundary", "action": action,
                                     "evidence": execution_decision.get("basis") or [detail],
                                     "owner_question": ""},
                "task_contract": package.get("task_contract"), "source_model": source_model,
                "execution_decision": execution_decision,
                "working_memory": package.get("working_memory"), "verified_memory": [],
                "draft_created": False, "expert_ran": False,
                "package_sha256": package.get("package_sha256")}

    expert_name = str(expert_name_override or (namespace + "_run_process"))
    draft_name = expert_name + "__draft_" + hashlib.sha256(str(build_id).encode("utf-8")).hexdigest()[:8]
    source_file = str(Path(sample_files[0]).parent if len(sample_files) > 1 else Path(sample_files[0]))
    attempts, feedback = [], ""
    creation_calls = run_repairs = acceptance_repairs = total_calls = 0
    runnable_attempts = 0
    previous_code_sha = ""
    seen_code_shas = set()
    strategy_failures = set()
    controller_history = []
    last_controller = {}
    retry_existing = False
    last_built = {}
    has_runnable = draft_created = expert_ran = False
    failure_class = "creation"
    last_applied_lesson = ""
    last_unapplied_lesson = ""
    final_code = "agentic_build_failed"
    final_failure_kind = "builder_defect"

    def can_continue():
        if time.monotonic() - started_monotonic >= time_budget:
            return False
        if total_calls >= total_budget:
            return False
        if not has_runnable:
            return creation_calls < create_budget
        if failure_class == "run":
            return run_repairs < run_budget
        return acceptance_repairs < accept_budget

    while total_calls < total_budget:
        if not can_continue() and total_calls:
            break
        total_calls += 1
        if not has_runnable:
            creation_calls += 1
        remaining = total_budget - total_calls
        phase_name = ("Повторяю тот же draft после временного сбоя" if retry_existing else
                      ("Qwen создаёт code-artifact draft-эксперта" if not has_runnable else
                       "Qwen создаёт исправленный code-artifact"))
        progress("agentic_reason", "%s · попытка текущего шага %d/%d" %
                 (phase_name, total_calls, total_budget),
                 "running", "draft: %s · осталось попыток: %d · repair run/accept: %d/%d, %d/%d" %
                 (draft_name, remaining, run_repairs, run_budget, acceptance_repairs, accept_budget))
        package["working_memory"] = package.get("working_memory") or {"version": 1, "entries": []}
        _refresh_package(package)
        reused_for_transient = retry_existing
        if retry_existing:
            built = dict(last_built)
            built.update({"ok": True, "generation_path": "transient_retry_same_expert",
                          "generation_seconds": 0.0})
            retry_existing = False
        else:
            built = _create_or_update(draft_name, package, feedback, llm)
        if not built.get("ok"):
            issues = [built.get("why") or "эксперт не создан"]
            rec = {"attempt": total_calls, "phase": "create" if not has_runnable else "repair_generate",
                   "failure_class": "creation", "issues": issues, "code_changed": False,
                   "budgets": {"creation": creation_calls, "run_repair": run_repairs,
                               "acceptance_repair": acceptance_repairs, "total": total_calls}}
            attempts.append(rec)
            lesson = _failure_lesson(total_calls, issues, previous_code_sha, source="builder")
            package["working_memory"] = _merge_memory(package.get("working_memory"), [lesson])
            last_unapplied_lesson = lesson.get("text") if lesson else issues[0]
            feedback = {**rec, "last_unapplied_lesson": last_unapplied_lesson,
                        "previous_code_sha256": previous_code_sha}
            _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                                   "repairing" if can_continue() else "failed", build_id)
            if can_continue():
                progress("agentic_reason", "Draft не создан · следующая попытка учтёт урок", "warn",
                         _repair_context(feedback, package))
                time.sleep(min(total_calls, 2))
                continue
            progress("agentic_reason", "Draft не создан · допустимые попытки исчерпаны", "error",
                     _repair_context(feedback, package))
            break

        draft_created = True
        code_sha = str(built.get("code_sha256") or "")
        strategy_sha = str(built.get("strategy_sha256") or "")
        code_changed = not previous_code_sha or code_sha != previous_code_sha
        duplicate_code = bool(has_runnable and not reused_for_transient and code_sha in seen_code_shas)
        output_contract_only = last_controller.get("action") == "repair_output_contract"
        duplicate_strategy = bool(has_runnable and not reused_for_transient and not output_contract_only and
                                  strategy_sha and
                                  (strategy_sha, str(last_controller.get("failure_class") or failure_class))
                                  in strategy_failures)
        if duplicate_code or duplicate_strategy:
            issues = ["Qwen повторила уже испытанный code hash" if duplicate_code else
                      "Qwen повторила ту же стратегию после того же класса сбоя"]
            controller = adaptive_failure_decision({
                "issues": issues, "code_seen": duplicate_code, "strategy_seen": duplicate_strategy,
                "input_count": len(sample_files),
                "input_formats": sorted({Path(x).suffix.casefold() for x in sample_files}),
                "can_split": bool(step_contract and step_can_partition(step_contract)),
            })
            controller_history.append(controller)
            last_controller = controller
            rec = {"attempt": total_calls, "phase": "no_progress",
                   "failure_class": controller["failure_class"], "controller": controller,
                   "issues": issues, "code_sha256": code_sha, "code_changed": False,
                   "budgets": {"creation": creation_calls, "run_repair": run_repairs,
                               "acceptance_repair": acceptance_repairs, "total": total_calls}}
            attempts.append(rec)
            lesson = _failure_lesson(total_calls, issues, code_sha, no_progress=True)
            package["working_memory"] = _merge_memory(package.get("working_memory"), [lesson])
            last_unapplied_lesson = lesson.get("text") if lesson else issues[0]
            feedback = {**rec, "last_unapplied_lesson": last_unapplied_lesson,
                        "previous_code_sha256": previous_code_sha}
            _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                                   "repairing" if can_continue() else "stalled", build_id)
            progress("agentic_reason",
                     "Код не изменился · это не засчитано как ремонт",
                     "warn" if can_continue() else "error",
                     "rejected-урок передан следующей попытке" if can_continue() else
                     "общий лимит исчерпан; ложного продолжения нет")
            final_code, final_failure_kind = "agentic_stalled", "no_progress"
            if controller.get("action") == "split_step":
                progress("agentic_control", "Контроллер меняет гранулярность шага", "warn",
                         "no_progress → split_step; повторный code/strategy запуск запрещён")
                break
            if can_continue():
                feedback["required_next_action"] = controller.get("action")
                continue
            break

        if has_runnable and not reused_for_transient:
            if failure_class == "run":
                run_repairs += 1
            else:
                acceptance_repairs += 1
            last_applied_lesson = last_unapplied_lesson
            last_unapplied_lesson = ""
        has_runnable = True
        runnable_attempts += 1
        previous_code_sha = code_sha
        seen_code_shas.add(code_sha)
        last_built = dict(built)
        progress("agentic_reason", "Draft-эксперт сохранён и независимо прочитан", "success",
                 "%s · код %s · путь: %s · %.1f сек · изменился: %s" %
                 (draft_name, code_sha[:10], built.get("generation_path") or "unknown",
                  float(built.get("generation_seconds") or 0), "да" if code_changed else "первый"))
        outdir = bdir / ("solution_attempt_%d" % total_calls)
        outdir.mkdir(parents=True, exist_ok=True)
        progress("agentic_run", "Запускаю настоящий draft через run_expert", "running",
                 "попытка %d · файлы: %s" %
                 (runnable_attempts, ", ".join(Path(p).name for p in sample_files)))
        params = {"source_file": source_file, "output_dir": str(outdir),
                  "api_token": CONFIG.get("auth_token", ""), "api_base": BASE,
                  "rules_json": json.dumps(package.get("rules") or [], ensure_ascii=False),
                  "fields_json": json.dumps(package.get("fields") or {}, ensure_ascii=False)}
        input_formats = sorted({Path(x).suffix.casefold() or "<none>" for x in sample_files})
        run_timeout = 240 if len(sample_files) >= 6 or len(input_formats) > 1 else 600
        run_started = time.monotonic()
        result = run_expert(draft_name, params, wait=run_timeout, glob=True)
        run_seconds = round(time.monotonic() - run_started, 1)
        expert_ran = True
        result = materialize_step_result(result, outdir)
        validation = validate_result(result, sample_files, package)
        if validation["ok"]:
            progress("agentic_run", "Draft обработал обязательные входы и создал артефакты", "success",
                     "файлов: %d · артефактов: %d" %
                     (len(sample_files), len(validation.get("artifacts") or [])))
            if step_contract and not semantic_criteria:
                # Для файлов/хэшей/schema/counts второй вызов Qwen ничего не доказывает сверх
                # детерминированного гейта. Семантические шаги по-прежнему проверяет независимая Qwen.
                judge = {"verdict": "pass", "confidence": 1.0, "issues": [], "owner_question": "",
                         "memory": {"concepts": [], "rules": []}, "rejected_hypotheses": [],
                         "source": "deterministic_contract"}
                progress("agentic_accept", "Детерминированная приёмка шага пройдена", "success",
                         "дополнительный вызов Qwen не требуется")
            else:
                progress("agentic_accept", "Независимо проверяю бизнес-результат по Task Contract", "running",
                         "порог: %d%%" % round(semantic_threshold * 100))
                judge = judge_result(package, result, validation)
            failure_class = "acceptance"
        else:
            judge = {"verdict": "fail", "confidence": 1.0, "issues": validation["issues"],
                     "owner_question": "", "memory": {"concepts": [], "rules": []},
                     "rejected_hypotheses": []}
            failure_class = "run"
        rec = {"attempt": total_calls, "runnable_attempt": runnable_attempts, "phase": "accept",
               "failure_class": failure_class, "code_sha256": code_sha, "code_changed": code_changed,
               "strategy_sha256": strategy_sha, "run_seconds": run_seconds,
               "validation": validation, "judge": judge, "result": _compact(result),
               "budgets": {"creation": creation_calls, "run_repair": run_repairs,
                           "acceptance_repair": acceptance_repairs, "total": total_calls}}
        attempts.append(rec)
        passed = (validation["ok"] and judge.get("verdict") == "pass" and
                  judge.get("confidence", 0) >= semantic_threshold)
        if passed:
            progress("agentic_publish", "Публикую stable-эксперта после полной приёмки", "running", expert_name)
            promoted = _promote_expert(draft_name, expert_name, package)
            if not promoted.get("ok"):
                final_code, final_failure_kind = "agent_publish_failed", "platform_error"
                last_unapplied_lesson = promoted.get("why") or "ошибка публикации stable-эксперта"
                progress("agentic_publish", "Stable-эксперт не опубликован", "error", last_unapplied_lesson)
                break
            package["working_memory"] = _merge_memory(
                package.get("working_memory"), _memory_from_judge(judge, total_calls, passed=True))
            package["working_memory"] = _promote_memory(package.get("working_memory"), source_ids, total_calls)
            _refresh_package(package)
            verified_memory = _verified_memory(package.get("working_memory"))
            _persist_agentic_state(sess_dir, session_id, bdir, package, source_model, "success", build_id)
            _delete_draft(draft_name, llm.get("agent_id") or qwen_agent())
            progress("agentic_memory", "Подтверждённая память готова к публикации вместе с агентом", "success",
                     "verified concepts/rules: %d" % len(verified_memory))
            progress("agentic_publish", "Stable-эксперт опубликован", "success", expert_name)
            progress("agentic_accept", "Бизнес-результат подтверждён", "success",
                     "уверенность: %d%%" % round(judge.get("confidence", 0) * 100))
            evidence = {"contract_version": 2, "built_at": datetime.now(timezone.utc).isoformat(),
                        "package_sha256": package.get("package_sha256"), "expert": expert_name,
                        "draft_expert": draft_name, "source_model": source_model,
                        "execution_decision": execution_decision,
                        "task_contract": package.get("task_contract"), "verified_memory": verified_memory,
                        "source_files": [{"name": Path(p).name, "sha256": _sha256(p)} for p in sample_files],
                        "attempts": attempts, "accepted_attempt": total_calls}
            (bdir / "agentic_evidence.json").write_text(
                json.dumps(evidence, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
            return {"ok": True, "expert": expert_name, "draft_expert": draft_name,
                    "source_file": source_file, "summary": validation.get("summary") or {},
                    "result": result, "validation": validation, "judge": judge, "attempts": attempts,
                    "task_contract": package.get("task_contract"), "source_model": source_model,
                    "execution_decision": execution_decision,
                    "strategy": source_model.get("strategy"), "working_memory": package.get("working_memory"),
                    "verified_memory": verified_memory, "draft_created": True, "expert_ran": True,
                    "last_applied_lesson": last_applied_lesson, "last_unapplied_lesson": "",
                    "package_sha256": package.get("package_sha256"),
                    "expert_code_sha256": promoted.get("code_sha256") or code_sha,
                    "failure_decision": last_controller,
                    "controller_history": controller_history,
                    "source_files": [Path(p).name for p in sample_files]}

        issues = judge.get("issues") or validation.get("issues") or ["бизнес-цель не доказана"]
        owner_question = str(judge.get("owner_question") or "").strip()
        controller = adaptive_failure_decision({
            "issues": issues, "message": result, "status": result.get("status") if isinstance(result, dict) else "",
            "duration_seconds": run_seconds, "run_timeout_seconds": run_timeout,
            "input_count": len(sample_files), "input_formats": input_formats,
            "previous_classes": [x.get("failure_class") for x in controller_history if isinstance(x, dict)],
            "owner_question": owner_question,
            "semantic_ambiguity": bool(validation.get("ok") and owner_question and
                                       owner_question_gate(package, owner_question).get("ask")),
            "semantic_failed": bool(validation.get("ok") and judge.get("verdict") == "fail"),
            "can_split": bool(step_contract and step_can_partition(step_contract)),
        })
        question_gate = owner_question_gate(package, owner_question)
        if owner_question and not question_gate.get("ask"):
            controller.setdefault("evidence", []).append(
                "вопрос человеку подавлен как " + str(question_gate.get("classification") or "technical"))
            rec["suppressed_owner_question"] = {
                "question": owner_question, "classification": question_gate.get("classification"),
                "reason": question_gate.get("reason")}
        controller_history.append(controller)
        last_controller = controller
        rec["controller"] = controller
        rec["failure_class"] = controller.get("failure_class")
        if strategy_sha:
            strategy_failures.add((strategy_sha, str(controller.get("failure_class") or "")))
        learned = _memory_from_judge(judge, total_calls, passed=False)
        lesson = _failure_lesson(total_calls, issues, code_sha, source="acceptance_gate")
        package["working_memory"] = _merge_memory(package.get("working_memory"), learned + [lesson])
        last_unapplied_lesson = lesson.get("text") if lesson else "; ".join(issues)
        feedback = _repair_feedback(result, validation, judge, controller)
        feedback.update({"previous_code_sha256": code_sha, "last_unapplied_lesson": last_unapplied_lesson,
                         "working_memory": package.get("working_memory"),
                         "remaining": {"run_repairs": max(0, run_budget - run_repairs),
                                       "acceptance_repairs": max(0, accept_budget - acceptance_repairs),
                                       "total": max(0, total_budget - total_calls)}})
        if controller.get("action") == "repair_output_contract":
            feedback["previous_expert_code"] = str(last_built.get("code") or "")
        _refresh_package(package)
        if controller.get("action") == "ask_owner" and controller.get("owner_question"):
            final_code, final_failure_kind = "needs_owner_input", "need_human"
            _persist_agentic_state(sess_dir, session_id, bdir, package, source_model, "need_human", build_id)
            progress("agentic_accept", "Нужен бизнес-выбор владельца", "warn",
                     controller.get("owner_question"))
            break
        if controller.get("action") == "split_step":
            final_code, final_failure_kind = "agentic_needs_decomposition", controller.get("failure_class")
            _persist_agentic_state(sess_dir, session_id, bdir, package, source_model, "decompose", build_id)
            progress("agentic_control", "Контроллер дробит перегруженный шаг", "warn",
                     "%s → split_step; принятые checkpoints других шагов сохраняются" %
                     controller.get("failure_class"))
            break
        if controller.get("action") == "retry_transient":
            retry_existing = True
            progress("agentic_control", "Временный сбой: повторяю тот же эксперт один раз", "warn",
                     "без повторной генерации кода")
        elif controller.get("action") == "repair_output_contract":
            progress("agentic_control", "Ремонтирую только выходной контракт", "warn",
                     "; ".join(str(x) for x in issues)[:420])
        elif controller.get("action") == "replace_implementation":
            progress("agentic_control", "Меняю стратегию реализации", "warn",
                     "повтор прежнего code/strategy hash запрещён")
        _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                               "repairing" if can_continue() else "failed", build_id)
        if can_continue():
            progress("agentic_accept", "Проверка отклонила результат · урок передан следующему ремонту", "warn",
                     "%s · код должен измениться · осталось run/accept: %d/%d" %
                     ("; ".join(str(x) for x in issues)[:420],
                      max(0, run_budget - run_repairs), max(0, accept_budget - acceptance_repairs)))
            time.sleep(min(total_calls, 2))
            continue
        final_code = "agentic_acceptance_failed" if validation.get("ok") else "agentic_run_failed"
        final_failure_kind = "acceptance" if validation.get("ok") else "runtime"
        progress("agentic_accept", "Результат отклонён · доступный repair budget исчерпан", "error",
                 "; ".join(str(x) for x in issues)[:500])
        break

    last = attempts[-1] if attempts else {}
    judge = last.get("judge") or {}
    issues = judge.get("issues") or last.get("issues") or ["решение не прошло приёмку"]
    if final_code == "agentic_build_failed" and has_runnable:
        final_code, final_failure_kind = "agentic_acceptance_failed", "acceptance"
    if time.monotonic() - started_monotonic >= time_budget and final_code in (
            "agentic_build_failed", "agentic_acceptance_failed", "agentic_run_failed"):
        final_code, final_failure_kind = "agentic_timeout", "timeout"
        last_unapplied_lesson = last_unapplied_lesson or "общий лимит времени сборки исчерпан"
    _persist_agentic_state(sess_dir, session_id, bdir, package, source_model,
                           "need_human" if final_code == "needs_owner_input" else "failed", build_id)
    _delete_draft(draft_name, llm.get("agent_id") or qwen_agent())
    control_action = str(last_controller.get("action") or "stop_fail_closed")
    safe_owner_question = (last_controller.get("owner_question") or "") \
        if control_action == "ask_owner" else ""
    return {"ok": False, "code": final_code, "failure_kind": final_failure_kind,
            "expert": expert_name, "draft_expert": draft_name, "detail": "; ".join(str(x) for x in issues)[:800],
            "owner_question": safe_owner_question, "attempts": attempts,
            "control_action": control_action, "failure_decision": last_controller,
            "controller_history": controller_history,
            "task_contract": package.get("task_contract"), "source_model": source_model,
            "execution_decision": execution_decision,
            "strategy": source_model.get("strategy"), "working_memory": package.get("working_memory"),
            "verified_memory": [], "draft_created": draft_created, "expert_ran": expert_ran,
            "files_processed": (last.get("validation") or {}).get("files_used") or [],
            "last_applied_lesson": last_applied_lesson,
            "last_unapplied_lesson": last_unapplied_lesson,
            "package_sha256": package.get("package_sha256"),
            "expert_code_sha256": previous_code_sha,
            "budgets": {"creation": {"used": creation_calls, "limit": create_budget},
                        "run_repair": {"used": run_repairs, "limit": run_budget},
                        "acceptance_repair": {"used": acceptance_repairs, "limit": accept_budget},
                        "total": {"used": total_calls, "limit": total_budget},
                        "time_seconds": {"used": round(time.monotonic() - started_monotonic, 1),
                                         "limit": time_budget}}}
